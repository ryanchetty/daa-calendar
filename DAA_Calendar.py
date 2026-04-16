import os, json, sqlite3, openpyxl, re, sys, threading, time, calendar, datetime, psutil, fitz, math, tempfile, urllib.request, subprocess, ssl, certifi
from datetime import datetime, timedelta, date
from PyQt5 import QtCore, QtGui, QtWidgets, sip
from PyQt5.QtCore import Qt, QTimer, QEvent, QObject, QMarginsF, QRect, pyqtSignal, pyqtSlot, QMetaObject, QDate, QPointF, QSize, QRectF, QStandardPaths, QSizeF
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog, QPrintPreviewDialog, QPrinterInfo
from PyQt5.QtGui import QColor, QBrush, QFont, QIcon, QPainter, QPageLayout, QPageSize, QPixmap, QTextDocument, QImage, qAlpha, QPen, QTransform
from PyQt5.QtWidgets import (QToolButton, QFrame,
    QAction, QApplication, QCheckBox, QComboBox, QDialog, QGridLayout, QGroupBox,
    QHBoxLayout, QHeaderView, QInputDialog, QLabel, QLineEdit, QMainWindow, QMenuBar,
    QMessageBox, QPushButton, QScrollArea, QSizePolicy, QSpacerItem, QSplitter,
    QTabWidget, QTableWidget, QTableWidgetItem, QTextEdit, QVBoxLayout, QWidget,
    QTabBar, QStackedWidget, QFileDialog, QStyleFactory, QTableView,
    QStyledItemDelegate, QRadioButton, QAbstractScrollArea, QLayout,
    QDateEdit, QListWidgetItem, QGraphicsView, QGraphicsScene, QListWidget,
    QButtonGroup, QFormLayout, QSizePolicy, QDoubleSpinBox, QSplashScreen
)
from PyQt5.QtNetwork import QLocalServer, QLocalSocket
from PyQt5.QtSvg import QSvgRenderer
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import Workbook, load_workbook


APP_VERSION = "1.2"
GITHUB_OWNER = "ryanchetty"
GITHUB_REPO = "daa-calendar"
INSTALLER_ASSET_NAME = "DAACal_Installer.exe"
APP_EXE_NAME = "DAA_Calendar.exe"

def parse_version(v: str):
    v = (v or "").strip().lstrip("vV")
    parts = []
    for p in v.split("."):
        try:
            parts.append(int(p))
        except ValueError:
            parts.append(0)
    return tuple(parts)

def is_newer_version(latest: str, current: str) -> bool:
    return parse_version(latest) > parse_version(current)

def get_latest_github_release():
    url = f"https://api.github.com/repos/ryanchetty/daa-calendar/releases/latest"
    req = urllib.request.Request(url, headers={"User-Agent": "DAACal-Updater"})

    ssl_context = ssl.create_default_context(cafile=certifi.where())

    with urllib.request.urlopen(req, timeout=15, context=ssl_context) as resp:
        return json.load(resp)
def find_installer_asset(release: dict):
    for asset in release.get("assets", []):
        name = asset.get("name", "")
        if name.lower() == INSTALLER_ASSET_NAME.lower():
            return asset

    for asset in release.get("assets", []):
        name = asset.get("name", "")
        if name.lower().endswith(".exe"):
            return asset

    return None

def download_update_installer(download_url: str, asset_name: str) -> str:
    out_path = os.path.join(tempfile.gettempdir(), asset_name)
    ssl_context = ssl.create_default_context(cafile=certifi.where())

    with urllib.request.urlopen(download_url, context=ssl_context) as resp, open(out_path, "wb") as f:
        f.write(resp.read())

    return out_path
def create_updater_bat(installer_path: str, app_exe_path: str) -> str:
    bat_path = os.path.join(tempfile.gettempdir(), "DAACal_RunUpdate.bat")

    script = f"""@echo off
setlocal

set "INSTALLER={installer_path}"
set "APP_EXE={app_exe_path}"

timeout /t 2 /nobreak >nul

"%INSTALLER%" /VERYSILENT /SUPPRESSMSGBOXES /NORESTART

timeout /t 2 /nobreak >nul

start "" "%APP_EXE%"

endlocal
"""
    with open(bat_path, "w", encoding="utf-8") as f:
        f.write(script)

    return bat_path

def prompt_for_update(parent, latest_version: str) -> bool:
    from PyQt5.QtWidgets import QMessageBox

    msg = QMessageBox(parent)
    msg.setWindowTitle("DAACal Update Available")
    msg.setIcon(QMessageBox.Information)
    msg.setText(
        f"A new version of DAACal is available.\n\n"
        f"Current version: {APP_VERSION}\n"
        f"New version: {latest_version.lstrip('vV')}\n\n"
        f"Update now?"
    )
    msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    return msg.exec_() == QMessageBox.Yes

def start_silent_update(parent=None):
    try:
        release = get_latest_github_release()
        latest_tag = (release.get("tag_name") or "").strip()
        if not latest_tag:
            return False

        if not is_newer_version(latest_tag, APP_VERSION):
            return False

        if not prompt_for_update(parent, latest_tag):
            return False

        asset = find_installer_asset(release)
        if not asset:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.warning(parent, "Update Error", "No installer asset was found in the latest GitHub release.")
            return False

        installer_url = asset.get("browser_download_url", "").strip()
        asset_name = asset.get("name", INSTALLER_ASSET_NAME).strip() or INSTALLER_ASSET_NAME
        if not installer_url:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.warning(parent, "Update Error", "The installer download URL is missing.")
            return False

        installer_path = download_update_installer(installer_url, asset_name)

        if getattr(sys, "frozen", False):
            app_exe_path = sys.executable
        else:
            app_exe_path = os.path.abspath(sys.argv[0])

        updater_bat = create_updater_bat(installer_path, app_exe_path)

        subprocess.Popen(
            ["cmd.exe", "/c", updater_bat],
            creationflags=subprocess.CREATE_NO_WINDOW
        )

        # ask current app instance to quit cleanly
        try:
            send_ipc_command("QUIT")
        except Exception:
            pass

        QApplication.quit()
        return True

    except Exception as e:
        print("Updater failed:", e)
        return False

def resource_path(relative_name: str) -> str:
    """
    Returns an absolute path to a bundled resource.

    Works for:
    - running as .py (uses folder beside the .py)
    - PyInstaller .exe (uses the temp extraction folder in one-file mode)
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        base_dir = sys._MEIPASS  # PyInstaller temp folder
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, relative_name)

class LoadingSpinner(QWidget):
    def __init__(self, parent=None, size=36):
        super().__init__(parent)
        self._angle = 0
        self.setFixedSize(size, size)

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.rotate)
        self.timer.start(80)

    def rotate(self):
        self._angle = (self._angle + 30) % 360
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        cx = self.width() / 2.0
        cy = self.height() / 2.0
        outer_r = min(self.width(), self.height()) / 2.0 - 3
        inner_r = outer_r - 8

        painter.translate(cx, cy)
        painter.rotate(self._angle)

        # 12 dots, one bright "head", trailing fade
        dot_count = 12
        dot_size = 4.0

        for i in range(dot_count):
            painter.save()
            painter.rotate(i * 30.0)

            alpha = int(255 * ((i + 1) / dot_count))
            painter.setPen(Qt.NoPen)
            painter.setBrush(QColor(0, 0, 0, alpha))

            y = -inner_r
            painter.drawEllipse(QRectF(-dot_size / 2.0, y - dot_size / 2.0, dot_size, dot_size))
            painter.restore()

class StartupSplash(QWidget):
    def __init__(self, pixmap):
        super().__init__(None, Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.SplashScreen)
        self.setAttribute(Qt.WA_TranslucentBackground, False)
        self.setStyleSheet("background: white;")

        self.image_label = QLabel()
        self.image_label.setPixmap(pixmap)
        self.image_label.setAlignment(Qt.AlignCenter)

        self.spinner = LoadingSpinner(size=36)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)
        layout.addWidget(self.image_label, 0, Qt.AlignCenter)
        layout.addWidget(self.spinner, 0, Qt.AlignHCenter | Qt.AlignTop)

        self.adjustSize()

        screen = QApplication.primaryScreen().availableGeometry()
        self.move(
            screen.center().x() - self.width() // 2,
            screen.center().y() - self.height() // 2
        )

class MPSBlisterWizardDialog(QDialog):
    CAL_ZERO_PAD_X = 1.30
    CAL_ZERO_PAD_Y = 1.90
    CAL_ZERO_NUDGE_X = 0.00
    CAL_ZERO_NUDGE_Y = 2.00

    def _is_header_token(self, t: str) -> bool:
        """
        Broad detector for candidate header-like text.
        We intentionally allow both:
          - meal names
          - times
          - common meal synonyms
        Non-meal headers like MST1 are NOT considered header tokens here, because
        we do not want them rewritten as meal slots.
        """
        t = (t or "").strip()
        if not t:
            return False

        u = t.upper().replace("’", "'").replace("‘", "'")
        u = " ".join(u.split())

        if u in {
            "B'FAST", "BFAST", "BREAKFAST", "MORNING",
            "LUNCH", "NOON", "MIDDAY",
            "DINNER", "EVENING", "SUPPER",
            "BED", "BEDTIME", "NIGHT"
        }:
            return True

        if "FAST" in u:
            return True

        # time tokens: 8:10, 08:10, 8AM, 8 AM, 10PM, 10 PM
        if re.fullmatch(r"\d{1,2}:\d{2}", u):
            return True
        if re.fullmatch(r"\d{1,2}\s?(AM|PM)", u):
            return True

        return False

    def _norm_table_header_text(self, t: str) -> str:
        t = (t or "").upper().strip()
        t = t.replace("’", "'").replace("‘", "'")
        t = " ".join(t.split())
        return t

    def _slot_for_named_table_header(self, t: str):
        """
        Semantic mapping for meal-name style table headers.
        Returns slot index 0..3 or None.
        """
        n = self._norm_table_header_text(t)

        if n in {"B'FAST", "BFAST", "BREAKFAST", "MORNING"}:
            return 0
        if n in {"LUNCH", "NOON", "MIDDAY"}:
            return 1
        if n in {"DINNER", "EVENING", "SUPPER"}:
            return 2
        if n in {"BED", "BEDTIME", "NIGHT"}:
            return 3

        return None

    def _is_time_header_text(self, t: str) -> bool:
        n = self._norm_table_header_text(t)
        return (
                re.fullmatch(r"\d{1,2}:\d{2}", n) is not None or
                re.fullmatch(r"\d{1,2}\s?(AM|PM)", n) is not None
        )

    def _extract_candidate_table_header_spans(self, page, top_slots: dict, bottom_slots: dict):
        """
        Return candidate spans from the middle table area only.

        Output items are:
            (cy, cx, rect, text, size)
        """
        spans = []
        raw = page.get_text("rawdict") or {}

        for block in raw.get("blocks", []):
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    txt = (span_text(span) or "").strip()
                    if not txt or not self._is_header_token(txt):
                        continue

                    x0, y0, x1, y1 = span.get("bbox", [0, 0, 0, 0])
                    rect = fitz.Rect(x0, y0, x1, y1)
                    cx = (x0 + x1) / 2.0
                    cy = (y0 + y1) / 2.0
                    size = float(span.get("size") or 9.0)

                    spans.append((cy, cx, rect, txt, size))

        top_y1 = max((r.y1 for r in top_slots.values()), default=0)
        bot_y0 = min((r.y0 for r in bottom_slots.values()), default=float(page.rect.height))

        mid_spans = [s for s in spans if top_y1 + 5 < s[0] < bot_y0 - 5]
        mid_spans.sort(key=lambda s: s[0])

        return mid_spans

    def _choose_best_table_header_band(self, mid_spans):
        """
        Pick the densest horizontal-ish row of candidate header spans.
        """
        if not mid_spans:
            return []

        best = []
        cur = []
        tol_y = 10.0

        for s in mid_spans:
            if not cur:
                cur = [s]
            elif abs(s[0] - cur[0][0]) <= tol_y:
                cur.append(s)
            else:
                if len(cur) > len(best):
                    best = cur
                cur = [s]

        if len(cur) > len(best):
            best = cur

        best_lr = sorted(best, key=lambda s: s[1])

        picked = []
        min_dx = 12.0
        for s in best_lr:
            if not picked or abs(s[1] - picked[-1][1]) > min_dx:
                picked.append(s)

        return picked

    def _build_table_header_cells(self, picked, labels, blanks, dbg=False):
        """
        Hybrid named-header + time-header strategy.

        Handles:
          - named anchors: B'FAST/LUNCH/DINNER/BED
          - time anchors: 8:05 / 8AM / etc mapped left-to-right
          - missing internal slots synthesized by interpolation
          - missing edge slots synthesized by extrapolation
          - non-meal columns ignored because picked only includes meal/time-like spans

        Returns:
          table_cells: dict slot_index -> fitz.Rect
        """
        if not picked:
            return {}

        band_y0 = min(s[2].y0 for s in picked) - 0.5
        band_y1 = max(s[2].y1 for s in picked) + 0.5

        named = []
        timed = []
        other = []

        for s in picked:
            txt = s[3]
            si = self._slot_for_named_table_header(txt)
            if si is not None:
                named.append((si, s))
            elif self._is_time_header_text(txt):
                timed.append(s)
            else:
                other.append(s)

        if dbg:
            print(" TABLE named =", [(si, x[3]) for si, x in named])
            print(" TABLE timed =", [x[3] for x in timed])
            print(" TABLE other =", [x[3] for x in other])

        detected = {}  # slot -> {"cx", "src_rect", "detected_text"}

        # Priority 1: semantic mapping if any named meal headers exist
        if named:
            for si, s in named:
                cx = float(s[1])
                src_rect = s[2]
                detected_text = s[3]

                prev = detected.get(si)
                if prev is None or src_rect.width > prev["src_rect"].width:
                    detected[si] = {
                        "cx": cx,
                        "src_rect": src_rect,
                        "detected_text": detected_text,
                    }

        # Priority 2: otherwise, time headers map left-to-right by position
        elif timed:
            timed = sorted(timed, key=lambda s: s[1])

            # Use up to 4 visible time columns as slot anchors 0..3
            for idx, s in enumerate(timed[:4]):
                detected[idx] = {
                    "cx": float(s[1]),
                    "src_rect": s[2],
                    "detected_text": s[3],
                }

        else:
            return {}

        if dbg:
            print(" TABLE detected slots =", sorted(detected.keys()))
            print(" TABLE detected texts =", {k: v["detected_text"] for k, v in sorted(detected.items())})

        table_cells = {}  # slot -> rect

        # Existing detected cells
        for si, info in detected.items():
            r = info["src_rect"]
            table_cells[si] = fitz.Rect(r.x0 - 2.0, band_y0, r.x1 + 2.0, band_y1)

        widths = [info["src_rect"].width for info in detected.values()]
        est_w = (sum(widths) / len(widths)) if widths else 36.0
        est_w = max(est_w, 24.0)

        def make_box(cx, width):
            half = width / 2.0
            return fitz.Rect(cx - half, band_y0, cx + half, band_y1)

        detected_slots = sorted(detected.keys())

        # Synthesize missing internal slots by interpolation
        for si in range(4):
            if si in table_cells:
                continue

            left = max((k for k in detected_slots if k < si), default=None)
            right = min((k for k in detected_slots if k > si), default=None)

            if left is not None and right is not None and right != left:
                left_cx = detected[left]["cx"]
                right_cx = detected[right]["cx"]
                step = (right_cx - left_cx) / (right - left)
                cx = left_cx + step * (si - left)
                table_cells[si] = make_box(cx, est_w)

                if dbg:
                    print(f" TABLE synthesized slot {si} by interpolation: cx={cx:.2f}")

        # Synthesize edges by extrapolation
        detected_slots = sorted(detected.keys())
        if len(detected_slots) >= 2:
            left_a, left_b = detected_slots[0], detected_slots[1]
            right_a, right_b = detected_slots[-2], detected_slots[-1]

            left_step = (detected[left_b]["cx"] - detected[left_a]["cx"]) / (left_b - left_a)
            right_step = (detected[right_b]["cx"] - detected[right_a]["cx"]) / (right_b - right_a)

            for si in range(4):
                if si in table_cells:
                    continue

                if si < detected_slots[0]:
                    cx = detected[left_a]["cx"] - left_step * (left_a - si)
                    table_cells[si] = make_box(cx, est_w)
                    if dbg:
                        print(f" TABLE synthesized edge slot {si} from left anchors: cx={cx:.2f}")

                elif si > detected_slots[-1]:
                    cx = detected[right_b]["cx"] + right_step * (si - right_b)
                    table_cells[si] = make_box(cx, est_w)
                    if dbg:
                        print(f" TABLE synthesized edge slot {si} from right anchors: cx={cx:.2f}")

        # Safety: keep boxes ordered and non-overlapping
        ordered = sorted((si, box) for si, box in table_cells.items())
        repaired = {}

        for idx, (si, box) in enumerate(ordered):
            x0, y0, x1, y1 = box
            if idx > 0:
                prev_si, prev_box = ordered[idx - 1]
                if x0 < prev_box.x1 + 1.0:
                    mid = (prev_box.x1 + x0) / 2.0
                    prev_box = fitz.Rect(prev_box.x0, prev_box.y0, mid - 0.5, prev_box.y1)
                    box = fitz.Rect(mid + 0.5, box.y0, box.x1, box.y1)
                    repaired[prev_si] = prev_box
            repaired[si] = box

        table_cells = repaired

        # Only keep slots that are actually wanted:
        # - if blank=True, still keep it, because we may need to whiteout old text
        # - if label is empty and blank=False, skip
        final_cells = {}
        for si, box in table_cells.items():
            txt = (labels[si] or "").strip()
            if blanks[si] or txt:
                final_cells[si] = box

        if dbg:
            print(" TABLE final boxes =", {si: tuple(box) for si, box in sorted(final_cells.items())})

        return final_cells

    def _apply_header_changes_from_list(self, headers_4: list, blank_flags_4=None):
        """
        Apply header changes to:
          1) top pill row
          2) bottom pill row
          3) middle table header row

        Hybrid table strategy:
          - named meal headers map by meaning
          - time headers map by left-to-right order
          - missing slot cells can be synthesized when needed
          - only meal/time header row is rewritten
        """
        if self.work_doc is None:
            return

        dbg = bool(getattr(self, "_hdr_debug", False))
        labels, blanks = self._norm_headers_and_blanks(headers_4, blank_flags_4)

        if dbg:
            print("\n=== _apply_header_changes_from_list ===")
            print("labels =", labels)
            print("blanks =", blanks)

        for pi in range(self.work_doc.page_count):
            page = self.work_doc.load_page(pi)
            page_w = float(page.rect.width)

            top_slots, bottom_slots = self._find_pill_slots_on_page(page)

            if dbg:
                print(f"\n--- Page {pi + 1}/{self.work_doc.page_count} ---")
                print("top_slots keys   =", sorted(top_slots.keys()))
                print("bottom_slots keys=", sorted(bottom_slots.keys()))
                for si in range(4):
                    tr = top_slots.get(si)
                    br = bottom_slots.get(si)
                    print(
                        f" slot {si}: top={tuple(tr) if tr else None}  "
                        f"bottom={tuple(br) if br else None}  "
                        f"blank={blanks[si]} label={labels[si]!r}"
                    )

                if len(top_slots) < 4:
                    print(f"[WARN] Only detected {len(top_slots)}/4 TOP pill bars on page {pi + 1}")
                if len(bottom_slots) < 4:
                    print(f"[WARN] Only detected {len(bottom_slots)}/4 BOTTOM pill bars on page {pi + 1}")

            # --- 1) Update TOP pills ---
            for si, pill_rect in top_slots.items():
                inner = self._pill_inner_box(pill_rect)
                if dbg:
                    print(f" TOP  slot {si}: whiteout inner={tuple(inner)}")
                self._whiteout_box(page, inner)

                if blanks[si]:
                    if dbg:
                        print(f" TOP  slot {si}: SKIP insert (blank=True)")
                    continue

                txt = (labels[si] or "").strip()
                if txt:
                    if dbg:
                        print(f" TOP  slot {si}: insert {txt!r} into={tuple(inner)}")
                    self._insert_centered_text(page, inner, txt, font_size=12.5, slot_index=si)

            # --- 2) Update BOTTOM pills ---
            for si, pill_rect in bottom_slots.items():
                inner = self._pill_inner_box(pill_rect)
                if dbg:
                    print(f" BOT  slot {si}: whiteout inner={tuple(inner)}")
                self._whiteout_box(page, inner)

                if blanks[si]:
                    if dbg:
                        print(f" BOT  slot {si}: SKIP insert (blank=True)")
                    continue

                txt = (labels[si] or "").strip()
                if txt:
                    if dbg:
                        print(f" BOT  slot {si}: insert {txt!r} into={tuple(inner)}")
                    self._insert_centered_text(page, inner, txt, font_size=12.5, slot_index=si)

            # --- 3) Update TABLE header row ---
            mid_spans = self._extract_candidate_table_header_spans(page, top_slots, bottom_slots)

            if dbg:
                top_y1 = max((r.y1 for r in top_slots.values()), default=0)
                bot_y0 = min((r.y0 for r in bottom_slots.values()), default=float(page.rect.height))
                print(
                    f"table header spans mid_spans={len(mid_spans)} "
                    f"(top_y1={round(top_y1, 2)} bot_y0={round(bot_y0, 2)})"
                )

            if not mid_spans:
                if dbg:
                    print("[WARN] TABLE: no candidate mid-row header spans found")
                continue

            picked = self._choose_best_table_header_band(mid_spans)

            if dbg:
                print(" TABLE picked =", [(round(s[1], 2), s[3]) for s in picked])

            if not picked:
                if dbg:
                    print("[WARN] TABLE: no distinct header spans detected; skipping table header rewrite")
                continue

            table_cells = self._build_table_header_cells(picked, labels, blanks, dbg=dbg)

            if not table_cells:
                if dbg:
                    print("[WARN] TABLE: no usable table cells resolved")
                continue

            for si in range(4):
                cell_box = table_cells.get(si)
                if cell_box is None:
                    continue

                if dbg:
                    print(f" TABLE slot {si}: whiteout cell_box={tuple(cell_box)}")
                self._whiteout_box(page, cell_box)

                if blanks[si]:
                    if dbg:
                        print(f" TABLE slot {si}: BLANK -> whiteout only")
                    continue

                txt = (labels[si] or "").strip()
                if not txt:
                    if dbg:
                        print(f" TABLE slot {si}: SKIP insert (empty text)")
                    continue

                table_font_size = self._table_font_size_for_text(txt)

                if dbg:
                    print(f" TABLE slot {si}: font_size={table_font_size} text={txt!r}")
                    print(f" TABLE slot {si}: insert {txt!r} into={tuple(cell_box)}")

                self._insert_centered_table_text(
                    page,
                    cell_box,
                    txt,
                    slot_index=si,
                    font_size=table_font_size,
                )

    def _pdf_slot_fg_rgb01(self, slot_index: int):
        """
        PDF-only: keep UI colors unchanged, but ensure slot 3 (dark blue pill)
        draws visible text after whiteout-to-white by forcing black text.
        """
        _, _, fg_hex = self._WIZ_HEADER_SLOT_COLOURS[int(slot_index)]
        if int(slot_index) == 3:
            fg_hex = "#000000"
        return self._rgb01_from_hex(fg_hex)

    def _norm_headers_and_blanks(self, headers, blanks):
        """
        Always return exactly 4 labels and 4 blank flags.
        - If headers missing: fall back to detected self.header_labels (from PDF), then meal defaults.
        - Never allow empty label with blank=False (it would whiteout and insert nothing).
        """
        base = list(getattr(self, "header_labels", None) or [])
        if len(base) < 4:
            base = (base + ["B'FAST", "LUNCH", "DINNER", "BED"])[:4]

        labels = list(headers or [])
        labels = (labels + base)[:4]
        labels = [(t or "").strip() for t in labels]

        blank_flags = list(blanks or [])
        blank_flags = (blank_flags + [False, False, False, False])[:4]
        blank_flags = [bool(x) for x in blank_flags]

        for i in range(4):
            if not blank_flags[i] and not labels[i]:
                labels[i] = base[i]

        return labels, blank_flags
    def _rgb01_from_hex(self, hx: str):
        hx = hx.lstrip("#")
        return tuple(int(hx[i:i + 2], 16) / 255.0 for i in (0, 2, 4))

    def _color_dist(self, a, b) -> float:
        return ((a[0] - b[0]) ** 2 + (a[1] - b[1]) ** 2 + (a[2] - b[2]) ** 2) ** 0.5

    def _slot_color_map(self):
        """
        Map slot_index -> expected fill RGB(0..1) derived from your _WIZ_HEADER_SLOT_COLOURS.
        slot_index: 0..3 corresponds to your UI order.
        """
        m = {}
        for i, (_, bg_hex, _) in enumerate(self._WIZ_HEADER_SLOT_COLOURS):
            m[i] = self._rgb01_from_hex(bg_hex)
        return m

    def _column_bounds_from_slots(self, slots: dict, page_width: float):
        """
        slots: dict {slot_index: rect} for TOP row (preferred).
        Returns 4 (l,r) bounds, stable, in slot index order.
        """
        # fallback to quarters if missing
        if len(slots) < 4:
            q = page_width / 4.0
            return [(0, q), (q, 2 * q), (2 * q, 3 * q), (3 * q, page_width)]

        # use centers from rects by slot index order
        centers = [((slots[i].x0 + slots[i].x1) / 2.0) for i in range(4)]
        mids = [(centers[i] + centers[i + 1]) / 2.0 for i in range(3)]
        return [(0, mids[0]), (mids[0], mids[1]), (mids[1], mids[2]), (mids[2], page_width)]

    def _pill_inner_box(self, rect: fitz.Rect):
        """
        Return an inner box inside the pill that leaves room for text.
        The previous iy (0.25 * pad_y) produced ~7-8pt tall boxes, too small for 10.5pt text.
        """
        pad_x = getattr(self, "_header_pad_x", 1.0)
        pad_y = getattr(self, "_header_pad_y", 1.0)

        # Horizontal inset can stay fairly strong (avoids rounded border)
        ix = rect.width * (0.18 * pad_x)

        # Vertical inset MUST be smaller to preserve height for text
        # 0.12 keeps ~13pt inner height for a 17pt pill before pad_y scaling.
        iy = rect.height * (0.12 * pad_y)

        # Clamp so we never invert
        ix = min(ix, rect.width * 0.45)
        iy = min(iy, rect.height * 0.35)

        return fitz.Rect(rect.x0 + ix, rect.y0 + iy, rect.x1 - ix, rect.y1 - iy)

    def _whiteout_box(self, page, box: fitz.Rect, fill=(1, 1, 1)):
        page.draw_rect(box, color=fill, fill=fill, overlay=True)

    def _slot_bg_rgb01(self, slot_index: int):
        _, bg_hex, _ = self._WIZ_HEADER_SLOT_COLOURS[int(slot_index)]
        return self._rgb01_from_hex(bg_hex)

    def _insert_centered_text(self, page, box: fitz.Rect, text: str, font_size: float, slot_index: int):
        """
        Deterministic, single-line centered text.
        Fixed font sizes:
          - Pills: 12
          - Table headers: 7
        Shrink only if it physically can't fit.
        """
        nudge_x = getattr(self, "_header_nudge_x", 0.0)
        nudge_y = getattr(self, "_header_nudge_y", 0.0)
        fg = self._pdf_slot_fg_rgb01(slot_index)
        dbg = bool(getattr(self, "_hdr_debug", False))

        box = fitz.Rect(box.x0 + nudge_x, box.y0 + nudge_y, box.x1 + nudge_x, box.y1 + nudge_y)

        # Fixed sizes: table row is short; pills are tall.
        base_fs = 7.0 if box.height <= 14.0 else 12.0
        fs_min = 6.0

        # Prefer bold; fall back to Base14
        font_candidates = ("hebo", "helvb", "helv")

        def text_width(fontname: str, fontsize: float) -> float:
            return fitz.get_text_length(text, fontname=fontname, fontsize=fontsize)

        for fontname in font_candidates:
            try:
                fs = base_fs
                while fs >= fs_min and text_width(fontname, fs) > box.width:
                    fs -= 0.5
                fs = max(fs_min, fs)

                tw = text_width(fontname, fs)
                x = box.x0 + max(0.0, (box.width - tw) / 2.0)

                # Stable baseline centering (doesn't shift when fs changes)
                y = box.y0 + (box.height / 2.0) + (fs * 0.33)

                page.insert_text(
                    fitz.Point(x, y),
                    text,
                    fontsize=fs,
                    fontname=fontname,
                    color=fg,
                    overlay=True,
                )
                return
            except Exception as e:
                if dbg:
                    print(f"[WARN] insert_text failed font={fontname}: {e}")

        if dbg:
            print(f"[WARN] all font candidates failed for text={text!r} box={tuple(box)}")

    def _table_font_size_for_text(self, text: str) -> float:
        """
        Table header font rule:
          - 6 chars or fewer: large
          - more than 6 chars: small
        """
        t = (text or "").strip()
        return 4 if len(t) > 6 else 6.0

    def _insert_centered_table_text(self, page, box: fitz.Rect, text: str, slot_index: int, font_size: float):
        """
        Table-only text insertion.
        Uses the exact font_size passed in. No auto-shrinking.
        """
        fg = self._pdf_slot_fg_rgb01(slot_index)
        dbg = bool(getattr(self, "_hdr_debug", False))
        font_candidates = ("hebo", "helvb", "helv")

        def text_width(fontname: str, fontsize: float) -> float:
            return fitz.get_text_length(text, fontname=fontname, fontsize=fontsize)

        for fontname in font_candidates:
            try:
                tw = text_width(fontname, font_size)
                x = box.x0 + max(0.0, (box.width - tw) / 2.0)
                y = box.y0 + (box.height / 2.0) + (font_size * 0.33)

                page.insert_text(
                    fitz.Point(x, y),
                    text,
                    fontsize=font_size,
                    fontname=fontname,
                    color=fg,
                    overlay=True,
                )
                return
            except Exception as e:
                if dbg:
                    print(f"[WARN] _insert_centered_table_text failed font={fontname}: {e}")

        if dbg:
            print(f"[WARN] all table font candidates failed for text={text!r} box={tuple(box)}")


    def _extract_header_spans(self, page):
        """
        Return list of candidate spans: (cy, cx, bbox, text, size)
        Only spans that look like headers.
        """
        raw = page.get_text("rawdict") or {}
        out = []
        for block in raw.get("blocks", []):
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    txt = (span_text(span) or "").strip()
                    if not self._is_header_token(txt):
                        continue
                    x0, y0, x1, y1 = span.get("bbox", [0, 0, 0, 0])
                    cx = (x0 + x1) / 2.0
                    cy = (y0 + y1) / 2.0
                    out.append((cy, cx, (x0, y0, x1, y1), txt, float(span.get("size") or 10.0)))
        return out

    def _cluster_by_y(self, spans, gap=28.0):
        """
        Cluster spans into bands by Y proximity.
        Returns list of bands (each band is list of span tuples), top-to-bottom.
        """
        if not spans:
            return []
        spans = sorted(spans, key=lambda s: s[0])
        bands = [[spans[0]]]
        for s in spans[1:]:
            if abs(s[0] - bands[-1][-1][0]) <= gap:
                bands[-1].append(s)
            else:
                bands.append([s])
        # sort each band left-to-right for determinism
        return [sorted(b, key=lambda s: s[1]) for b in bands]

    def _pick_header_bands(self, page):
        """
        We expect 2-3 header bands:
          - top pills
          - table header row
          - bottom pills
        Choose the top 3 bands by (span count), but keep vertical order.
        """
        spans = self._extract_header_spans(page)
        bands = self._cluster_by_y(spans, gap=28.0)
        if not bands:
            return []

        # take bands with meaningful density
        bands_scored = [(len(b), i, b) for i, b in enumerate(bands)]
        bands_scored.sort(reverse=True, key=lambda x: x[0])

        chosen = sorted([b for _, _, b in bands_scored[:3]], key=lambda b: sum(s[0] for s in b) / max(1, len(b)))
        return chosen

    def _band_slot_bounds(self, page, band):
        """
        For a given band (list of spans), compute 4 slot x-bounds using the 4 strongest x-clusters.
        This works even if order is reversed (bottom row).
        """
        width = float(page.rect.width)
        xs = sorted([s[1] for s in band])
        if len(xs) < 4:
            q = width / 4.0
            return [(0, q), (q, 2 * q), (2 * q, 3 * q), (3 * q, width)]

        # pick 4 representative centers by quantiles
        # (robust vs duplicates / extra tokens)
        qpos = [0.125, 0.375, 0.625, 0.875]
        centers = []
        for qp in qpos:
            idx = int(round(qp * (len(xs) - 1)))
            centers.append(xs[idx])
        centers = sorted(centers)

        mids = [(centers[i] + centers[i + 1]) / 2.0 for i in range(3)]
        return [(0, mids[0]), (mids[0], mids[1]), (mids[1], mids[2]), (mids[2], width)]

    def _slot_index_for_x(self, slot_bounds, cx):
        for i, (l, r) in enumerate(slot_bounds):
            if l <= cx <= r:
                return i
        return 0 if cx < slot_bounds[0][0] else 3

    def _page_header_targets(self, page):
        """
        Returns list of targets:
          (slot_index, bbox, font_size)
        across all chosen header bands.
        """
        bands = self._pick_header_bands(page)
        targets = []
        for band in bands:
            slot_bounds = self._band_slot_bounds(page, band)
            for cy, cx, bbox, txt, size in band:
                si = self._slot_index_for_x(slot_bounds, cx)
                targets.append((si, bbox, size))
        return targets

    def _blank_rect(self, page, rect, slot_bounds, slot_index):
        pad_x = getattr(self, "_header_pad_x", 1.0)
        pad_y = getattr(self, "_header_pad_y", 1.0)

        l, r = slot_bounds[slot_index]
        x0, y0, x1, y1 = rect

        dx = (x1 - x0) * 0.25 * pad_x
        dy = (y1 - y0) * 0.45 * pad_y

        rx0 = max(l, x0 - dx)
        rx1 = min(r, x1 + dx)
        ry0 = y0 - dy
        ry1 = y1 + dy

        page.draw_rect(fitz.Rect(rx0, ry0, rx1, ry1), color=(1, 1, 1), fill=(1, 1, 1), overlay=True)

    # ---------------------------
    # Header options persistence
    # ---------------------------

    def _ensure_header_options_table(self):
        conn = self._db_conn()
        cur = conn.cursor()
        cur.execute("""
                    CREATE TABLE IF NOT EXISTS mps_header_options
                    (
                        option_text
                        TEXT
                        PRIMARY
                        KEY
                    )
                    """)
        # seed defaults
        for t in ["B'FAST", "BREAKFAST", "MORNING", "LUNCH", "DINNER", "BED"]:
            cur.execute("INSERT OR IGNORE INTO mps_header_options(option_text) VALUES (?)", (t,))
        conn.commit()
        if conn is not getattr(self.parent(), "conn", None):
            conn.close()

    def _load_header_options(self) -> list:
        self._ensure_header_options_table()
        conn = self._db_conn()
        cur = conn.cursor()
        cur.execute("SELECT option_text FROM mps_header_options ORDER BY option_text COLLATE NOCASE")
        rows = cur.fetchall()
        if conn is not getattr(self.parent(), "conn", None):
            conn.close()
        return [r[0] for r in rows if r and r[0]]

    def _save_header_option(self, text: str) -> None:
        text = (text or "").strip()
        if not text:
            return
        self._ensure_header_options_table()
        conn = self._db_conn()
        cur = conn.cursor()
        cur.execute("INSERT OR IGNORE INTO mps_header_options(option_text) VALUES (?)", (text,))
        conn.commit()
        if conn is not getattr(self.parent(), "conn", None):
            conn.close()

    def _refresh_all_header_dropdowns(self):
        opts = self._load_header_options()
        for cb in getattr(self, "header_combos", []):
            cur = cb.currentText()
            cb.blockSignals(True)
            cb.clear()
            cb.addItems(opts)
            cb.setCurrentText(cur)
            cb.blockSignals(False)

    # ---------------------------------
    # Calibration persistence (global)
    # ---------------------------------



    """
    Multi-patient aware MPS blister wizard.
    Each patient gets independent weeks/blister + header settings.
    """
    # --- Wizard UI header colours (UI only; does NOT affect PDF output) ---
    _WIZ_HEADER_SLOT_COLOURS = [
        ("B'FAST", "#F4C430", "#000000"),  # yellow, black text
        ("LUNCH", "#8BC34A", "#000000"),  # green, black text
        ("DINNER", "#18B7C3", "#000000"),  # teal, black text
        ("BED", "#0E2A47", "#FFFFFF"),  # navy, white text
    ]

    def _ensure_header_calibration_table(self):
        conn = self._db_conn()
        cur = conn.cursor()

        cur.execute("""
                    CREATE TABLE IF NOT EXISTS mps_header_calibration
                    (
                        id
                        INTEGER
                        PRIMARY
                        KEY
                        CHECK
                    (
                        id =
                        1
                    ),
                        pad_x REAL DEFAULT 1.00,
                        pad_y REAL DEFAULT 1.00
                        )
                    """)

        cur.execute("PRAGMA table_info(mps_header_calibration)")
        cols = {row[1] for row in cur.fetchall()}

        if "nudge_x" not in cols:
            cur.execute("ALTER TABLE mps_header_calibration ADD COLUMN nudge_x REAL DEFAULT 0.0")
        if "nudge_y" not in cols:
            cur.execute("ALTER TABLE mps_header_calibration ADD COLUMN nudge_y REAL DEFAULT 0.0")

        cur.execute("""
                    INSERT
                    OR IGNORE INTO mps_header_calibration(id, pad_x, pad_y, nudge_x, nudge_y)
            VALUES (1, 1.00, 1.00, 0.0, 0.0)
                    """)

        conn.commit()
        if conn is not getattr(self.parent(), "conn", None):
            conn.close()

    def _load_header_calibration(self):
        self._ensure_header_calibration_table()
        conn = self._db_conn()
        cur = conn.cursor()
        cur.execute("SELECT pad_x, pad_y, nudge_x, nudge_y FROM mps_header_calibration WHERE id = 1")
        row = cur.fetchone() or (
            self.CAL_ZERO_PAD_X,
            self.CAL_ZERO_PAD_Y,
            self.CAL_ZERO_NUDGE_X,
            self.CAL_ZERO_NUDGE_Y,
        )
        if conn is not getattr(self.parent(), "conn", None):
            conn.close()
        return float(row[0] or 1.0), float(row[1] or 1.0), float(row[2] or 0.0), float(row[3] or 0.0)

    def _save_header_calibration(self, pad_x: float, pad_y: float, nudge_x: float, nudge_y: float):
        self._ensure_header_calibration_table()
        conn = self._db_conn()
        cur = conn.cursor()
        cur.execute(
            "UPDATE mps_header_calibration SET pad_x = ?, pad_y = ?, nudge_x = ?, nudge_y = ? WHERE id = 1",
            (float(pad_x), float(pad_y), float(nudge_x), float(nudge_y)),
        )
        conn.commit()
        if conn is not getattr(self.parent(), "conn", None):
            conn.close()

    def _print(self):
        # Ensure current UI is captured
        self._save_ui_to_patient()

        # Ensure work_doc is up to date
        self._build_final_doc()

        if not self.work_doc or self.work_doc.page_count == 0:
            QMessageBox.warning(self, "Nothing to print", "No pages available to print.")
            return

        # --- Validate printer using saved MPS printer name ---
        available_printers = {
            p.printerName().strip(): p
            for p in QPrinterInfo.availablePrinters()
            if p.printerName().strip()
        }

        mps_printer_name = ""
        mps_orientation = "portrait"
        try:
            settings = _load_daacal_settings()
            mps_printer_name = (settings.get("mps_default_printer") or "").strip()
            mps_orientation = (settings.get("mps_default_orientation") or "portrait").strip().lower()
        except Exception:
            mps_printer_name = ""
            mps_orientation = "portrait"

        if mps_orientation not in ("portrait", "landscape"):
            mps_orientation = "portrait"

        if not mps_printer_name:
            QMessageBox.warning(
                self,
                "MPS Printer Not Set",
                "Please set an MPS default printer in Manage -> Set Printer."
            )
            return

        if mps_printer_name not in available_printers:
            QMessageBox.warning(
                self,
                "MPS Printer Unavailable",
                f"The saved MPS printer '{mps_printer_name}' is not currently available.\n"
                f"Use Manage -> Refresh Printers, then Set Printer."
            )
            return

        qt_orientation = QPageLayout.Landscape if mps_orientation == "landscape" else QPageLayout.Portrait
        legacy_orientation = QPrinter.Landscape if mps_orientation == "landscape" else QPrinter.Portrait
        expected_orientation_text = "landscape" if mps_orientation == "landscape" else "portrait"

        qt_printer_info = available_printers[mps_printer_name]
        printer = QPrinter(qt_printer_info)
        printer.setPrinterName(mps_printer_name)
        printer.setOutputFormat(QPrinter.NativeFormat)
        printer.setDocName("MPS Blister PDF")
        printer.setCopyCount(1)

        # Force A5 / selected orientation / single-sided as hard as Qt allows
        a5_layout = QPageLayout(
            QPageSize(QPageSize.A5),
            qt_orientation,
            QMarginsF(0, 0, 0, 0),
            QPageLayout.Millimeter
        )

        try:
            printer.setPageLayout(a5_layout)
        except Exception:
            pass

        try:
            printer.setPageSize(QPageSize(QPageSize.A5))
        except Exception:
            try:
                printer.setPageSize(QPrinter.A5)
            except Exception:
                pass

        try:
            printer.setPageOrientation(qt_orientation)
        except Exception:
            try:
                printer.setOrientation(legacy_orientation)
            except Exception:
                pass

        try:
            printer.setDuplex(QPrinter.DuplexNone)
        except Exception:
            pass

        # Use full physical page coordinates so we can place the image ourselves.
        # This avoids "fit to printable rect" behavior hiding an A4 driver override.
        try:
            printer.setFullPage(True)
        except Exception:
            pass

        painter = QPainter()
        if not painter.begin(printer):
            QMessageBox.warning(
                self,
                "Print Failed",
                f"Could not start print job on '{mps_printer_name}'."
            )
            return

        try:
            # ------------------------------------------------------------
            # HARD VERIFY what the driver actually accepted AFTER begin()
            # ------------------------------------------------------------
            try:
                actual_layout = printer.pageLayout()
                full_mm = actual_layout.fullRect(QPageLayout.Millimeter)
                paint_mm = actual_layout.paintRect(QPageLayout.Millimeter)
                page_id = actual_layout.pageSize().id()
            except Exception:
                actual_layout = None
                full_mm = None
                paint_mm = None
                page_id = None

            if full_mm is None:
                painter.end()
                QMessageBox.warning(
                    self,
                    "Print Failed",
                    "Could not read printer page size after starting the print job."
                )
                return

            actual_w = float(full_mm.width())
            actual_h = float(full_mm.height())

            print(
                "MPS PRINT DEBUG | printer=", mps_printer_name,
                "| requested_orientation=", expected_orientation_text,
                "| full_mm=", round(actual_w, 2), "x", round(actual_h, 2),
                "| paint_mm=", round(float(paint_mm.width()), 2), "x", round(float(paint_mm.height()), 2)
                if paint_mm is not None else "n/a",
                "| page_id=", page_id,
                "| dpi=", printer.resolution()
            )

            # Allow a few mm tolerance for driver rounding.
            is_a5_portrait = abs(actual_w - 148.0) <= 5.0 and abs(actual_h - 210.0) <= 5.0
            is_a5_landscape = abs(actual_w - 210.0) <= 5.0 and abs(actual_h - 148.0) <= 5.0
            is_a4_portrait = abs(actual_w - 210.0) <= 5.0 and abs(actual_h - 297.0) <= 5.0
            is_a4_landscape = abs(actual_w - 297.0) <= 5.0 and abs(actual_h - 210.0) <= 5.0

            if is_a4_portrait or is_a4_landscape:
                painter.end()
                QMessageBox.critical(
                    self,
                    "Printer Forced A4",
                    f"The printer driver ignored the required A5 setting and switched to A4.\n\n"
                    f"Printer: {mps_printer_name}\n"
                    f"Expected orientation: {expected_orientation_text}\n"
                    f"Actual job size reported by driver: {actual_w:.1f} x {actual_h:.1f} mm\n\n"
                    f"Open the Windows printer defaults for this printer and set:\n"
                    f"- Paper size: A5\n"
                    f"- Orientation: {expected_orientation_text.capitalize()}\n"
                    f"- Duplex: Off / Simplex\n"
                    f"- Scaling: 100% / Actual Size\n"
                    f"- No Fit to Page / Shrink / Expand\n"
                    f"- No 'Use printer default paper' override"
                )
                return

            actual_is_expected = is_a5_landscape if mps_orientation == "landscape" else is_a5_portrait

            if not actual_is_expected:
                painter.end()
                QMessageBox.critical(
                    self,
                    "Printer Orientation Mismatch",
                    f"The printer driver did not accept the required A5 {expected_orientation_text} job.\n\n"
                    f"Printer: {mps_printer_name}\n"
                    f"Actual job size reported by driver: {actual_w:.1f} x {actual_h:.1f} mm"
                )
                return

            # ------------------------------------------------------------
            # Print each PDF page at ACTUAL SIZE on verified A5 page
            # ------------------------------------------------------------
            dpi = printer.resolution() or 300

            # Full physical page rect in device pixels after driver finalization
            try:
                page_rect = printer.pageRect(QPrinter.DevicePixel)
            except Exception:
                page_rect = printer.pageRect()

            page_rect_w = int(page_rect.width())
            page_rect_h = int(page_rect.height())

            for i in range(self.work_doc.page_count):
                if i > 0:
                    if not printer.newPage():
                        raise RuntimeError("Failed to start a new printer page.")
                    try:
                        page_rect = printer.pageRect(QPrinter.DevicePixel)
                    except Exception:
                        page_rect = printer.pageRect()
                    page_rect_w = int(page_rect.width())
                    page_rect_h = int(page_rect.height())

                page = self.work_doc.load_page(i)

                # PDF size in points; A5 should be ~419.53 x 595.28 pt
                pdf_rect = page.rect
                pdf_w_pt = float(pdf_rect.width)
                pdf_h_pt = float(pdf_rect.height)

                print(
                    f"MPS PRINT DEBUG | page {i + 1} pdf_pt={pdf_w_pt:.2f} x {pdf_h_pt:.2f} "
                    f"| printer_px={page_rect_w} x {page_rect_h}"
                )

                # Render at printer DPI so 1 PDF point (1/72") maps to exact printer pixels.
                mat = fitz.Matrix(dpi / 72.0, dpi / 72.0)
                pix = page.get_pixmap(matrix=mat, alpha=False)

                img = QImage(
                    pix.samples,
                    pix.width,
                    pix.height,
                    pix.stride,
                    QImage.Format_RGB888,
                ).copy()


                if mps_orientation == "portrait":
                    try:
                        transform = QTransform()
                        transform.rotate(90)
                        img = img.transformed(transform, Qt.SmoothTransformation)
                    except Exception:
                        pass

                # ACTUAL SIZE:
                # use rendered PDF size directly, do NOT scale to target rect
                draw_w = int(img.width())
                draw_h = int(img.height())

                # Center on the physical A5 page
                x = int(page_rect.x() + (page_rect_w - draw_w) / 2)
                y = int(page_rect.y() + (page_rect_h - draw_h) / 2)

                painter.drawImage(QRectF(x, y, draw_w, draw_h), img)

        except Exception as e:
            try:
                painter.end()
            except Exception:
                pass
            QMessageBox.critical(
                self,
                "Print Failed",
                f"An error occurred while printing:\n\n{e}"
            )
            return

        try:
            painter.end()
        except Exception:
            pass

    def _persist_all_settings_to_db(self):
        """Persist all patient_overrides to DB (same logic as Save As…)."""
        for norm, disp in (self.ordered_patients or []):
            ov = self.patient_overrides.get(norm, {"weeks": 1, "headers": [], "headers_blank": [False] * 4})
            self._save_patient_settings(
                norm,
                disp,
                ov.get("weeks") or 1,
                ov.get("headers") or [],
                ov.get("headers_blank") or [False, False, False, False],
            )
    def _apply(self):
        """Apply current settings to preview and save per-patient settings to DB."""
        # Ensure current patient's latest UI is captured
        self._save_ui_to_patient()

        # Build output doc and refresh preview
        self._build_final_doc()

        # Persist settings (same as Save As…)
        self._persist_all_settings_to_db()

    def _style_header_combo_as_pill(self, cb: QComboBox, slot_index: int):
        """
        Style a header QComboBox like the weeks/blister pills.
        Colour is FIXED by slot_index (0..3) so it always matches the PDF header bar,
        regardless of selected/custom text. UI-only.
        """
        slot_index = max(0, min(3, int(slot_index)))
        _, bg, fg = self._WIZ_HEADER_SLOT_COLOURS[slot_index]

        # Must be editable to support custom headers
        cb.setEditable(True)

        # Center the editable text
        if cb.lineEdit():
            cb.lineEdit().setAlignment(Qt.AlignCenter)

        cb.setStyleSheet(f"""
            QComboBox {{
                background: {bg};
                color: {fg};
                border: 2px solid #000;
                border-radius: 14px;
                padding: 6px 14px;
                font-weight: 700;
                font-size: 11pt;
                min-height: 30px;
            }}
            QComboBox::drop-down {{
                border: 0px;
                width: 26px;
            }}
            QComboBox QAbstractItemView {{
                border: 2px solid #000;
                border-radius: 10px;
                padding: 6px;
                background: #FFFFFF;   /* dropdown list stays readable */
                color: #000000;
            }}
            QLineEdit {{
                border: none;
                background: transparent;
                color: {fg};
                font-weight: 700;
                font-size: 11pt;
            }}
        """)

    def _svg_icon(self, path: str, w: int, h: int) -> QIcon:
        """
        Rasterise SVG -> QPixmap (high-res) -> scale down.
        Preserves aspect ratio and prevents QtSvg from "stretching" the artwork.
        """
        renderer = QSvgRenderer(path)
        if not renderer.isValid():
            pm = QPixmap(w, h)
            pm.fill(Qt.transparent)
            return QIcon(pm)

        # Render at higher resolution for clean edges
        scale = 4
        rw, rh = w * scale, h * scale

        img = QImage(rw, rh, QImage.Format_ARGB32_Premultiplied)
        img.fill(Qt.transparent)

        painter = QPainter(img)
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setRenderHint(QPainter.SmoothPixmapTransform, True)

        # IMPORTANT: keep aspect ratio by rendering into a centred rect
        # that matches the icon's aspect ratio.
        target = QRectF(0, 0, rw, rh)

        # Use the SVG's default size to determine its native aspect ratio
        ds = renderer.defaultSize()
        if ds.width() > 0 and ds.height() > 0:
            svg_ar = ds.width() / ds.height()
            tgt_ar = rw / rh

            if svg_ar > tgt_ar:
                # SVG is wider: fit width, reduce height
                new_h = rw / svg_ar
                y = (rh - new_h) / 2
                target = QRectF(0, y, rw, new_h)
            else:
                # SVG is taller: fit height, reduce width
                new_w = rh * svg_ar
                x = (rw - new_w) / 2
                target = QRectF(x, 0, new_w, rh)

        renderer.render(painter, target)
        painter.end()

        pm_hi = QPixmap.fromImage(img)

        # Scale down to final icon size (smooth)
        pm = pm_hi.scaled(w, h, Qt.KeepAspectRatio, Qt.SmoothTransformation)

        # Centre into exact WxH pixmap (so iconSize works perfectly)
        final_pm = QPixmap(w, h)
        final_pm.fill(Qt.transparent)
        p = QPainter(final_pm)
        xoff = (w - pm.width()) // 2
        yoff = (h - pm.height()) // 2
        p.drawPixmap(xoff, yoff, pm)
        p.end()

        return QIcon(final_pm)

    def _parse_page_dates(self, doc):
        """
        Return list of (first_day, last_day, expiry) as datetime.date per page.

        Preferred method:
          Parse the labeled fields:
            First Day: DD Mon YY
            Last Day:  DD Mon YY
            Expiry:    DD Mon YY

        Fallback:
          Use generic date extraction, but allow the valid case where
          first_day == last_day and expiry is the only other distinct date.
        """
        import datetime as dt
        pages = []

        label_patterns = {
            "first": re.compile(r"First Day:\s*(\d{1,2} \w{3} \d{2})"),
            "last": re.compile(r"Last Day:\s*(\d{1,2} \w{3} \d{2})"),
            "expiry": re.compile(r"Expiry:\s*(\d{1,2} \w{3} \d{2})"),
        }

        generic_pattern = re.compile(r"\d{1,2} \w{3} \d{2}")

        for pi in range(doc.page_count):
            txt = doc.load_page(pi).get_text("text") or ""

            # --- Best path: parse labeled fields directly ---
            m_first = label_patterns["first"].search(txt)
            m_last = label_patterns["last"].search(txt)
            m_exp = label_patterns["expiry"].search(txt)

            if m_first and m_last and m_exp:
                first = dt.datetime.strptime(m_first.group(1), "%d %b %y").date()
                last = dt.datetime.strptime(m_last.group(1), "%d %b %y").date()
                expiry = dt.datetime.strptime(m_exp.group(1), "%d %b %y").date()
                pages.append((first, last, expiry))
                continue

            # --- Fallback: generic extraction ---
            dates = generic_pattern.findall(txt)

            # drop DOB-ish stuff: year < 20
            filtered = [d for d in dates if int(d.split()[-1]) >= 20]

            # keep order, but do NOT require 3 unique dates
            seen = set()
            unique = []
            for d in filtered:
                if d not in seen:
                    seen.add(d)
                    unique.append(d)

            if len(unique) >= 3:
                first = dt.datetime.strptime(unique[0], "%d %b %y").date()
                last = dt.datetime.strptime(unique[1], "%d %b %y").date()
                expiry = dt.datetime.strptime(unique[-1], "%d %b %y").date()
            elif len(unique) == 2:
                # valid single-day page: first == last, expiry is second
                first = dt.datetime.strptime(unique[0], "%d %b %y").date()
                last = first
                expiry = dt.datetime.strptime(unique[1], "%d %b %y").date()
            else:
                raise ValueError(
                    f"Page {pi + 1} does not contain enough usable date fields: {unique}"
                )

            pages.append((first, last, expiry))

        return pages

    def _replace_dates_on_page(self, page, replacements: dict):
        """
        replacements: dict {old_text_str: new_text_str}
        Uses your existing span_text / replace_span_text helpers.
        """
        raw = page.get_text("rawdict")
        for block in raw.get("blocks", []):
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    txt = span_text(span).strip()
                    new = replacements.get(txt)
                    if new is not None:
                        replace_span_text(page, span, new)

    def _build_weeks_doc_from_source(self, source_doc, option: int):
        """
        Build and return a NEW fitz.Document according to option (1/2/4 weeks),
        using the same date-replacement logic as your original wizard.

        This does NOT touch UI state and does NOT mutate self.orig_doc.
        """
        option = int(option or 1)
        pages = self._parse_page_dates(source_doc)
        out = fitz.open()

        # --- 1 week / blister → no change ---
        if option == 1:
            out.insert_pdf(source_doc)
            return out

        # --- 4 weeks / blister → 1 page ---
        if option == 4:
            first_start = pages[0][0]
            last_end = pages[-1][1]

            # Your current code uses latest_expiry = max(...).
            # (If you later want oldest, change max->min here.)
            latest_expiry = max(p[2] for p in pages)

            out.insert_pdf(source_doc, from_page=0, to_page=0)
            page = out[0]
            old_first, old_last, old_exp = pages[0]
            repl = {
                old_first.strftime("%d %b %y"): first_start.strftime("%d %b %y"),
                old_last.strftime("%d %b %y"): last_end.strftime("%d %b %y"),
                old_exp.strftime("%d %b %y"): latest_expiry.strftime("%d %b %y"),
            }
            self._replace_dates_on_page(page, repl)
            return out

        # --- 2 weeks / blister → 2 pages ---
        from datetime import timedelta

        first_start = pages[0][0]
        last_end = pages[-1][1]

        # Page 1 last day = start + 13 (14 days inclusive)
        page1_last = first_start + timedelta(days=13)
        # Page 2 first day = start + 14
        page2_first = first_start + timedelta(days=14)

        # Expiry for page 1 and 2 follows your existing logic
        if len(pages) >= 2:
            expiry1 = pages[1][2]
        else:
            expiry1 = pages[-1][2]

        if len(pages) >= 4:
            expiry2 = pages[3][2]
        elif len(pages) >= 3:
            expiry2 = pages[2][2]
        else:
            expiry2 = pages[-1][2]

        # Page 1 uses original page 1 layout
        out.insert_pdf(source_doc, from_page=0, to_page=0)
        p0 = out[0]
        old_first0, old_last0, old_exp0 = pages[0]
        repl0 = {
            old_first0.strftime("%d %b %y"): first_start.strftime("%d %b %y"),
            old_last0.strftime("%d %b %y"): page1_last.strftime("%d %b %y"),
            old_exp0.strftime("%d %b %y"): expiry1.strftime("%d %b %y"),
        }
        self._replace_dates_on_page(p0, repl0)

        # Page 2 uses original page 3 layout (week 3) if available
        base_index = 2 if len(pages) >= 3 else 1
        base_index = min(base_index, source_doc.page_count - 1)

        out.insert_pdf(source_doc, from_page=base_index, to_page=base_index)
        p1 = out[1]
        old_first2, old_last2, old_exp2 = pages[base_index]
        repl1 = {
            old_first2.strftime("%d %b %y"): page2_first.strftime("%d %b %y"),
            old_last2.strftime("%d %b %y"): last_end.strftime("%d %b %y"),
            old_exp2.strftime("%d %b %y"): expiry2.strftime("%d %b %y"),
        }
        self._replace_dates_on_page(p1, repl1)

        return out



    def _find_pill_slots_on_page(self, page, tol=0.20):
        """
        Debug:
          - set `self._hdr_debug = True` to print fill colors + match distances.
        """
        dbg = bool(getattr(self, "_hdr_debug", False))

        draws = page.get_drawings()
        slot_colors = self._slot_color_map()

        if dbg:
            print("\n=== _find_pill_slots_on_page ===")
            print("tol =", tol)
            print("expected slot colors (rgb01):")
            for i in range(4):
                c = slot_colors[i]
                print(f" slot {i}: {tuple(round(x, 4) for x in c)}")

        def looks_like_pill_bar(r: fitz.Rect) -> bool:
            w, h = float(r.width), float(r.height)
            if not (60.0 <= w <= 110.0):
                return False
            if not (12.0 <= h <= 26.0):
                return False
            if w / max(h, 1.0) < 2.5:
                return False
            return True

        cand = {i: [] for i in range(4)}

        # Track closest misses so you can see why something didn't match
        closest_misses = []  # (best_dist, best_i, fill, rect)

        for d in draws:
            fill = d.get("fill")
            rect = d.get("rect")
            if fill is None or rect is None:
                continue

            # ignore white fills
            if self._color_dist(fill, (1.0, 1.0, 1.0)) < 1e-6:
                continue

            if not looks_like_pill_bar(rect):
                continue

            best_i, best_dist = None, 1e9
            for i, c in slot_colors.items():
                dist = self._color_dist(fill, c)
                if dist < best_dist:
                    best_i, best_dist = i, dist

            if dbg:
                print(
                    " pill?",
                    f"rect(w={rect.width:.1f},h={rect.height:.1f},y0={rect.y0:.1f})",
                    "fill=", tuple(round(x, 4) for x in fill),
                    "-> best_slot=", best_i,
                    "dist=", round(best_dist, 4),
                    "ACCEPT" if best_dist <= float(tol) else "REJECT",
                )

            if best_dist <= float(tol):
                cand[best_i].append(rect)
            else:
                closest_misses.append((best_dist, best_i, fill, rect))

        if dbg and closest_misses:
            closest_misses.sort(key=lambda x: x[0])
            print("\nclosest rejected (top 8):")
            for best_dist, best_i, fill, rect in closest_misses[:8]:
                print(
                    f" dist={best_dist:.4f} slot={best_i} fill={tuple(round(x, 4) for x in fill)} "
                    f"rect(w={rect.width:.1f},h={rect.height:.1f},y0={rect.y0:.1f})"
                )

        top, bottom = {}, {}
        for i in range(4):
            if not cand[i]:
                continue

            rs = sorted(cand[i], key=lambda r: r.y0)
            top_rect = rs[0]
            bot_rect = rs[-1] if len(rs) > 1 else None

            top[i] = top_rect
            if bot_rect is not None and abs(bot_rect.y0 - top_rect.y0) > 20:
                bottom[i] = bot_rect

        if dbg:
            print("\nselected slots:")
            for i in range(4):
                tr = top.get(i)
                br = bottom.get(i)
                print(f" slot {i}: top={tuple(tr) if tr else None} bottom={tuple(br) if br else None}")

        return top, bottom

    def _render_preview(self):
        """Render all pages of self.work_doc (or orig_doc if work_doc is None) into the left scroll area."""
        # clear layout
        while self.preview_layout.count():
            item = self.preview_layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.deleteLater()

        doc = self.work_doc if (self.work_doc is not None) else self.orig_doc
        if doc is None:
            return

        from PyQt5.QtGui import QImage, QPixmap

        for page_index in range(doc.page_count):
            page = doc.load_page(page_index)
            pix = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
            img = QImage(
                pix.samples,
                pix.width,
                pix.height,
                pix.stride,
                QImage.Format_RGBA8888 if pix.alpha else QImage.Format_RGB888,
            )
            lbl = QLabel()
            lbl.setPixmap(QPixmap.fromImage(img))
            self.preview_layout.addWidget(lbl)

        self.preview_layout.addStretch(1)

    # ------------------------------------------------------------------
    # Patient name handling
    # ------------------------------------------------------------------
    def _norm_name(self, s: str) -> str:
        s = (s or "").strip()
        s = re.sub(r"\s+", " ", s)
        return s.casefold()

    def _extract_patient_from_page(self, page):
        """
        Returns (norm, display) or (None, None)
        """
        try:
            text = page.get_text("text") or ""
        except Exception:
            return None, None

        lines = [ln.strip() for ln in text.splitlines()]
        dob_idx = None
        for i, ln in enumerate(lines):
            if re.search(r"\bDOB\b", ln, re.IGNORECASE):
                dob_idx = i
                break
        if dob_idx is None:
            return None, None

        for j in range(dob_idx - 1, -1, -1):
            cand = lines[j].strip()
            if not cand:
                continue
            if re.search(r"\b(week|blister|address|expiry|phone)\b", cand, re.I):
                continue
            if re.search(r"\d{1,2} \w{3} \d{2}", cand):
                continue
            return self._norm_name(cand), cand

        return None, None

    def _map_pages_to_patients(self):
        """
        Returns:
          blocks: [(patient_norm, patient_display, [page_idxs])]
          ordered_patients: [(norm, display)]
        """
        blocks = []
        ordered = []
        seen = set()

        cur_norm = None
        cur_disp = None
        cur_pages = []

        for i in range(len(self.orig_doc)):
            norm, disp = self._extract_patient_from_page(self.orig_doc[i])

            if cur_norm is None:
                cur_norm, cur_disp = norm, disp
                cur_pages = [i]
                continue

            if norm == cur_norm:
                cur_pages.append(i)
            else:
                blocks.append((cur_norm, cur_disp, cur_pages[:]))
                if cur_norm and cur_norm not in seen:
                    ordered.append((cur_norm, cur_disp))
                    seen.add(cur_norm)
                cur_norm, cur_disp = norm, disp
                cur_pages = [i]

        if cur_pages:
            blocks.append((cur_norm, cur_disp, cur_pages))
            if cur_norm and cur_norm not in seen:
                ordered.append((cur_norm, cur_disp))

        return blocks, ordered

    # ------------------------------------------------------------------
    # DB helpers
    # ------------------------------------------------------------------
    def _db_conn(self):
        if hasattr(self.parent(), "conn") and self.parent().conn:
            return self.parent().conn
        return sqlite3.connect(DB_FILE)

    def _load_patient_settings(self, norm):
        if not norm:
            return None
        conn = self._db_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT weeks_per_blister, header_labels_json "
            "FROM mps_patient_settings WHERE patient_name_norm = ?",
            (norm,)
        )
        row = cur.fetchone()
        if conn is not getattr(self.parent(), "conn", None):
            conn.close()
        if not row:
            return None

        weeks, hdr = row
        headers = None
        blanks = None

        if hdr:
            try:
                parsed = json.loads(hdr)
                # Backward compat:
                # - old format: ["B'FAST","LUNCH",...]
                # - new format: {"labels":[...], "blank":[...]}
                if isinstance(parsed, dict):
                    headers = parsed.get("labels")
                    blanks = parsed.get("blank")
                elif isinstance(parsed, list):
                    headers = parsed
                    blanks = [not bool(x) for x in parsed]  # empty string => blank
            except Exception:
                headers = None
                blanks = None

        return {
            "weeks": int(weeks or 1),
            "headers": headers,
            "headers_blank": blanks,
        }

    def _save_patient_settings(self, norm, display, weeks, headers, headers_blank=None):
        payload = {
            "labels": list(headers or ["", "", "", ""])[:4],
            "blank": [bool(x) for x in (headers_blank or [False, False, False, False])][:4],
        }

        conn = self._db_conn()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO mps_patient_settings
            (patient_name_norm, patient_name_display,
             weeks_per_blister, header_labels_json, updated_at)
            VALUES (?, ?, ?, ?, datetime('now')) ON CONFLICT(patient_name_norm) DO
            UPDATE SET
                patient_name_display = excluded.patient_name_display,
                weeks_per_blister = excluded.weeks_per_blister,
                header_labels_json = excluded.header_labels_json,
                updated_at = excluded.updated_at
            """,
            (norm, display, int(weeks), json.dumps(payload))
        )
        conn.commit()
        if conn is not getattr(self.parent(), "conn", None):
            conn.close()

    def _detect_header_labels(self):
        """
        Robustly detect 4 header labels on page 1.
        Handles duplicates, mixed time labels, missing/blank slots.
        """
        if not getattr(self, "orig_doc", None):
            return ["B'FAST", "LUNCH", "DINNER", "BED"]

        try:
            page = self.orig_doc.load_page(0)
        except Exception:
            return ["B'FAST", "LUNCH", "DINNER", "BED"]

        raw = page.get_text("rawdict") or {}
        width = float(page.rect.width) if hasattr(page, "rect") else 420.0

        known = {"B'FAST", "BREAKFAST", "MORNING", "LUNCH", "DINNER", "BED"}

        def is_time(tok: str) -> bool:
            t = tok.strip().upper()
            return (
                    re.fullmatch(r"\d{1,2}:\d{2}", t) is not None or
                    re.fullmatch(r"\d{1,2}\s?(AM|PM)", t) is not None
            )

        def norm_tok(tok: str) -> str:
            t = (tok or "").strip()
            t = re.sub(r"\s+", " ", t)
            return t

        candidates = []
        for block in raw.get("blocks", []):
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    txt = norm_tok(span_text(span))
                    if not txt:
                        continue
                    up = txt.upper()
                    if up in known or is_time(up) or "FAST" in up:
                        x0, y0, x1, y1 = span.get("bbox", [0, 0, 0, 0])
                        cx = (x0 + x1) / 2.0
                        cy = (y0 + y1) / 2.0
                        candidates.append((cy, cx, txt))

        if not candidates:
            return ["B'FAST", "LUNCH", "DINNER", "BED"]

        # pick densest y-band
        candidates.sort(key=lambda t: t[0])
        band_tol = 8.0
        best_band = []
        band = []

        for c in candidates:
            if not band:
                band = [c]
                continue
            if abs(c[0] - band[0][0]) <= band_tol:
                band.append(c)
            else:
                if len(band) > len(best_band):
                    best_band = band
                band = [c]
        if len(band) > len(best_band):
            best_band = band

        row = sorted(best_band, key=lambda t: t[1])

        # bin into 4 slots by x quarters
        slots = [[] for _ in range(4)]
        for cy, cx, txt in row:
            idx = int(min(3, max(0, (cx / max(width, 1.0)) * 4)))
            slots[idx].append(txt)

        defaults = ["B'FAST", "LUNCH", "DINNER", "BED"]
        out = []
        for i in range(4):
            tok_list = slots[i]
            chosen = None
            if tok_list:
                for t in tok_list:
                    up = t.upper()
                    if up in known or "FAST" in up:
                        chosen = t
                        break
                if chosen is None:
                    for t in tok_list:
                        if is_time(t.upper()):
                            chosen = t
                            break
                if chosen is None:
                    chosen = tok_list[0]
            out.append(chosen or defaults[i])

        return out

    # ------------------------------------------------------------------
    # Init
    # ------------------------------------------------------------------
    def __init__(self, parent, pdf_path):
        super().__init__(parent)
        self._hdr_debug = True
        self.setWindowTitle("Blister PDF Wizard")
        self.resize(1500, 800)
        self.setMinimumWidth(1400)

        self.orig_doc = fitz.open(pdf_path)
        self.work_doc = None

        # --- header label baseline (ONLY if detector exists) ---
        if hasattr(self, "_detect_header_labels"):
            try:
                self.header_labels = self._detect_header_labels()
            except Exception:
                self.header_labels = []
        else:
            self.header_labels = []

        if not self.header_labels or len(self.header_labels) < 4:
            self.header_labels = ["B'FAST", "LUNCH", "DINNER", "BED"]

        # Detect patients
        self.patient_blocks, self.ordered_patients = self._map_pages_to_patients()

        # Per-patient overrides (session-local)
        # norm -> {"weeks": int|None, "headers": list|None}
        self.patient_overrides = {
            norm: {
                "weeks": 1,
                "headers": list(self.header_labels),  # <- IMPORTANT: uses PDF-detected headers (e.g. 08:05/08:10/08:15)
                "headers_blank": [False, False, False, False]
            }
            for norm, _ in (self.ordered_patients or [])
        }

        # Load remembered settings into overrides (do NOT touch UI yet)
        for norm, _ in (self.ordered_patients or []):
            st = self._load_patient_settings(norm)
            if st:
                hdrs, blks = self._norm_headers_and_blanks(st.get("headers"), st.get("headers_blank"))
                self.patient_overrides[norm]["weeks"] = int(st.get("weeks") or 1)
                self.patient_overrides[norm]["headers"] = hdrs
                self.patient_overrides[norm]["headers_blank"] = blks
        # ---------------- UI ----------------
        outer = QHBoxLayout(self)

        # Preview (left)
        self.preview_scroll = QScrollArea()
        self.preview_scroll.setWidgetResizable(True)
        self.preview_container = QWidget()
        self.preview_layout = QVBoxLayout(self.preview_container)
        self.preview_scroll.setWidget(self.preview_container)
        outer.addWidget(self.preview_scroll, 2)

        # Right panel
        right = QVBoxLayout()
        outer.addLayout(right, 1)

        self.stack = QStackedWidget()
        right.addWidget(self.stack)

        # Step 0 – patients
        step0 = QWidget()
        l0 = QVBoxLayout(step0)
        l0.addWidget(QLabel("Recognised patients:"))
        lst = QListWidget()
        if self.ordered_patients:
            for _, disp in self.ordered_patients:
                lst.addItem(disp)
        else:
            lst.addItem("No patient names detected.")
        l0.addWidget(lst)
        self.stack.addWidget(step0)

        # Step 1 – per-patient settings
        step1 = QWidget()
        l1 = QVBoxLayout(step1)

        l1.addWidget(QLabel("Patient:"))
        self.patient_combo = QComboBox()

        # Prevent signals during initial population
        self.patient_combo.blockSignals(True)
        for norm, disp in (self.ordered_patients or []):
            self.patient_combo.addItem(disp, norm)
        self.patient_combo.blockSignals(False)

        l1.addWidget(self.patient_combo)

        l1.addWidget(QLabel("Weeks per blister:"))

        # --- Icon toggle buttons (SVG) ---
        icons_dir = os.path.join(BASE_DIR, "icons")

        self.weeks_group = QButtonGroup(self)
        self.weeks_group.setExclusive(True)

        self.w1 = QToolButton()
        self.w2 = QToolButton()
        self.w4 = QToolButton()

        for btn in (self.w1, self.w2, self.w4):
            btn.setCheckable(True)
            btn.setAutoRaise(True)
            btn.setToolButtonStyle(Qt.ToolButtonIconOnly)

        icon_w, icon_h = 220, 64

        # Render SVGs to icons (your helper already insets with margin)
        self.w1.setIcon(self._svg_icon(os.path.join(icons_dir, "1weekperblister.svg"), icon_w, icon_h))
        self.w2.setIcon(self._svg_icon(os.path.join(icons_dir, "2weeksperblister.svg"), icon_w, icon_h))
        self.w4.setIcon(self._svg_icon(os.path.join(icons_dir, "4weeksperblister.svg"), icon_w, icon_h))

        # Single source of truth for geometry (this is what centres properly)
        for btn in (self.w1, self.w2, self.w4):
            btn.setIconSize(QSize(icon_w, icon_h))
            btn.setFixedSize(icon_w + 20, icon_h + 20)  # symmetric padding via geometry
            btn.setStyleSheet("""
                QToolButton {
                    border: 2px solid transparent;
                    border-radius: 10px;
                    background: transparent;
                }
                QToolButton:checked {
                    border: 2px solid #4CAF50;
                    background: rgba(76, 175, 80, 0.12);
                }
            """)

        self.w1.setToolTip("1 week per blister")
        self.w2.setToolTip("2 weeks per blister")
        self.w4.setToolTip("4 weeks per blister")

        self.weeks_group.addButton(self.w1, 1)
        self.weeks_group.addButton(self.w2, 2)
        self.weeks_group.addButton(self.w4, 4)

        weeks_col = QVBoxLayout()
        weeks_col.setAlignment(Qt.AlignHCenter)
        weeks_col.setSpacing(8)

        weeks_col.addWidget(self.w1)
        weeks_col.addWidget(self.w2)
        weeks_col.addWidget(self.w4)
        l1.addLayout(weeks_col)

        # small, controlled gap before Headers (tune if needed)
        l1.addSpacing(8)

        # --- after weeks_col is created and configured ---
        weeks_gap = weeks_col.spacing()  # = 8
        pill_w = icon_w + 20  # = 240 (same as w1/w2/w4 fixed width)

        # --- after weeks_col is created and configured ---
        weeks_gap = weeks_col.spacing()  # matches the weeks/blister gap (8)
        pill_w = icon_w + 20  # matches the pills' black outline width (240)

        l1.addWidget(QLabel("Headers:"))

        headers_col = QVBoxLayout()
        headers_col.setAlignment(Qt.AlignHCenter)
        headers_col.setSpacing(48)

        # --- load persisted dropdown options + blank calibration ---
        self._ensure_header_options_table()
        options = self._load_header_options()
        self._header_pad_x, self._header_pad_y, self._header_nudge_x, self._header_nudge_y = self._load_header_calibration()
        self.header_combos = []
        self.header_blank_checks = []
        self._header_restore_text = [""] * 4

        for i in range(4):
            roww = QWidget()
            rowl = QHBoxLayout(roww)
            rowl.setContentsMargins(0, 0, 0, 0)
            rowl.setSpacing(10)

            cb = QComboBox()
            cb.setEditable(True)
            cb.addItems(options)

            # baseline detected labels
            cb.setCurrentText(self.header_labels[i] if i < len(self.header_labels) else "")

            # keep your existing pill styling (recommended)
            self._style_header_combo_as_pill(cb, i)

            # sizing matches the weeks pills
            cb.setFixedWidth(pill_w)
            cb.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

            blank = QCheckBox("Blank")
            blank.setChecked(False)

            def _on_blank_toggled(checked, combo=cb, slot=i):
                # Tick Blank: remember current text (if any) then clear + disable.
                # Untick Blank: restore remembered text (or fallback default) then enable.
                if not hasattr(self, "_header_restore_text"):
                    self._header_restore_text = [""] * 4

                if checked:
                    cur = ""
                    if combo.isEditable() and combo.lineEdit():
                        cur = combo.lineEdit().text().strip()
                    else:
                        cur = combo.currentText().strip()

                    if cur:
                        self._header_restore_text[slot] = cur

                    combo.blockSignals(True)
                    try:
                        if combo.isEditable():
                            combo.setEditText("")
                        combo.setCurrentText("")
                    finally:
                        combo.blockSignals(False)

                    combo.setEnabled(False)

                else:
                    combo.setEnabled(True)

                    restore = (self._header_restore_text[slot] or "").strip()
                    if not restore:
                        restore = self.header_labels[slot] if slot < len(self.header_labels) else ""

                    combo.blockSignals(True)
                    try:
                        if combo.isEditable():
                            combo.setEditText(restore)
                        combo.setCurrentText(restore)
                    finally:
                        combo.blockSignals(False)

            blank.toggled.connect(_on_blank_toggled)

            def _maybe_apply():
                if getattr(self, "_ui_ready", False):
                    self._apply()

            blank.toggled.connect(_maybe_apply)
            # persist any new typed/selected option so it appears next time
            def _persist_option(_=None, combo=cb):
                t = combo.currentText().strip()
                if t:
                    self._save_header_option(t)
                    self._refresh_all_header_dropdowns()

            cb.activated.connect(_persist_option)
            if cb.lineEdit():
                cb.lineEdit().editingFinished.connect(_persist_option)

            rowl.addWidget(cb)
            rowl.addWidget(blank)

            self.header_combos.append(cb)
            self.header_blank_checks.append(blank)
            headers_col.addWidget(roww, alignment=Qt.AlignHCenter)

        l1.addLayout(headers_col)

        # --- calibration controls (affects blank masking only) ---
        # --- calibration controls (pad + nudge) ---
        cal_box = QGroupBox("Label calibration")
        cal_l = QHBoxLayout(cal_box)

        # pad controls (how much white box expands around old header)
        self.padx_spin = QDoubleSpinBox()
        self.padx_spin.setRange(-20.0, 20.0)
        self.padx_spin.setSingleStep(0.1)
        self.padx_spin.setValue(self._header_pad_x - self.CAL_ZERO_PAD_X)

        self.pady_spin = QDoubleSpinBox()
        self.pady_spin.setRange(-20.0, 20.0)
        self.pady_spin.setSingleStep(0.1)
        self.pady_spin.setValue(self._header_pad_y - self.CAL_ZERO_PAD_Y)

        # nudge controls (move replacement text in PDF points)
        self.nudgex_spin = QDoubleSpinBox()
        self.nudgex_spin.setRange(-50.0, 50.0)
        self.nudgex_spin.setSingleStep(0.5)
        self.nudgex_spin.setValue(self._header_nudge_x - self.CAL_ZERO_NUDGE_X)

        self.nudgey_spin = QDoubleSpinBox()
        self.nudgey_spin.setRange(-50.0, 50.0)
        self.nudgey_spin.setSingleStep(0.5)
        self.nudgey_spin.setValue(self._header_nudge_y - self.CAL_ZERO_NUDGE_Y)
        save_cal = QPushButton("Save")

        def _save_calibration():
            self._header_pad_x = self.CAL_ZERO_PAD_X + float(self.padx_spin.value())
            self._header_pad_y = self.CAL_ZERO_PAD_Y + float(self.pady_spin.value())
            self._header_nudge_x = self.CAL_ZERO_NUDGE_X + float(self.nudgex_spin.value())
            self._header_nudge_y = self.CAL_ZERO_NUDGE_Y + float(self.nudgey_spin.value())
            self._save_header_calibration(
                self._header_pad_x,
                self._header_pad_y,
                self._header_nudge_x,
                self._header_nudge_y,
            )
            self._apply()  # rebuild + preview

        save_cal.clicked.connect(_save_calibration)

        cal_l.addWidget(QLabel("Pad X"))
        cal_l.addWidget(self.padx_spin)
        cal_l.addWidget(QLabel("Pad Y"))
        cal_l.addWidget(self.pady_spin)
        cal_l.addWidget(QLabel("Nudge X"))
        cal_l.addWidget(self.nudgex_spin)
        cal_l.addWidget(QLabel("Nudge Y"))
        cal_l.addWidget(self.nudgey_spin)
        cal_l.addWidget(save_cal)

        l1.addWidget(cal_box)
        weeks_col.setContentsMargins(0, 0, 0, 0)
        headers_col.setContentsMargins(0, 0, 0, 0)

        self.stack.addWidget(step1)

        # Step 2 – preview
        step2 = QWidget()
        l2 = QVBoxLayout(step2)
        l2.addWidget(QLabel("Preview on the left"))
        self.stack.addWidget(step2)

        # Nav
        nav = QHBoxLayout()
        right.addLayout(nav)
        self.back = QPushButton("Back")
        self.next = QPushButton("Next")
        self.print_btn = QPushButton("Print…")
        self.save = QPushButton("Save As…")
        nav.addWidget(self.back)
        nav.addWidget(self.next)
        nav.addWidget(self.print_btn)
        nav.addWidget(self.save)

        # Wiring
        self.back.clicked.connect(self._back)
        self.next.clicked.connect(self._next)
        self.print_btn.clicked.connect(self._print)
        self.save.clicked.connect(self._save)

        # Guard flag to prevent overwriting patient[0] with blank UI defaults on first load
        self._ui_ready = False
        # Track active patient so we can save the one we are leaving
        self._active_patient_norm = None
        # Connect AFTER _ui_ready exists
        self.patient_combo.currentIndexChanged.connect(self._load_patient_into_ui)

        # Load patient 0 into UI WITHOUT triggering "save current UI into patient 0"
        if self.patient_combo.count() > 0:
            self.patient_combo.blockSignals(True)
            self.patient_combo.setCurrentIndex(0)
            self.patient_combo.blockSignals(False)

            # Load first patient into UI
            self._load_patient_into_ui(0)
            self._active_patient_norm = self.patient_combo.currentData()
        self._ui_ready = True
        self.pdf_path = pdf_path

        # Initial preview
        self._render_preview()

    # ------------------------------------------------------------------
    # UI sync
    # ------------------------------------------------------------------
    def _save_ui_to_patient(self, norm=None):
        """
        Persist current UI selections into overrides for the given patient norm.
        If norm is None, uses current combo selection.
        """

        if not getattr(self, "_ui_ready", False):
            return

            # Force any in-progress edits to commit before reading currentText()
        for cb in getattr(self, "header_combos", []):
            if cb.isEditable() and cb.lineEdit():
                cb.lineEdit().clearFocus()

        if norm is None:
            norm = self.patient_combo.currentData()
        if not norm:
            return

        # Ensure structure exists
        if norm not in self.patient_overrides:
            self.patient_overrides[norm] = {
                "weeks": None,
                "headers": None,
                "headers_blank": [False, False, False, False],
            }

        weeks = int(self.weeks_group.checkedId() or 1)

        # _get_ui_headers must return (labels, blanks)
        headers, blank_flags = self._get_ui_headers()
        if getattr(self, "_hdr_debug", False):
            print("UI readback:")
            for i, (cb, chk) in enumerate(zip(self.header_combos, self.header_blank_checks)):
                t = cb.lineEdit().text().strip() if (cb.isEditable() and cb.lineEdit()) else cb.currentText().strip()
                print(f" slot {i}: ui_text={t!r} ui_blank={chk.isChecked()} ui_enabled={cb.isEnabled()}")
        self.patient_overrides[norm]["weeks"] = weeks
        self.patient_overrides[norm]["headers"] = headers
        self.patient_overrides[norm]["headers_blank"] = blank_flags

    def _load_patient_into_ui(self, idx):
        """
        On patient dropdown change:
          - Save previous patient's UI state into overrides (only when ui ready)
          - Load new patient's state into UI WITHOUT triggering _apply mid-load
        """
        if idx < 0:
            return

        # Save the patient we are leaving
        prev_norm = getattr(self, "_active_patient_norm", None)
        if prev_norm and getattr(self, "_ui_ready", False):
            self._save_ui_to_patient(prev_norm)

        new_norm = self.patient_combo.itemData(idx)
        self._active_patient_norm = new_norm

        ov = self.patient_overrides.get(new_norm, {})
        weeks = int(ov.get("weeks") or 1)

        headers, blank_flags = self._norm_headers_and_blanks(
            ov.get("headers"),
            ov.get("headers_blank"),
        )

        # Block signals so blank toggles / combo signals can't call _apply during load
        for w in (self.w1, self.w2, self.w4, *self.header_combos, *self.header_blank_checks):
            w.blockSignals(True)

        try:
            self._set_weeks_ui(weeks)
            self._set_headers_ui(headers, blank_flags=blank_flags)
        finally:
            for w in (self.w1, self.w2, self.w4, *self.header_combos, *self.header_blank_checks):
                w.blockSignals(False)
    def _set_weeks_ui(self, w):
        {1: self.w1, 2: self.w2, 4: self.w4}.get(int(w or 1), self.w1).setChecked(True)

    def _get_ui_headers(self):
        labels = []
        blanks = []
        for cb, chk in zip(self.header_combos, self.header_blank_checks):
            txt = ""
            if cb.isEditable() and cb.lineEdit():
                txt = cb.lineEdit().text().strip()
            else:
                txt = cb.currentText().strip()

            is_blank = bool(chk.isChecked())

            blanks.append(is_blank)
            labels.append("" if is_blank else txt)

        return labels, blanks

    def _set_headers_ui(self, headers, blank_flags=None):
        headers, blank_flags = self._norm_headers_and_blanks(headers, blank_flags)

        if not hasattr(self, "_header_restore_text"):
            self._header_restore_text = [""] * 4

        # If something is being set as blank but we have a header value, keep it for restore.
        for i in range(4):
            if bool(blank_flags[i]) and (headers[i] or "").strip():
                self._header_restore_text[i] = (headers[i] or "").strip()

        # Apply combo text (blank => empty)
        for i, cb in enumerate(self.header_combos):
            value = "" if bool(blank_flags[i]) else (headers[i] or "")
            cb.blockSignals(True)
            try:
                if cb.isEditable():
                    cb.setEditText(value)
                cb.setCurrentText(value)
            finally:
                cb.blockSignals(False)

        # Apply checkbox state
        for i, chk in enumerate(self.header_blank_checks):
            chk.blockSignals(True)
            try:
                chk.setChecked(bool(blank_flags[i]))
            finally:
                chk.blockSignals(False)

        # Keep enabled/disabled consistent (signals might be blocked)
        for cb, chk in zip(self.header_combos, self.header_blank_checks):
            cb.setEnabled(not chk.isChecked())
    # Navigation / build
    # ------------------------------------------------------------------
    def _next(self):
        # Step 0 -> Step 1
        if self.stack.currentIndex() == 0:
            self.stack.setCurrentIndex(1)
            self.next.setText("Apply")
            return

        # Step 1 -> APPLY (do NOT navigate away)
        if self.stack.currentIndex() == 1:
            self._apply()
            return

    def _back(self):
        # Step 1 -> Step 0
        if self.stack.currentIndex() == 1:
            self.stack.setCurrentIndex(0)
            self.next.setText("Next")
            return

        self.stack.setCurrentIndex(0)
        self.next.setText("Next")

    def _build_final_doc(self):
        out = fitz.open()

        for norm, disp, pages in self.patient_blocks:
            if not pages:
                continue

            temp = fitz.open()
            patient_doc = None
            try:
                temp.insert_pdf(self.orig_doc, from_page=pages[0], to_page=pages[-1])

                ov = self.patient_overrides.get(norm, {}) or {}
                weeks = int(ov.get("weeks") or 1)

                headers, blank_flags = self._norm_headers_and_blanks(
                    ov.get("headers"),
                    ov.get("headers_blank"),
                )

                patient_doc = self._build_weeks_doc_from_source(temp, weeks)

                backup = self.work_doc
                try:
                    self.work_doc = patient_doc
                    self._apply_header_changes_from_list(headers, blank_flags_4=blank_flags)
                finally:
                    self.work_doc = backup

                out.insert_pdf(patient_doc)

            finally:
                try:
                    temp.close()
                except Exception:
                    pass
                try:
                    if patient_doc is not None and not patient_doc.is_closed:
                        patient_doc.close()
                except Exception:
                    pass

        self.work_doc = out
        self._render_preview()
    def _save(self):
        # Default filename: date + time (24h)
        stamp = datetime.now().strftime("%d-%m-%Y_%H%M")
        default_name = f"{stamp}_DAACal-MPS.pdf"

        start_dir = os.path.dirname(self.pdf_path) if getattr(self, "pdf_path", "") else ""
        suggested_path = os.path.join(start_dir, default_name) if start_dir else default_name

        path, _ = QFileDialog.getSaveFileName(self, "Save PDF", suggested_path, "PDF (*.pdf)")
        if not path:
            return

        if not path.lower().endswith(".pdf"):
            path += ".pdf"

        # If target exists, confirm overwrite and remove it first (prevents crash)
        if os.path.exists(path):
            resp = QMessageBox.question(
                self,
                "Overwrite file?",
                f"The file already exists:\n\n{path}\n\nDo you want to replace it?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if resp != QMessageBox.Yes:
                return

            try:
                os.remove(path)
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Cannot overwrite",
                    f"Unable to remove the existing file:\n\n{path}\n\n{e}",
                )
                return

        # Capture current UI -> overrides
        self._save_ui_to_patient()

        # Ensure we have an up-to-date work_doc
        self._build_final_doc()

        # Save PDF (guard against IO / fitz errors)
        try:
            self.work_doc.save(path)
        except Exception as e:
            QMessageBox.critical(self, "Save error", f"Failed to save PDF:\n\n{e}")
            return

        # Persist settings (same as before)
        self._persist_all_settings_to_db()

        self.accept()
class PdfGraphicsView(QGraphicsView):
    """
    QGraphicsView that emits a signal with scene coordinates when clicked.
    """
    clicked = pyqtSignal(QPointF)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            scene_pos = self.mapToScene(event.pos())
            self.clicked.emit(scene_pos)
        super().mousePressEvent(event)


def span_text(span):
    """
    Helper to safely get the text content of a span from PyMuPDF's 'rawdict' output.
    Newer PyMuPDF versions put text in span['chars'] instead of span['text'].
    """
    txt = span.get("text")
    if isinstance(txt, str):
        return txt
    chars = span.get("chars")
    if isinstance(chars, list):
        return "".join(ch.get("c", "") for ch in chars)
    return ""

def replace_span_text(page, span, new_text: str):
    new_text = str(new_text)
    full_rect = fitz.Rect(span["bbox"])
    font_size = span.get("size") or 10

    # --- BLANK MODE: draw white box instead of text ---
    if not new_text.strip():
        h = full_rect.height
        w = full_rect.width

        size = span.get("size") or font_size

        if size < 8:
            # TABLE HEADERS / TIMES (small text)
            # → small, proportional patch just covering the time/text
            margin_y = h * 0.15      # a bit taller than glyphs
            margin_x = w * 0.25      # a bit wider left/right
        else:
            # PILL LABELS (big text in coloured capsules)
            # → TALL and WIDE to cover the whole green/yellow bar
            margin_y = h * 0.80      # back to tall box
            margin_x = w * 1.25      # very wide horizontally

        erase_rect = fitz.Rect(
            full_rect.x0 - margin_x,
            full_rect.y0 - margin_y,
            full_rect.x1 + margin_x,
            full_rect.y1 + margin_y,
        )
        page.draw_rect(erase_rect, fill=(1, 1, 1), color=None, overlay=True)
        return

    # --- NORMAL MODE (non-blank text): unchanged behaviour ---
    h = full_rect.height
    pad_top = h * 0.01
    pad_bottom = h * 0.07

    erase_rect = fitz.Rect(
        full_rect.x0,
        full_rect.y0 + pad_top,
        full_rect.x1,
        full_rect.y1 - pad_bottom,
    )
    page.draw_rect(erase_rect, fill=(1, 1, 1), color=None, overlay=True)

    orig_text = span.get("text", "")
    pill_center_x = (full_rect.x0 + full_rect.x1) / 2.0

    if orig_text:
        avg_char_width = full_rect.width / max(1, len(orig_text))
        new_text_width = avg_char_width * len(new_text)
    else:
        avg_char_width = font_size * 0.5
        new_text_width = avg_char_width * len(new_text)

    # HORIZONTAL: centre, then nudge slightly LEFT
    x = pill_center_x - new_text_width / 2.0
    x -= font_size * 0.20

    # VERTICAL: keep original baseline
    base_x, base_y = span["origin"]
    y = base_y

    page.insert_text(
        (x, y),
        new_text,
        fontsize=font_size,
        color=(0, 0, 0),
    )



class BlisterHeadingDialog(QDialog):
    """
    Dialog to let the user pick which headings (e.g. 08:05, 08:10, 08:15)
    to replace, and what to replace them with (e.g. B'FAST).
    """
    def __init__(self, parent, labels):
        super().__init__(parent)
        self.setWindowTitle("Blister headings")
        self.selected_labels = []
        self.replace_text = "B'FAST"

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            "Select which headings to replace,\n"
            "then choose/enter the replacement text."
        ))

        self.list = QListWidget()
        for lbl in labels:
            item = QListWidgetItem(lbl)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked)  # checked by default
            self.list.addItem(item)
        layout.addWidget(self.list)

        layout.addWidget(QLabel("Replace with:"))

        # --- DROPDOWN for B'FAST / LUNCH / DINNER / BED / Blank ---
        self.replace_combo = QComboBox()
        self.replace_combo.setEditable(True)
        self.replace_combo.addItems(["B'FAST", "LUNCH", "DINNER", "BED", "Blank"])
        self.replace_combo.setCurrentText(self.replace_text)
        layout.addWidget(self.replace_combo)

        btn_row = QHBoxLayout()
        ok_btn = QPushButton("OK")
        cancel_btn = QPushButton("Cancel")
        ok_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

    def accept(self):
        self.selected_labels = []
        for i in range(self.list.count()):
            item = self.list.item(i)
            if item.checkState() == Qt.Checked:
                self.selected_labels.append(item.text())

        txt = self.replace_combo.currentText().strip()

        # Special handling: "Blank" → empty string,
        # which triggers the BIG WHITE BOX behaviour in replace_span_text.
        if txt.upper() == "BLANK":
            txt = ""

        self.replace_text = txt
        super().accept()

    def get_selected_labels(self):
        return self.selected_labels

    def get_replace_text(self):
        return self.replace_text



def _extract_page_dates(doc):
    """
    For each page, return a list [start, end, expiry] as strings like '08 Dec 25',
    ignoring DOB (assumed to have year < 20).
    """
    page_dates = []
    date_pattern = re.compile(r"\d{1,2} \w{3} \d{2}")

    for i in range(doc.page_count):
        txt = doc.load_page(i).get_text("text")
        dates = date_pattern.findall(txt)

        # Filter out DOB-ish stuff (year < 20)
        filtered = [d for d in dates if int(d.split()[-1]) >= 20]

        # Deduplicate, keep order
        seen = set()
        unique = []
        for d in filtered:
            if d not in seen:
                seen.add(d)
                unique.append(d)

        if len(unique) < 3:
            raise ValueError(
                f"Page {i + 1} appears to have fewer than 3 date-like fields: {unique}"
            )

        page_dates.append(unique)

    return page_dates


def combine_4week_blister_to_single(input_path: str,
                                    output_path: str,
                                    times_to_change=None,
                                    new_time_label="B'FAST"):

    doc = fitz.open(input_path)
    page_dates = _extract_page_dates(doc)

    starts = [p[0] for p in page_dates]
    ends = [p[1] for p in page_dates]
    expiries = [p[2] for p in page_dates]

    def parse_date(s: str):
        return datetime.strptime(s, "%d %b %y").date()

    overall_start = min(starts, key=parse_date)
    overall_end = max(ends, key=parse_date)

    #  ---- FIXED LINE ----
    overall_expiry = min(expiries, key=parse_date)     # oldest expiry (printed on last blister)

    out = fitz.open()
    out.insert_pdf(doc, from_page=0, to_page=0)
    page = out[0]

    old_start, old_end, old_expiry = page_dates[0][:3]
    date_replacements = {
        old_start: overall_start,
        old_end: overall_end,
        old_expiry: overall_expiry,
    }




def combine_4week_blister_to_two(input_path: str,
                                 output_path: str,
                                 times_to_change=None,
                                 new_time_label="B'FAST"):
    """
    Take a 4-page weekly blister PDF and produce a 2-page PDF with 2 weeks per page.

    - Page 1 of output = combined dates for original pages 1+2 (base layout = original page 1)
    - Page 2 of output = combined dates for original pages 3+4 (base layout = original page 3)
    - Replaces any headings in `times_to_change` on each base page with `new_time_label`.
    """
    doc = fitz.open(input_path)
    page_dates = _extract_page_dates(doc)

    def parse_date(s: str):
        return datetime.strptime(s, "%d %b %y").date()

    out = fitz.open()
    halves_meta = []

    def _build_half(page_indices, base_index):
        page_indices = [i for i in page_indices if i < len(page_dates)]
        if not page_indices:
            return None

        starts = [page_dates[i][0] for i in page_indices]
        ends = [page_dates[i][1] for i in page_indices]

        overall_start = min(starts, key=parse_date)
        overall_end = max(ends, key=parse_date)
        # For this half, expiry = expiry from the LAST original page in this half
        last_idx = page_indices[-1]
        overall_expiry = page_dates[last_idx][2]

        out.insert_pdf(doc, from_page=base_index, to_page=base_index)
        page = out[-1]

        old_start, old_end, old_expiry = page_dates[base_index][:3]
        date_replacements = {
            old_start: overall_start,
            old_end: overall_end,
            old_expiry: overall_expiry,
        }

        raw = page.get_text("rawdict")
        for block in raw.get("blocks", []):
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    txt = span_text(span)
                    rep = date_replacements.get(txt)
                    if rep:
                        replace_span_text(page, span, rep)

        local_times_to_change = times_to_change
        if local_times_to_change is None:
            txt0 = page.get_text("text")
            time_pattern = re.compile(r"\d{2}:\d{2}")
            local_times_to_change = sorted(set(time_pattern.findall(txt0)))

        raw = page.get_text("rawdict")
        times_set = set(local_times_to_change or [])

        for block in raw.get("blocks", []):
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    txt = span_text(span)
                    if txt in times_set:
                        replace_span_text(page, span, new_time_label)

        return overall_start, overall_end, overall_expiry

    # First 2 weeks: pages 0+1 (base = 0)
    first_meta = _build_half([0, 1], 0)
    if first_meta:
        halves_meta.append(first_meta)

    # Second 2 weeks: pages 2+3 (base = 2) if present
    if doc.page_count >= 3:
        second_meta = _build_half([2, 3], 2)
        if second_meta:
            halves_meta.append(second_meta)

    if not halves_meta:
        raise ValueError("Unable to build any 2-week halves from the input PDF.")

    out.save(output_path)
    out.close()
    doc.close()

    return halves_meta, (times_to_change or [])


class PdfEditorDialog(QDialog):
    """
    In-app PDF viewer/editor:

    - Left: page list
    - Centre: rendered page
    - Click any visible text to edit it (span-level, same font/size)
    - Right: Save tools + 4-week blister layout (4→1 page, 4→2 pages)
    """
    def __init__(self, parent, pdf_path: str):
        super().__init__(parent)
        self.setWindowTitle(f"PDF Editor - {os.path.basename(pdf_path)}")
        self.resize(1000, 700)

        self.doc = None
        self.current_page_index = 0
        self.zoom = 2.0
        self.pdf_path = pdf_path

        layout = QHBoxLayout(self)

        # Page list
        self.page_list = QListWidget()
        self.page_list.setMaximumWidth(120)
        self.page_list.currentRowChanged.connect(self.on_page_selected)
        layout.addWidget(self.page_list)

        # Graphics view + scene
        self.scene = QGraphicsScene()
        self.view = PdfGraphicsView()
        self.view.setScene(self.scene)
        self.view.clicked.connect(self.handle_click)
        layout.addWidget(self.view, 1)

        # Right panel
        right = QVBoxLayout()


        # Save buttons
        #self.save_btn = QPushButton("Save")
        self.save_as_btn = QPushButton("Save As…")
        #self.save_btn.clicked.connect(self.save)
        self.save_as_btn.clicked.connect(self.save_as)
        #right.addWidget(self.save_btn)
        right.addWidget(self.save_as_btn)

        # 4-week blister layout box
        blister_box = QGroupBox("4-week blister")
        blister_layout = QVBoxLayout()

        self.combine_1page_btn = QPushButton("4 weeks → 1 page")
        self.combine_2page_btn = QPushButton("4 weeks → 2 pages")

        self.combine_1page_btn.clicked.connect(self.combine_to_one_page)
        self.combine_2page_btn.clicked.connect(self.combine_to_two_pages)

        blister_layout.addWidget(self.combine_1page_btn)
        blister_layout.addWidget(self.combine_2page_btn)
        blister_box.setLayout(blister_layout)
        right.addWidget(blister_box)

        right.addStretch(1)
        layout.addLayout(right)

        self.load_pdf(pdf_path)

    # ---------- load/render ----------
    def load_pdf(self, path: str):
        try:
            self.doc = fitz.open(path)
        except Exception as e:
            QMessageBox.critical(self, "Error opening PDF", str(e))
            self.reject()
            return

        self.page_list.clear()
        for i in range(self.doc.page_count):
            self.page_list.addItem(f"Page {i + 1}")
        if self.doc.page_count:
            self.page_list.setCurrentRow(0)
            self.show_page(0)

    def on_page_selected(self, index: int):
        if self.doc is None or index < 0:
            return
        self.current_page_index = index
        self.show_page(index)

    def show_page(self, index: int):
        if self.doc is None or not (0 <= index < self.doc.page_count):
            return

        page = self.doc.load_page(index)
        mat = fitz.Matrix(self.zoom, self.zoom)
        pix = page.get_pixmap(matrix=mat)

        fmt = QImage.Format_RGBA8888 if pix.alpha else QImage.Format_RGB888
        img = QImage(pix.samples, pix.width, pix.height, pix.stride, fmt)
        img = img.copy()
        pixmap = QPixmap.fromImage(img)

        self.scene.clear()
        self.scene.addPixmap(pixmap)
        self.scene.setSceneRect(0, 0, pixmap.width(), pixmap.height())

    # ---------- click-to-edit ----------
    def handle_click(self, scene_pos: QPointF):
        if self.doc is None:
            return

        page = self.doc.load_page(self.current_page_index)

        pdf_x = scene_pos.x() / self.zoom
        pdf_y = scene_pos.y() / self.zoom

        raw = page.get_text("rawdict")

        best_span = None
        best_area = None

        for block in raw.get("blocks", []):
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    txt = span_text(span).strip()
                    if not txt:
                        continue
                    x0, y0, x1, y1 = span.get("bbox", (0, 0, 0, 0))

                    margin = 1.5
                    if (x0 - margin) <= pdf_x <= (x1 + margin) and (y0 - margin) <= pdf_y <= (y1 + margin):
                        area = (x1 - x0) * (y1 - y0)
                        if best_span is None or area < best_area:
                            best_span = span
                            best_area = area

        if best_span is None:
            return

        old_text = span_text(best_span)

        new_text, ok = QInputDialog.getText(
            self,
            "Edit text",
            "New text (leave blank to erase):",
            text=old_text
        )
        if not ok:
            return

        replace_span_text(page, best_span, new_text)
        self.show_page(self.current_page_index)

    # ---------- save ----------
    def save(self):
        if not self.doc or not self.pdf_path:
            return
        try:
            self.doc.save(self.pdf_path, incremental=False)
            QMessageBox.information(self, "Saved", f"Saved to:\n{self.pdf_path}")
        except Exception as e:
            QMessageBox.critical(self, "Save error", str(e))

    def save_as(self):
        if not self.doc:
            return
        out_path, _ = QFileDialog.getSaveFileName(
            self, "Save PDF As", self.pdf_path, "PDF files (*.pdf)"
        )
        if not out_path:
            return
        if not out_path.lower().endswith(".pdf"):
            out_path += ".pdf"
        try:
            self.doc.save(out_path, incremental=False)
            QMessageBox.information(self, "Saved", f"Saved to:\n{out_path}")
        except Exception as e:
            QMessageBox.critical(self, "Save error", str(e))

    def closeEvent(self, event):
        if self.doc is not None:
            self.doc.close()
        super().closeEvent(event)

    # ---------- blister helpers inside viewer ----------
    def _pick_headings_for_blister(self):
        """
        Inspect page 1 of the currently loaded doc and let the user choose which
        HH:MM headings to convert (e.g. 08:05 / 08:10 / 08:15 → B'FAST).
        """
        try:
            if self.doc is None or self.doc.page_count == 0:
                return None, None
            first_page = self.doc.load_page(0)
            first_text = first_page.get_text("text")
        except Exception:
            return None, None

        times = sorted(set(re.findall(r"\d{2}:\d{2}", first_text)))
        if not times:
            QMessageBox.information(
                self,
                "No headings found",
                "No time-like headings (HH:MM) were detected on the first page."
            )
            return None, None

        dlg = BlisterHeadingDialog(self, times)
        if dlg.exec_() != QDialog.Accepted:
            return None, None

        selected = dlg.get_selected_labels()
        new_label = dlg.get_replace_text()
        if not selected or not new_label:
            QMessageBox.information(
                self,
                "Nothing to do",
                "Please select at least one heading and enter replacement text."
            )
            return None, None
        return selected, new_label

    def combine_to_one_page(self):
        """4 weeks / 4 pages → 4 weeks / 1 page (summary card), then
        immediately reload that new PDF in the viewer so you see the result."""
        if not self.pdf_path:
            return

        selected, new_label = self._pick_headings_for_blister()
        if not selected:
            return

        base, ext = os.path.splitext(self.pdf_path)
        default_out = base + "_1page.pdf"
        out_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save 4-week 1-page PDF as",
            default_out,
            "PDF files (*.pdf)"
        )
        if not out_path:
            return
        if not out_path.lower().endswith(".pdf"):
            out_path += ".pdf"

        try:
            overall_start, overall_end, overall_expiry, used = combine_4week_blister_to_single(
                self.pdf_path,
                out_path,
                times_to_change=selected,
                new_time_label=new_label,
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error creating combined blister",
                f"An error occurred while processing the PDF:\n{e}",
            )
            return

        # Reload the new PDF straight into the viewer
        self.pdf_path = out_path
        self.load_pdf(out_path)

        QMessageBox.information(
            self,
            "Combined blister created",
            (
                f"Created:\n{out_path}\n\n"
                f"Range: {overall_start} → {overall_end}\n"
                f"Expiry: {overall_expiry}\n"
                f"Replaced headings: {', '.join(used)}"
            ),
        )

    def combine_to_two_pages(self):
        """4 weeks / 4 pages → 4 weeks / 2 pages (2 weeks / blister),
        and then reload that result in the viewer."""
        if not self.pdf_path:
            return

        selected, new_label = self._pick_headings_for_blister()
        if not selected:
            return

        base, ext = os.path.splitext(self.pdf_path)
        default_out = base + "_2pages.pdf"
        out_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save 4-week 2-page PDF as",
            default_out,
            "PDF files (*.pdf)"
        )
        if not out_path:
            return
        if not out_path.lower().endswith(".pdf"):
            out_path += ".pdf"

        try:
            overall_start, overall_end, overall_expiry, used = combine_4week_blister_to_two(
                self.pdf_path,
                out_path,
                times_to_change=selected,
                new_time_label=new_label,
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error creating combined blister",
                f"An error occurred while processing the PDF:\n{e}",
            )
            return

        # Reload the new PDF in the viewer
        self.pdf_path = out_path
        self.load_pdf(out_path)

        QMessageBox.information(
            self,
            "Combined blister created",
            (
                f"Created:\n{out_path}\n\n"
                f"Range: {overall_start} → {overall_end}\n"
                f"Expiry: {overall_expiry}\n"
                f"Replaced headings: {', '.join(used)}"
            ),
        )



def mondays_in_month(year: int, month: int):
    """Return list[date] of all Mondays in the given month/year."""
    cal = calendar.Calendar(firstweekday=calendar.MONDAY)
    return [
        d for d in cal.itermonthdates(year, month)
        if d.month == month and d.weekday() == 0
    ]

SINGLE_INSTANCE_KEY = "DAACAL_SINGLE_INSTANCE"

def ensure_single_instance(server_name=SINGLE_INSTANCE_KEY):
    """
    Returns a QLocalServer if this is the first instance.
    Returns None if another instance of DAACal is already running.
    """
    # Try to connect to an existing instance
    socket = QLocalSocket()
    socket.connectToServer(server_name)
    if socket.waitForConnected(100):
        # Another instance is already running
        socket.disconnectFromServer()
        return None

    # No existing instance: create a server for this one
    QLocalServer.removeServer(server_name)  # clean up stale server from crash, if any
    server = QLocalServer()
    if not server.listen(server_name):
        # Could not listen for some reason -> treat as "already running"
        return None

    return server

def send_ipc_command(command: str, server_name: str = SINGLE_INSTANCE_KEY, timeout_ms: int = 500) -> bool:
    """
    Sends a short command to the already-running instance over QLocalSocket.
    Returns True if sent, False if couldn't connect.
    """
    sock = QLocalSocket()
    sock.connectToServer(server_name)
    if not sock.waitForConnected(timeout_ms):
        return False

    payload = (command.strip() + "\n").encode("utf-8")
    sock.write(payload)
    sock.flush()
    sock.waitForBytesWritten(timeout_ms)
    sock.disconnectFromServer()
    return True


def attach_ipc_listener(server: QLocalServer, app: QApplication) -> None:
    """
    Makes THIS instance respond to IPC commands from later-launched instances.
    Supported:
      - QUIT: gracefully exits this instance.
    """
    def on_new_connection():
        while server.hasPendingConnections():
            conn = server.nextPendingConnection()
            if conn is None:
                return

            # Read whatever was sent (small, one-shot)
            conn.waitForReadyRead(200)
            raw = bytes(conn.readAll()).decode("utf-8", errors="ignore").strip().upper()

            conn.disconnectFromServer()
            conn.deleteLater()

            if raw == "QUIT":
                app.quit()

    server.newConnection.connect(on_new_connection)

class ClaimsWindow(QDialog):
    """
    Lists patients with partial_supply='Y'.
    - Master checkbox per patient = exclude patient
    - When excluded, show per-week checkboxes to optionally exclude specific weeks instead.
    - Month/Year picker auto-generates Monday start dates.
    - Export to Downloads as 'DAA Claims [Month YYYY].xlsx'
    """
    def __init__(self, parent, db_cursor):
        super().__init__(parent)
        self.setWindowTitle("DAA Claims")
        self.setMinimumWidth(720)
        self.cur = db_cursor

        outer = QVBoxLayout(self)

        # --- Month/Year selector ---
        top = QHBoxLayout()
        top.addWidget(QLabel("Claim month:"))
        self.month_combo = QComboBox()
        self.year_combo = QComboBox()

        months = [calendar.month_name[m] for m in range(1, 13)]
        for m in months: self.month_combo.addItem(m)
        current = date.today()
        years = list(range(current.year - 1, current.year + 3))
        for y in years: self.year_combo.addItem(str(y))

        # default to current month/year
        self.month_combo.setCurrentIndex(current.month - 1)
        self.year_combo.setCurrentText(str(current.year))

        self.month_combo.currentIndexChanged.connect(self._rebuild_week_columns)
        self.year_combo.currentIndexChanged.connect(self._rebuild_week_columns)

        top.addWidget(self.month_combo)
        top.addWidget(self.year_combo)
        top.addStretch(1)
        outer.addLayout(top)

        # --- Scrollable patient list ---
        self.container = QWidget()
        self.grid = QGridLayout(self.container)
        # --- table header (single grid used by header + rows) ---
        # columns: [0]=Exclude (fixed), [1]=VLine (1px), [2]=Name (stretch)
        self.header_exclude = QLabel("<b>Exclude</b>")
        self.header_name = QLabel("<b>Name</b>")

        # compute exclude column width from header text
        exclude_w = self.header_exclude.sizeHint().width() + 18  # padding for checkbox area
        self.grid.addWidget(self.header_exclude, 0, 0, alignment=Qt.AlignCenter)

        header_div = QFrame()
        header_div.setFrameShape(QFrame.VLine)
        header_div.setLineWidth(1)
        header_div.setStyleSheet("color:#b0b0b0;")
        self.grid.addWidget(header_div, 0, 1)

        self.grid.addWidget(self.header_name, 0, 2, alignment=Qt.AlignLeft | Qt.AlignVCenter)

        # enforce column metrics so rows line up perfectly (QGridLayout doesn’t have setColumnMaximumWidth)
        self.grid.setColumnMinimumWidth(0, exclude_w)
        self.grid.setColumnStretch(0, 0)  # fixed-like column (Exclude)

        self.grid.setColumnMinimumWidth(1, 1)
        self.grid.setColumnStretch(1, 0)  # divider column stays narrow

        self.grid.setColumnStretch(2, 1)  # Name column fills remaining width

        # header bottom line
        hdr_line = QFrame()
        hdr_line.setFrameShape(QFrame.HLine)
        hdr_line.setStyleSheet("color:#b0b0b0;")
        self.grid.addWidget(hdr_line, 1, 0, 1, 3)

        # tidy spacing; white like calendar
        self.container.setStyleSheet("background:white; color:black; font-size:10pt;")
        self.grid.setHorizontalSpacing(0)
        self.grid.setVerticalSpacing(0)

        self.rows = []  # list of dicts per patient row

        self._load_patients()

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self.container)
        outer.addWidget(scroll)

        # --- Buttons ---
        btns = QHBoxLayout()
        btns.addStretch(1)
        self.export_btn = QPushButton("Export XLSX")
        self.export_btn.clicked.connect(self._export_xlsx)
        btns.addWidget(self.export_btn)
        outer.addLayout(btns)

        self._rebuild_week_columns()

    def _selected_year_month(self):
        y = int(self.year_combo.currentText())
        m = self.month_combo.currentIndex() + 1
        return y, m

    def _load_patients(self):
        """
        Same grid for header + rows:
          col0 = Exclude (fixed), col1 = VLine (1px), col2 = Name (stretch)
        Checkbox is top-center in its cell. Per-week panel expands under the name.
        """
        self.cur.execute("""
                         SELECT name, COALESCE(medicare, ''), paused, flagged
                         FROM patients
                         WHERE UPPER(TRIM(partial_supply)) = 'Y'
                           AND (
                             ceased IS NULL
                                 OR ceased = 0
                                 OR TRIM(ceased) = ''
                                 OR LOWER(ceased) IN ('n', 'no', 'false')
                             )
                         ORDER BY name COLLATE NOCASE
                         """)
        results = self.cur.fetchall()

        # clear all rows under the header area (keep rows 0-1 which are header + underline)
        # grid rows now start at 2 for data
        # remove everything with row >= 2
        to_delete = []
        for i in range(self.grid.count()):
            item = self.grid.itemAt(i)
            r, c, rs, cs = self.grid.getItemPosition(i)
            if r >= 2:
                w = item.widget()
                if w: to_delete.append(w)
        for w in to_delete:
            w.setParent(None)
            w.deleteLater()

        self.rows.clear()

        # start placing rows at grid row index 2
        r = 2
        for (name, medicare, paused, flagged) in results:
            # --- col 0: Exclude checkbox (top-center)
            # --- col 0: Exclude checkbox with stable vertical position
            # --- col 0: Exclude checkbox perfectly centered (stays steady when expanded) ---
            exclude_cb = QCheckBox()

            cb_wrapper = QWidget()
            cb_layout = QVBoxLayout(cb_wrapper)
            cb_layout.setContentsMargins(0, 4, 0, 4)  # equal padding top/bottom for vertical centering
            cb_layout.setSpacing(0)

            # balanced layout: stretch – checkbox – stretch keeps it dead center
            cb_layout.addStretch(1)
            cb_layout.addWidget(exclude_cb, alignment=Qt.AlignHCenter)
            cb_layout.addStretch(1)

            # when expanded, the checkbox stays visually stable
            self.grid.addWidget(cb_wrapper, r, 0, alignment=Qt.AlignHCenter | Qt.AlignTop)

            # --- col 1: vertical divider (1px)
            divider = QFrame()
            divider.setFrameShape(QFrame.VLine)
            divider.setLineWidth(1)
            divider.setStyleSheet("color:#e0e0e0;")
            self.grid.addWidget(divider, r, 1)

            # --- col 2: Name cell with expandable per-week panel (stacked)
            display_name = name.replace(",", "").strip()
            name_lbl = QLabel(display_name)
            name_lbl.setWordWrap(True)

            # highlight like calendar
            if paused:
                name_lbl.setStyleSheet("background-color:#FF0000; color:white; padding:3px;")
                name_lbl.setToolTip("Paused")
            elif flagged:
                name_lbl.setStyleSheet("background-color:#FFC0CB; color:black; padding:3px;")
                name_lbl.setToolTip("Flagged")
            else:
                name_lbl.setStyleSheet("background-color:white; color:black; padding:3px;")

            week_box = QWidget()
            week_box.setVisible(False)
            week_layout = QVBoxLayout(week_box)
            week_layout.setContentsMargins(24, 2, 6, 6)  # indent under name
            week_layout.setSpacing(2)

            name_container = QWidget()
            nlay = QVBoxLayout(name_container)
            nlay.setContentsMargins(0, 0, 0, 0)
            nlay.setSpacing(0)
            nlay.addWidget(name_lbl)
            nlay.addWidget(week_box)

            self.grid.addWidget(name_container, r, 2)

            # horizontal line under the whole row (spans all 3 columns)
            row_line = QFrame()
            row_line.setFrameShape(QFrame.HLine)
            row_line.setStyleSheet("color:#e0e0e0;")
            self.grid.addWidget(row_line, r + 1, 0, 1, 3)

            # record row refs
            self.rows.append({
                "name": display_name,
                "medicare": (medicare or "").replace(" ", ""),
                "exclude_cb": exclude_cb,
                "week_box": week_box,
                "week_layout": week_layout,
                "week_cbs": []
            })

            # keep checkbox pinned top when week_box grows
            exclude_cb.stateChanged.connect(
                lambda state, rb=self.rows[-1]: rb["week_box"].setVisible(state == Qt.Checked))

            # next data row starts two rows down (we inserted an HLine at r+1)
            r += 2

    def _rebuild_week_columns(self):
        """Build per-week checkboxes (one per Monday in selected month) for each row."""
        y, m = self._selected_year_month()
        mondays = mondays_in_month(y, m)

        for row in self.rows:
            # clear old
            for cb in row["week_cbs"]:
                cb.deleteLater()
            row["week_cbs"].clear()

            # new cbs (default all checked -> if excluding the patient, we’ll exclude all weeks unless the user unticks some)
            for d in mondays:
                cb = QCheckBox(d.strftime("Exclude %d/%m/%Y"))
                cb.setChecked(True)
                row["week_layout"].addWidget(cb)
                row["week_cbs"].append(cb)

            row["week_box"].setVisible(row["exclude_cb"].isChecked())

    def _export_xlsx(self):
        from PyQt5.QtWidgets import QFileDialog
        import calendar, os
        from openpyxl import load_workbook

        # --- derive date info ---
        y, m = self._selected_year_month()
        month_name = calendar.month_name[m]
        mondays = mondays_in_month(y, m)

        # --- locate template file ---
        template_path = resource_path("DAA Claiming Spreadsheet.xlsx")

        if not os.path.exists(template_path):
            return  # silent fail if missing

        # --- load workbook from template ---
        wb = load_workbook(template_path)
        ws = wb.active

        # --- clear existing data rows below row 4 ---
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        # --- populate spreadsheet, starting at row 4 ---
        out_row = 4
        for row in self.rows:
            name = row["name"].strip()
            med = (row["medicare"] or "").replace(" ", "")

            if not row["exclude_cb"].isChecked():
                for d in mondays:
                    ws.cell(out_row, 1).value = name
                    ws.cell(out_row, 2).value = med
                    ws.cell(out_row, 3).value = d.strftime("%d/%m/%Y")
                    out_row += 1
            else:
                for d, cb in zip(mondays, row["week_cbs"]):
                    if cb.isChecked():
                        continue  # skip excluded week
                    ws.cell(out_row, 1).value = name
                    ws.cell(out_row, 2).value = med
                    ws.cell(out_row, 3).value = d.strftime("%d/%m/%Y")
                    out_row += 1

        # --- native Save As dialog (no popups after save) ---
        suggested_name = f"DAA Claims {month_name} {y}.xlsx"
        default_dir = os.path.expanduser("~/Downloads")
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save DAA Claim Spreadsheet",
            os.path.join(default_dir, suggested_name),
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        # --- save silently ---
        try:
            wb.save(file_path)
        except Exception:
            pass


class DBChangeHandler(FileSystemEventHandler):
    def __init__(self, parent):
        super().__init__()
        self.parent = parent

    def on_modified(self, event):
        if event.src_path.endswith("webster_calendar.db"):
            print("DEBUG: DB file changed, scheduling reload")
            QMetaObject.invokeMethod(
                self.parent,
                "emit_db_changed_signal",
                Qt.QueuedConnection
            )



class EditDrugsDialog(QDialog):
    drugs_updated = pyqtSignal()
    def __init__(self, patient_id, med_names, conn, parent=None):
        super().__init__(parent)
        self.setWindowModality(Qt.WindowModal)
        self.setWindowFlag(Qt.WindowStaysOnTopHint, True)
        self.setWindowTitle("Edit Drugs")
        self.patient_id = patient_id
        self.med_names = med_names
        self.conn = conn
        self.checkboxes = []
        layout = QVBoxLayout()
        self.drug_entries = []  # List to store (checkbox, lineedit, original_name, excel_row)

        # Build initial drug list
        cur = self.conn.cursor()
        for idx, drug in enumerate(med_names):
            excel_row = idx + 4  # ALWAYS tied to sheet position

            if drug is None or str(drug).strip() == "":
                continue  # skip UI row but preserve mapping for non-blank entries

            drug = str(drug).strip()

            cur.execute(
                "SELECT active FROM patient_drug_status WHERE patient_id=? AND drug_name=?",
                (patient_id, drug),
            )
            row = cur.fetchone()
            active_flag = 1 if row is None else row[0]

            hbox = QHBoxLayout()
            checkbox = QCheckBox()
            checkbox.setChecked(bool(active_flag))
            lineedit = QLineEdit(drug)

            hbox.addWidget(checkbox)
            hbox.addWidget(lineedit)
            layout.addLayout(hbox)

            self.drug_entries.append((checkbox, lineedit, drug, excel_row))

        # Add Drug button
        add_drug_btn = QPushButton("Add Drug")
        def handle_add_drug():
            new_drug, ok = QInputDialog.getText(self, "Add Drug", "Enter new drug name:")
            if not ok or not new_drug.strip():
                return
            new_drug = new_drug.strip()
            # Avoid duplicate entry in the dialog
            if any(lineedit.text().strip().lower() == new_drug.lower() for _, lineedit, _, _ in self.drug_entries):
                QMessageBox.warning(self, "Duplicate", f"{new_drug} is already listed.")
                return
            checkbox = QCheckBox()
            checkbox.setChecked(True)
            lineedit = QLineEdit(new_drug)
            hbox = QHBoxLayout()
            hbox.addWidget(checkbox)
            hbox.addWidget(lineedit)
            layout.insertLayout(len(self.drug_entries), hbox)
            # Next available Excel row (4 + count)
            next_row = 4 + len(self.drug_entries)
            self.drug_entries.append((checkbox, lineedit, new_drug, next_row))
        add_drug_btn.clicked.connect(handle_add_drug)
        layout.addWidget(add_drug_btn)

        # Save/Cancel row
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Save")
        save_btn.clicked.connect(self.save_and_accept)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

    def save_and_accept(self):
        if self.save_changes():
            self.drugs_updated.emit()
            self.accept()

    def save_changes(self):
        try:
            excel_path = self.excel_path  # must be set externally
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            cur = self.conn.cursor()

            # Write drug names back to their original Excel rows
            used_rows = set()

            for checkbox, lineedit, original_name, excel_row in self.drug_entries:
                new_name = (lineedit.text() or "").strip()
                is_active = int(checkbox.isChecked())

                if not isinstance(excel_row, int):
                    raise ValueError(f"Invalid excel_row for drug '{original_name}': {excel_row}")

                used_rows.add(excel_row)

                # If blank, clear the cell and remove status row
                if not new_name:
                    ws.cell(row=excel_row, column=1).value = None
                    if original_name:
                        cur.execute(
                            "DELETE FROM patient_drug_status WHERE patient_id=? AND drug_name=?",
                            (self.patient_id, original_name),
                        )
                    continue

                # Write to Excel at the correct row
                ws.cell(row=excel_row, column=1).value = new_name

                # If renamed, update DB row key
                if original_name and new_name != original_name:
                    cur.execute(
                        "UPDATE patient_drug_status SET drug_name=? WHERE patient_id=? AND drug_name=?",
                        (new_name, self.patient_id, original_name),
                    )

                # Upsert active status for the (possibly renamed) drug
                cur.execute(
                    "INSERT INTO patient_drug_status (patient_id, drug_name, active) VALUES (?, ?, ?) "
                    "ON CONFLICT(patient_id, drug_name) DO UPDATE SET active=excluded.active",
                    (self.patient_id, new_name, is_active),
                )

            # Clear any leftover drug rows below the list (keep your old behavior: clear rows 4..23 that aren't used)
            for r in range(4, 24):
                if r not in used_rows:
                    ws.cell(row=r, column=1).value = ws.cell(row=r, column=1).value  # leave as-is (no wipe)

            self.conn.commit()
            wb.save(excel_path)
            return True

        except Exception as e:
            QMessageBox.critical(self, "Save Drugs Error", str(e))
            return False

class OneBackspaceClearsLineEdit(QLineEdit):
    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Backspace:
            self.clear()
        else:
            super().keyPressEvent(event)

class HighlightDelegate(QStyledItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.highlighted_row = -1

    def paint(self, painter, option, index):
        if index.row() == self.highlighted_row:
            painter.save()
            painter.fillRect(option.rect, QColor(240, 240, 240))
            painter.restore()
        super().paint(painter, option, index)

class HighlightTableWidget(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.highlight_overlay = QFrame(self.viewport())
        self.highlight_overlay.setStyleSheet("background-color: rgb(240, 240, 240);")
        self.highlight_overlay.lower()
        self.highlight_overlay.hide()
        self.highlighted_row = -1
        self.cellClicked.connect(self.update_highlight_overlay)

    def update_highlight_overlay(self, row, column):
        self.highlighted_row = row
        self.reposition_overlay()
        self.highlight_overlay.show()

    def reposition_overlay(self):
        if self.highlighted_row < 0 or self.columnCount() == 0:
            self.highlight_overlay.hide()
            return
        top = self.rowViewportPosition(self.highlighted_row)
        left = self.columnViewportPosition(0)
        height = self.rowHeight(self.highlighted_row)
        width = sum(self.columnWidth(col) for col in range(self.columnCount()))
        self.highlight_overlay.setGeometry(left, top, width, height)

class CustomTabBar(QTabBar):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setTabsClosable(True)
    def tabSizeHint(self, index):
        size = super().tabSizeHint(index)
        size.setWidth(size.width() + 30)
        return size
        size = super().tabSizeHint(index)
        size.setWidth(size.width() + 30)
        return size

def parse_batch_expiry(batch_exp):
    batch = ""
    expiry = ""
    if isinstance(batch_exp, str):
        cleaned = batch_exp.upper().replace("(", "").replace(")", "").replace("EXP", "E").replace(":", "")
        match = re.search(r"B[     ]*([A-Z0-9]+)[     ,;]*E[     ]*([0-9]{1,2}/[0-9]{2,4})", cleaned)
        if match:
            batch = match.group(1)
            expiry = match.group(2)
        else:
            # Also handle "ABC123, 11/26" pattern
            simple = re.match(r"([A-Z0-9]+)[     ,]+([0-9]{1,2}/[0-9]{2,4})", cleaned)
            if simple:
                batch = simple.group(1)
                expiry = simple.group(2)
            else:
                parts = re.split(r'[     ,;]+', cleaned)
                b_idx = [i for i, p in enumerate(parts) if p.startswith("B")]
                e_idx = [i for i, p in enumerate(parts) if p.startswith("E")]
                if b_idx and e_idx:
                    try:
                        batch = parts[b_idx[0]][1:]
                        expiry = parts[e_idx[0]][1:]
                    except:
                        pass
    return batch, expiry

USER_ID_MAP = {
  "rc": "Ryan Chetty"
}

def sync_user_map_to_db(conn):
    """Ensure the DB user list matches USER_ID_MAP and vice versa."""
    cur = conn.cursor()
    # Ensure users table exists
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id TEXT PRIMARY KEY,
            name TEXT NOT NULL
        )
    """)
    conn.commit()

    # Load DB users into dict
    cur.execute("SELECT user_id, name FROM users")
    db_users = {uid.lower(): name for uid, name in cur.fetchall()}

    # --- Step 1: Add missing users from map to DB
    for uid, name in USER_ID_MAP.items():
        if uid.lower() not in db_users:
            print(f"SYNC: Adding missing user {uid} → {name}")
            cur.execute("INSERT INTO users (user_id, name) VALUES (?, ?)", (uid.lower(), name))

    # --- Step 2: Add missing DB users back into USER_ID_MAP (dynamic sync)
    for uid, name in db_users.items():
        if uid.lower() not in USER_ID_MAP:
            print(f"SYNC: Adding {uid} → {name} to USER_ID_MAP (from DB)")
            USER_ID_MAP[uid.lower()] = name

    conn.commit()

# === Global Event Filter to Reset Inactivity Timer ===
class GlobalActivityFilter(QObject):
    def __init__(self, main_app):
        super().__init__()
        self.main_app = main_app

    def eventFilter(self, obj, event):
        if event.type() in (QEvent.KeyPress, QEvent.MouseMove, QEvent.MouseButtonPress):
            if hasattr(self.main_app, 'inactivity_timer'):
                self.main_app.inactivity_timer.start()
        return False

# === Per-user settings (saved locally on each PC/user) ===

_DAACAL_SETTINGS_PATH = os.path.join(os.path.expanduser("~"), ".daacal_settings.json")


def _load_daacal_settings() -> dict:
    if not os.path.exists(_DAACAL_SETTINGS_PATH):
        return {}
    try:
        with open(_DAACAL_SETTINGS_PATH, "r", encoding="utf-8") as f:
            return json.load(f) or {}
    except Exception:
        return {}


def _save_daacal_settings(update: dict) -> None:
    settings = _load_daacal_settings()
    settings.update(update or {})
    with open(_DAACAL_SETTINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(settings, f, indent=4)
# === Dynamic base directory resolution with settings override + fallback prompt ===

def resolve_data_directory():
    settings = _load_daacal_settings()
    dir_candidate = (settings.get("data_directory") or "").strip()

    if os.path.isdir(dir_candidate) and os.path.exists(os.path.join(dir_candidate, "webster_calendar.db")):
        return dir_candidate

    # Prompt user if DB file missing
    app = QApplication.instance() or QApplication([])
    while True:
        dir_selected = QFileDialog.getExistingDirectory(None, "Select Data Directory")
        if not dir_selected:
            QMessageBox.critical(None, "Missing DB", "No valid directory selected. DAACal will now exit.")
            sys.exit(1)

        if os.path.exists(os.path.join(dir_selected, "webster_calendar.db")):
            _save_daacal_settings({"data_directory": dir_selected})
            return dir_selected

        QMessageBox.warning(None, "Invalid Directory", "Selected folder does not contain webster_calendar.db")

BASE_DIR = resolve_data_directory()
DB_FILE = os.path.join(BASE_DIR, 'webster_calendar.db')
COLLECTION_LOGS_DIR = os.path.join(BASE_DIR, 'collection logs')
# ===============================================================

class DateItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            self_date = datetime.strptime(self.text().strip(), "%d/%m/%Y")
        except Exception:
            return False
        try:
            other_date = datetime.strptime(other.text().strip(), "%d/%m/%Y")
        except Exception:
            return True
        return self_date < other_date

class NumericItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            self_val = float(self.text())
        except ValueError:
            return False
        try:
            other_val = float(other.text())
        except ValueError:
            return True
        return self_val < other_val

class NumericItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            self_val = int(self.text()) if self.text().strip() else 999
            other_val = int(other.text()) if other.text().strip() else 999
            return self_val < other_val
        except ValueError:
            return self.text() < other.text()

class DateItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            self_date = datetime.strptime(self.text().strip(), "%d/%m/%Y")
        except Exception:
            return False  # Treat self as 'greater' (bottom) when descending

        try:
            other_date = datetime.strptime(other.text().strip(), "%d/%m/%Y")
        except Exception:
            return True  # Treat other as 'greater'

        return self_date > other_date  # Descending

class LandscapePrintPreview(QPrintPreviewDialog):
    """QPrintPreviewDialog that always re-applies landscape orientation before printing."""
    def __init__(self, printer, parent=None):
        super().__init__(printer, parent)
        self.printer = printer
        # Ensure landscape before actual printing
        for btn in self.findChildren(QPushButton):
            btn.clicked.connect(self._force_landscape)

    def _force_landscape(self):
        self.printer.setOrientation(QPrinter.Landscape)
        self.printer.setPageLayout(
            QPageLayout(
                QPageSize(QPageSize.A4),
                QPageLayout.Landscape,
                QMarginsF(0, 0, 0, 0)
            )
        )

class PrintPreviewDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Print Preview")
        self.parent = parent

        # --- Capture creation date and week start ---
        self.date_printed_str = datetime.today().strftime("%d/%m/%Y")
        today = datetime.today()
        self.this_week_start = (today - timedelta(days=today.weekday())).strftime("%d/%m/%Y")
        self.next_week_start = (today - timedelta(days=today.weekday()) + timedelta(days=7)).strftime("%d/%m/%Y")

        # === Layout ===
        main_layout = QVBoxLayout()


        self.week_label = QLabel(f"Week Starting: {self.this_week_start}")

        # --- Week selection controls ---
        week_layout = QHBoxLayout()
        week_layout.addWidget(QLabel("Select week to preview:"))
        self.this_week_radio = QRadioButton("This Week")
        self.next_week_radio = QRadioButton("Next Week")
        self.overdue_checkbox = QCheckBox("Overdue")
        self.this_week_radio.setChecked(True)
        week_layout.addWidget(self.this_week_radio)
        week_layout.addWidget(self.next_week_radio)
        week_layout.addWidget(self.overdue_checkbox)
        main_layout.addLayout(week_layout)

        # --- Main table label ---
        self.main_label = QLabel()
        font = self.main_label.font()
        font.setBold(True)
        self.main_label.setFont(font)
        main_layout.addWidget(self.main_label)

        # --- Main table ---
        self.preview_table = QTableWidget()
        self.setup_table_headers(self.preview_table)
        main_layout.addWidget(self.preview_table)

        # --- Overdue label + table ---
        self.overdue_label = QLabel("Overdue")
        self.overdue_label.setFont(font)
        self.overdue_label.hide()
        main_layout.addWidget(self.overdue_label)

        self.overdue_table = QTableWidget()
        self.setup_table_headers(self.overdue_table)
        self.overdue_table.hide()
        main_layout.addWidget(self.overdue_table)

        # --- Buttons ---
        btn_layout = QHBoxLayout()
        print_btn = QPushButton("Print")
        cancel_btn = QPushButton("Close")
        btn_layout.addWidget(print_btn)
        btn_layout.addWidget(cancel_btn)
        main_layout.addLayout(btn_layout)

        self.setLayout(main_layout)

        # --- Signals ---
        self.this_week_radio.toggled.connect(self.update_preview)
        self.next_week_radio.toggled.connect(self.update_preview)
        self.overdue_checkbox.toggled.connect(self.update_preview)
        print_btn.clicked.connect(self.do_print)
        cancel_btn.clicked.connect(self.reject)

        self.update_preview()

    def setup_table_headers(self, table):
        table.setColumnCount(9)
        table.setHorizontalHeaderLabels([
            "", "Name", "Due Date", "Weeks/Blister", "Tray Size", "Packed", "Logged", "Checked", "Notes"
        ])
        table.verticalHeader().setVisible(False)
        table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)

    def populate_table(self, table, rows):
        table.setRowCount(0)
        for data in rows:
            r = table.rowCount()
            table.insertRow(r)
            for col, (text, align_left) in enumerate(data):
                item = QTableWidgetItem(text)
                if align_left:
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                else:
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignHCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                item.setToolTip(text)
                table.setItem(r, col, item)
        # Resize columns
        header = table.horizontalHeader()
        # Apply fixed widths for preview mode
        self.apply_fixed_column_widths_preview(table)

    def apply_fixed_column_widths_preview(self, table):
        """Sets fixed widths in preview mode based on table viewport width."""
        total_width_px = table.viewport().width()
        col_percentages = [0.05, 0.20, 0.10, 0.10, 0.10, 0.07, 0.07, 0.07, 0.24]  # sum = 1.0
        header = table.horizontalHeader()
        for col, pct in enumerate(col_percentages):
            header.setSectionResizeMode(col, QHeaderView.Fixed)
            table.setColumnWidth(col, int(total_width_px * pct))

    def apply_fixed_column_widths_print(self, table, printer):
        printable_width_px = printer.pageRect(QPrinter.DevicePixel).width()
        left_margin_px = printer.pageLayout().margins(QPageLayout.Point).left() * printer.logicalDpiX() / 72.0
        right_margin_px = printer.pageLayout().margins(QPageLayout.Point).right() * printer.logicalDpiX() / 72.0
        usable_width = printable_width_px - left_margin_px - right_margin_px

        col_percentages = [0.05, 0.20, 0.10, 0.10, 0.10, 0.07, 0.07, 0.07, 0.24]
        header = table.horizontalHeader()
        for col, pct in enumerate(col_percentages):
            header.setSectionResizeMode(col, QHeaderView.Fixed)
            table.setColumnWidth(col, int(usable_width * pct))

    def update_preview(self):
        # Determine week range
        today = datetime.today()
        start_of_week = today - timedelta(days=today.weekday())
        if self.next_week_radio.isChecked():
            start_of_week += timedelta(days=7)
            self.week_label.setText(f"Week Starting: {self.next_week_start}")
        else:
            self.week_label.setText(f"Week Starting: {self.this_week_start}")
        end_of_week = start_of_week + timedelta(days=6)

        self.main_label.setText("DAA Packing Order")

        # Build main rows
        main_rows = []
        included_names = set()
        for row in range(self.parent.table.rowCount()):
            due_item = self.parent.table.item(row, 6)
            if not due_item or not due_item.text().strip():
                continue

            # --- NEW: Skip paused or flagged ---
            name_item = self.parent.table.item(row, 3)
            patient_name = name_item.text() if name_item else ""
            self.parent.cur.execute(
                "SELECT flagged, paused FROM patients WHERE name = ?",
                (patient_name,)
            )
            db_flags = self.parent.cur.fetchone()
            if db_flags and (db_flags[0] or db_flags[1]):
                continue

            try:
                due_date = datetime.strptime(due_item.text().strip(), "%d/%m/%Y").date()
                if start_of_week.date() <= due_date <= end_of_week.date():
                    charge_item = self.parent.table.item(row, 2)
                    notes_item = self.parent.table.item(row, 12)
                    charge_flag = "$" if charge_item and charge_item.text().strip().upper() == "$" else ""
                    notes_text = notes_item.text() if notes_item else ""

                    self.parent.cur.execute(
                        "SELECT weeks_per_blister, pack_size FROM patients WHERE UPPER(name) = ?",
                        (patient_name.upper(),)
                    )
                    db_row = self.parent.cur.fetchone()
                    weeks_per_blister = db_row[0] if db_row and db_row[0] else ""
                    tray_size = db_row[1] if db_row and db_row[1] else ""

                    main_rows.append([
                        (charge_flag, False),
                        (patient_name, True),
                        (due_item.text().strip(), False),
                        (str(weeks_per_blister), False),
                        (str(tray_size), False),
                        ("", False),
                        ("", False),
                        ("", False),
                        (notes_text, True)
                    ])
                    included_names.add(patient_name)
            except:
                pass
        self.populate_table(self.preview_table, main_rows)

        # Build overdue rows
        if self.overdue_checkbox.isChecked():
            overdue_rows = []
            for row in range(self.parent.table.rowCount()):
                days_item = self.parent.table.item(row, 7)
                if not days_item or not days_item.text().strip():
                    continue

                name_item = self.parent.table.item(row, 3)
                patient_name = name_item.text() if name_item else ""

                # --- NEW: Skip paused or flagged ---
                self.parent.cur.execute(
                    "SELECT flagged, paused FROM patients WHERE name = ?",
                    (patient_name,)
                )
                db_flags = self.parent.cur.fetchone()
                if db_flags and (db_flags[0] or db_flags[1]):
                    continue

                try:
                    days_val = int(days_item.text().strip())
                except:
                    continue
                if days_val < 0:
                    if patient_name in included_names:
                        continue
                    charge_item = self.parent.table.item(row, 2)
                    notes_item = self.parent.table.item(row, 12)
                    charge_flag = "$" if charge_item and charge_item.text().strip().upper() == "$" else ""
                    notes_text = notes_item.text() if notes_item else ""

                    self.parent.cur.execute(
                        "SELECT weeks_per_blister, pack_size FROM patients WHERE UPPER(name) = ?",
                        (patient_name.upper(),)
                    )

                    db_row = self.parent.cur.fetchone()
                    weeks_per_blister = db_row[0] if db_row and db_row[0] else ""
                    tray_size = db_row[1] if db_row and db_row[1] else ""

                    overdue_rows.append([
                        (charge_flag, False),
                        (patient_name, True),
                        (self.parent.table.item(row, 6).text().strip(), False),
                        (str(weeks_per_blister), False),
                        (str(tray_size), False),
                        ("", False),
                        ("", False),
                        ("", False),
                        (notes_text, True)
                    ])
            if overdue_rows:
                self.overdue_label.show()
                self.overdue_table.show()
                self.populate_table(self.overdue_table, overdue_rows)
            else:
                self.overdue_label.hide()
                self.overdue_table.hide()
        else:
            self.overdue_label.hide()
            self.overdue_table.hide()

    def do_print(self):
        print("DEBUG: Preparing direct print in Landscape...")
        QTimer.singleShot(0, self._execute_print_job)

    def _execute_print_job(self):
        # --- Validate printer using cached printer info ---
        app = self.parent
        available_printers = app._get_cached_printer_info_map()
        printer_name = app.get_saved_printer()

        if not printer_name:
            QMessageBox.warning(
                self,
                "Printer Not Set",
                "Please set a default printer in the Manage menu before printing."
            )
            return

        if printer_name not in available_printers:
            QMessageBox.warning(
                self,
                "Printer Unavailable",
                f"The saved printer '{printer_name}' is not currently available.\n"
                f"Use Manage -> Refresh Printers, then Set Printer."
            )
            return

        # --- Setup printer ---
        qt_printer_info = available_printers[printer_name]
        printer = QPrinter(qt_printer_info)
        printer.setOutputFormat(QPrinter.NativeFormat)
        printer.setPageSize(QPrinter.A4)
        printer.setPageLayout(
            QPageLayout(QPageSize(QPageSize.A4), QPageLayout.Landscape, QMarginsF(15, 5, 15, 30))
        )
        printer.setOrientation(QPrinter.Landscape)
        printer.setFullPage(False)
        printer.setDocName("DAACal Printout")

        painter = QPainter(printer)

        # --- Margins in pixels ---
        left_margin_px = printer.pageLayout().margins(QPageLayout.Point).left() * printer.logicalDpiX() / 72.0
        right_margin_px = printer.pageLayout().margins(QPageLayout.Point).right() * printer.logicalDpiX() / 72.0
        top_margin_px = printer.pageLayout().margins(QPageLayout.Point).top() * printer.logicalDpiY() / 72.0

        # --- Date labels ---
        date_layout = QVBoxLayout()
        date_layout.setSpacing(0)
        date_layout.setContentsMargins(0, 0, 0, 0)

        date_printed_label = QLabel(f"Date Printed: {self.date_printed_str}")
        date_printed_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        date_printed_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        if self.next_week_radio.isChecked():
            week_start_date = datetime.today() - timedelta(days=datetime.today().weekday()) + timedelta(days=7)
        else:
            week_start_date = datetime.today() - timedelta(days=datetime.today().weekday())

        week_start_label = QLabel(f"Week Starting: {week_start_date.strftime('%d/%m/%Y')}")
        week_start_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        week_start_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        date_layout.addWidget(date_printed_label)
        date_layout.addWidget(week_start_label)

        # --- Logo ---
        icon_label = QLabel()
        icon_path = os.path.join(BASE_DIR, "daacal.png")
        icon_pixmap = QPixmap()
        if os.path.exists(icon_path):
            icon_pixmap = QPixmap(icon_path).scaledToHeight(48, Qt.SmoothTransformation)
            icon_label.setPixmap(icon_pixmap)
        else:
            print(f"DEBUG: Icon not found at {icon_path}")

        # --- Horizontal layout: dates first, logo last ---
        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(0, 0, 0, 0)
        header_layout.setSpacing(10)
        header_layout.setSizeConstraint(QLayout.SetFixedSize)
        header_layout.addLayout(date_layout)
        header_layout.addWidget(icon_label, alignment=Qt.AlignVCenter)

        # --- Wrap in QWidget for rendering ---
        header_widget = QWidget()
        header_widget.setAutoFillBackground(False)
        header_widget.setStyleSheet("background: transparent;")
        header_widget.setLayout(header_layout)
        header_widget.adjustSize()
        header_widget.resize(header_widget.sizeHint())

        # --- Vertical gap before first table ---
        gap_height = 70

        # --- Position header in the gap ---
        page_width = printer.pageRect(QPrinter.DevicePixel).width()
        header_x = page_width - header_widget.width() - right_margin_px
        header_y = top_margin_px + (gap_height - header_widget.height()) / 2

        painter.translate(header_x, header_y)
        header_widget.render(painter)
        painter.translate(-header_x, -header_y)

        # --- Start tables after gap ---
        y_offset = top_margin_px + gap_height

        # --- Main table label ---
        painter.setFont(QFont("Arial", 12, QFont.Bold))
        painter.drawText(int(left_margin_px), int(y_offset), self.main_label.text())
        y_offset += 20

        # --- Print main table ---
        if self.preview_table.isVisible():
            self.apply_fixed_column_widths_print(self.preview_table, printer)
            table_width = sum(self.preview_table.columnWidth(i) for i in range(self.preview_table.columnCount()))
            table_height = (
                self.preview_table.horizontalHeader().height()
                + sum(self.preview_table.rowHeight(i) for i in range(self.preview_table.rowCount()))
            )
            self.preview_table.resize(table_width, table_height)
            painter.translate(left_margin_px, y_offset)
            self.preview_table.render(painter)
            painter.translate(-left_margin_px, -y_offset)
            y_offset += table_height + 40

        # --- Print overdue table ---
        if self.overdue_table.isVisible():
            painter.setFont(QFont("Arial", 12, QFont.Bold))
            painter.drawText(int(left_margin_px), int(y_offset), self.overdue_label.text())
            y_offset += 20
            self.apply_fixed_column_widths_print(self.overdue_table, printer)
            table_width = sum(self.overdue_table.columnWidth(i) for i in range(self.overdue_table.columnCount()))
            table_height = (
                self.overdue_table.horizontalHeader().height()
                + sum(self.overdue_table.rowHeight(i) for i in range(self.overdue_table.rowCount()))
            )
            self.overdue_table.resize(table_width, table_height)
            painter.translate(left_margin_px, y_offset)
            self.overdue_table.render(painter)
            painter.translate(-left_margin_px, -y_offset)
            y_offset += table_height + 40

        painter.end()
        print("DEBUG: Direct print complete with balanced margins and reduced top gap.")


    def get_saved_printer(self):
        settings = _load_daacal_settings()
        saved_printer = (settings.get("default_printer") or "").strip()

        available_printers = [p.printerName() for p in QPrinterInfo.availablePrinters()]
        if saved_printer and saved_printer in available_printers:
            return saved_printer
        return ""

class PrinterRefreshWorker(QObject):
    finished = pyqtSignal(list, object)  # (infos, error)

    @pyqtSlot()
    def run(self):
        try:
            infos = list(QPrinterInfo.availablePrinters())
            self.finished.emit(infos, None)
        except Exception as e:
            self.finished.emit([], e)


class WebsterCalendarApp(QMainWindow):
    db_changed_signal = pyqtSignal()

    def _startup_step(self, text):
        splash = getattr(self, "splash", None)
        if splash:
            splash.raise_()
            splash.activateWindow()
            QApplication.processEvents()
    def showEvent(self, event):
        super().showEvent(event)

        splash = getattr(self, "splash", None)
        if splash is None:
            return

        self.splash = None
        QTimer.singleShot(150, splash.close)
    def _apply_printer_cache(self, infos):
        """
        Apply a freshly enumerated printer list to cache.
        """
        infos = list(infos or [])
        names = sorted(
            {p.printerName().strip() for p in infos if p.printerName().strip()},
            key=str.casefold
        )

        self._cached_printer_infos = infos
        self._cached_printers = names
        self._printer_cache_ready = True

        print(f"DEBUG: Printer cache refreshed: {names}")
        return names

    def _refresh_printer_cache(self):
        """
        Synchronous fallback only.
        Prefer _start_printer_refresh() for normal use.
        """
        try:
            infos = list(QPrinterInfo.availablePrinters())
            return self._apply_printer_cache(infos)
        except Exception as e:
            print(f"DEBUG: Failed to refresh printer cache: {e}")
            self._cached_printer_infos = []
            self._cached_printers = []
            self._printer_cache_ready = True
            return []

    def _set_printer_dialog_busy(self, busy: bool):
        """
        Update printer dialog UI while refresh is running.
        """
        dlg = getattr(self, "_printer_dialog", None)
        if dlg is None:
            return

        refresh_btn = getattr(self, "_printer_refresh_btn", None)
        combo = getattr(self, "_printer_combo_box", None)
        ok_btn = getattr(self, "_printer_ok_btn", None)
        status_lbl = getattr(self, "_printer_status_label", None)

        if refresh_btn is not None:
            refresh_btn.setEnabled(not busy)
        if combo is not None:
            combo.setEnabled(not busy)
        if ok_btn is not None:
            ok_btn.setEnabled((not busy) and combo is not None and combo.count() > 0)

        if status_lbl is not None:
            status_lbl.setText("Refreshing printers..." if busy else "")

    def _populate_printer_combo(self, preserve_selection=True):
        """
        Fill the printer combo from cache.
        """
        combo = getattr(self, "_printer_combo_box", None)
        if combo is None:
            return

        previous = combo.currentText().strip() if preserve_selection else ""
        saved = self.get_saved_printer_name_only()

        combo.blockSignals(True)
        combo.clear()
        combo.addItems(list(getattr(self, "_cached_printers", []) or []))

        target = ""
        if previous and previous in self._cached_printers:
            target = previous
        elif saved and saved in self._cached_printers:
            target = saved
        elif self._cached_printers:
            target = self._cached_printers[0]

        if target:
            combo.setCurrentText(target)

        combo.blockSignals(False)

        ok_btn = getattr(self, "_printer_ok_btn", None)
        if ok_btn is not None:
            ok_btn.setEnabled(combo.count() > 0)

    def _cleanup_printer_refresh_thread(self):
        """
        Safely tear down any previous refresh thread objects.
        """
        thread = getattr(self, "_printer_refresh_thread", None)
        worker = getattr(self, "_printer_refresh_worker", None)

        self._printer_refresh_thread = None
        self._printer_refresh_worker = None

        if thread is not None:
            try:
                thread.quit()
            except Exception:
                pass
            try:
                thread.wait(200)
            except Exception:
                pass
            try:
                thread.deleteLater()
            except Exception:
                pass

        if worker is not None:
            try:
                worker.deleteLater()
            except Exception:
                pass

    def _start_printer_refresh(self, show_dialog_feedback=False):
        """
        Refresh printers on a background QThread.
        If a printer picker dialog is open, update it when finished.
        """
        if getattr(self, "_printer_refresh_in_progress", False):
            return

        self._printer_refresh_in_progress = True

        if show_dialog_feedback:
            self._set_printer_dialog_busy(True)

        thread = QtCore.QThread(self)
        worker = PrinterRefreshWorker()
        worker.moveToThread(thread)

        self._printer_refresh_thread = thread
        self._printer_refresh_worker = worker

        thread.started.connect(worker.run)
        worker.finished.connect(self._on_printer_refresh_finished)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        thread.start()

    @pyqtSlot(list, object)
    def _on_printer_refresh_finished(self, infos, error):
        """
        Apply results from threaded refresh.
        """
        self._printer_refresh_in_progress = False

        if error is not None:
            print(f"DEBUG: Failed to refresh printer cache: {error}")
            self._cached_printer_infos = []
            self._cached_printers = []
            self._printer_cache_ready = True
        else:
            self._apply_printer_cache(infos)

        # If printer picker dialog is open, update it live
        if getattr(self, "_printer_dialog", None) is not None:
            self._populate_printer_combo(preserve_selection=True)
            self._set_printer_dialog_busy(False)

            # Also refresh the MPS combo if it exists
            mps_combo = getattr(self, "_printer_mps_combo_box", None)
            if mps_combo is not None:
                current = mps_combo.currentText().strip()
                settings = _load_daacal_settings()
                saved_mps = (settings.get("mps_default_printer") or "").strip()

                mps_combo.blockSignals(True)
                mps_combo.clear()
                mps_combo.addItems(list(getattr(self, "_cached_printers", []) or []))

                target = ""
                if current and current in self._cached_printers:
                    target = current
                elif saved_mps and saved_mps in self._cached_printers:
                    target = saved_mps
                else:
                    for name in (self._cached_printers or []):
                        if "MPS" in name.upper():
                            target = name
                            break

                if target:
                    mps_combo.setCurrentText(target)

                mps_combo.blockSignals(False)

            if error is not None:
                QMessageBox.warning(
                    self._printer_dialog,
                    "Printer Refresh Failed",
                    f"Could not refresh printers.\n\n{error}"
                )
            elif not self._cached_printers:
                QMessageBox.warning(
                    self._printer_dialog,
                    "No Printers Found",
                    "No printers are currently available."
                )

        self._cleanup_printer_refresh_thread()

    def refresh_printers(self):
        """
        Optional external/manual refresh entrypoint.
        Safe to keep for startup warmup or any other internal trigger.
        """
        self._start_printer_refresh(show_dialog_feedback=False)

    def _ensure_printer_cache(self):
        """
        Make sure some cache exists before using it.
        Uses synchronous fallback only if startup refresh has not completed yet.
        """
        if not getattr(self, "_printer_cache_ready", False):
            self._refresh_printer_cache()

    def _get_cached_printer_info_map(self):
        """
        Return {printer_name: QPrinterInfo} from cache.
        """
        self._ensure_printer_cache()

        infos = list(getattr(self, "_cached_printer_infos", []) or [])
        return {
            p.printerName().strip(): p
            for p in infos
            if p.printerName().strip()
        }

    def get_saved_printer_name_only(self):
        """
        Read the saved printer name from settings only.
        Does not enumerate printers.
        """
        settings = _load_daacal_settings()
        return (settings.get("default_printer") or "").strip()

    def get_saved_printer(self):
        """
        Return saved printer name only if it exists in cached printers.
        """
        saved_printer = self.get_saved_printer_name_only()
        if not saved_printer:
            return ""

        self._ensure_printer_cache()
        return saved_printer if saved_printer in (self._cached_printers or []) else ""

    def set_printer(self):
        """
        Show a custom printer picker dialog with:
          - default printer dropdown
          - MPS default printer dropdown
          - MPS orientation dropdown
          - one refresh button
          - threaded refresh
        """
        self._ensure_printer_cache()

        dlg = QDialog(self)
        dlg.setWindowTitle("Set Printer")
        dlg.setModal(True)
        dlg.resize(520, 220)

        main_layout = QVBoxLayout(dlg)

        # --- Default printer ---
        main_layout.addWidget(QLabel("Default printer:"))
        default_row = QHBoxLayout()

        default_combo = QComboBox()
        default_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        refresh_btn = QToolButton()
        refresh_btn.setAutoRaise(False)
        refresh_btn.setToolTip("Refresh printers")
        refresh_btn.setCursor(Qt.PointingHandCursor)
        refresh_btn.setFixedSize(32, 32)
        refresh_btn.setIcon(self.style().standardIcon(QtWidgets.QStyle.SP_BrowserReload))
        refresh_btn.setIconSize(QSize(18, 18))

        default_row.addWidget(default_combo, 1)
        default_row.addWidget(refresh_btn, 0)
        main_layout.addLayout(default_row)

        # --- MPS default printer ---
        main_layout.addWidget(QLabel("MPS default printer:"))
        mps_combo = QComboBox()
        mps_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        main_layout.addWidget(mps_combo)

        # --- MPS orientation ---
        main_layout.addWidget(QLabel("MPS orientation:"))
        mps_orientation_combo = QComboBox()
        mps_orientation_combo.addItems(["Portrait", "Landscape"])
        main_layout.addWidget(mps_orientation_combo)

        status_lbl = QLabel("")
        status_lbl.setStyleSheet("color: #666;")
        main_layout.addWidget(status_lbl)

        button_row = QHBoxLayout()
        button_row.addStretch(1)
        ok_btn = QPushButton("OK")
        cancel_btn = QPushButton("Cancel")
        button_row.addWidget(ok_btn)
        button_row.addWidget(cancel_btn)
        main_layout.addLayout(button_row)

        ok_btn.setDefault(True)

        # expose dialog widgets to helper methods
        self._printer_dialog = dlg
        self._printer_combo_box = default_combo
        self._printer_refresh_btn = refresh_btn
        self._printer_ok_btn = ok_btn
        self._printer_status_label = status_lbl
        self._printer_mps_combo_box = mps_combo

        # populate first combo using existing helper
        self._populate_printer_combo(preserve_selection=False)

        # populate second combo manually
        mps_combo.blockSignals(True)
        mps_combo.clear()
        mps_combo.addItems(list(getattr(self, "_cached_printers", []) or []))

        settings = _load_daacal_settings()
        saved_default = (settings.get("default_printer") or "").strip()
        saved_mps = (settings.get("mps_default_printer") or "").strip()
        saved_mps_orientation = (settings.get("mps_default_orientation") or "portrait").strip().lower()

        if saved_default and saved_default in (self._cached_printers or []):
            default_combo.setCurrentText(saved_default)

        if saved_mps and saved_mps in (self._cached_printers or []):
            mps_combo.setCurrentText(saved_mps)
        else:
            # fallback: auto-pick the first printer containing MPS
            for name in (self._cached_printers or []):
                if "MPS" in name.upper():
                    mps_combo.setCurrentText(name)
                    break

        if saved_mps_orientation == "landscape":
            mps_orientation_combo.setCurrentText("Landscape")
        else:
            mps_orientation_combo.setCurrentText("Portrait")

        mps_combo.blockSignals(False)

        if default_combo.count() == 0:
            status_lbl.setText("No cached printers yet. Click refresh.")

        def _refresh_both():
            self._start_printer_refresh(show_dialog_feedback=True)

        refresh_btn.clicked.connect(_refresh_both)
        ok_btn.clicked.connect(dlg.accept)
        cancel_btn.clicked.connect(dlg.reject)

        result = dlg.exec_()

        default_chosen = default_combo.currentText().strip()
        mps_chosen = mps_combo.currentText().strip()
        mps_orientation = mps_orientation_combo.currentText().strip().lower()

        # clear dialog refs no matter how it closes
        self._printer_dialog = None
        self._printer_combo_box = None
        self._printer_refresh_btn = None
        self._printer_ok_btn = None
        self._printer_status_label = None
        self._printer_mps_combo_box = None

        if result != QDialog.Accepted:
            return

        if not default_chosen:
            QMessageBox.warning(self, "No Printer Selected", "Please select a default printer.")
            return

        if not mps_chosen:
            QMessageBox.warning(self, "No MPS Printer Selected", "Please select an MPS default printer.")
            return

        _save_daacal_settings({
            "default_printer": default_chosen,
            "mps_default_printer": mps_chosen,
            "mps_default_orientation": mps_orientation,
        })

        print(f"DEBUG: Default printer set to: {default_chosen}")
        print(f"DEBUG: MPS default printer set to: {mps_chosen}")
        print(f"DEBUG: MPS orientation set to: {mps_orientation}")

    def _pack_entry_is_live(self) -> bool:
        table = getattr(self, "pack_entry_drug_table", None)
        ws = getattr(self, "pack_entry_ws", None)
        if table is None or ws is None:
            return False
        try:
            return not sip.isdeleted(table)
        except Exception:
            # If sip isn't available for some reason, best-effort: assume live
            return True

    def _clear_pack_entry_context(self) -> None:
        self.pack_entry_ws = None
        self.pack_entry_drug_table = None
        self.blank_pack_window = None

    def refresh_current_patient_tab(self):
        """
        Rebuilds the currently open patient tab (patient profile) so pack viewer / active drug-dependent
        UI reflects latest Excel + DB changes.
        """
        pid = getattr(self, "current_patient_id", None)
        if not pid:
            return

        # Find the patient row in the main table
        row = self.find_patient_row_by_id(pid)
        if row is None:
            return

        # Close existing tab for this patient (so open_patient_tab rebuilds it)
        number = str(getattr(self, "current_patient_number", "") or "").strip()
        if number:
            for i in range(self.tabs.count()):
                tab = self.tabs.widget(i)
                if tab and tab.property("patient_number") == number:
                    self.tabs.removeTab(i)
                    tab.deleteLater()
                    break

        # Re-open (rebuild) the patient tab
        self.open_patient_tab(row, 0)

    def open_mps_blister_wizard(self):
        downloads = QStandardPaths.writableLocation(QStandardPaths.DownloadLocation)

        pdf_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select blister PDF",
            downloads,
            "PDF files (*.pdf)"
        )
        if not pdf_path:
            return

        dlg = MPSBlisterWizardDialog(self, pdf_path)
        dlg.exec_()

    def open_pdf_editor_dialog(self):
        downloads = QStandardPaths.writableLocation(QStandardPaths.DownloadLocation)

        pdf_path, _ = QFileDialog.getOpenFileName(
            self,
            "Open PDF for editing",
            downloads,
            "PDF files (*.pdf)"
        )
        if not pdf_path:
            return

        dlg = PdfEditorDialog(self, pdf_path)
        dlg.exec_()


    def to_db_name(self, display_name):
        # Convert "First Last" → "Last, First"
        parts = display_name.strip().split()
        if len(parts) < 2:
            return display_name  # cannot convert
        first = parts[0]
        last = " ".join(parts[1:])
        return f"{last}, {first}"

    def resolve_excel_path(self, patient_name):
        """
        Attempts both FIRSTNAME LASTNAME.xlsx and LASTNAME FIRSTNAME.xlsx.
        Returns the first existing path, or None if none found.
        """

        base = patient_name.upper().replace(",", "").strip()
        first_last = f"{base}.xlsx"

        # Reverse order → LASTNAME FIRSTNAME
        parts = base.split()
        if len(parts) >= 2:
            last_first = f"{parts[-1]} {' '.join(parts[:-1])}.xlsx"
        else:
            last_first = first_last  # fallback

        path1 = os.path.join(COLLECTION_LOGS_DIR, first_last)
        path2 = os.path.join(COLLECTION_LOGS_DIR, last_first)

        if os.path.exists(path1):
            return path1
        if os.path.exists(path2):
            return path2

        return None

    def open_claims_window(self):
        dlg = ClaimsWindow(self, self.cur)
        dlg.exec_()

    def is_valid_date(self, s: str) -> bool:
        """Return True only if s is a real dd/MM/yyyy date."""
        if not s or not s.strip():
            return False
        d = QDate.fromString(s.strip(), "dd/MM/yyyy")
        return d.isValid()

    def is_valid_medicare(self, val: str) -> bool:
        """
        Validate Australian Medicare number.
        Format: 11 digits (8 base + 1 check + 1 issue + 1 IRN)
        Example: 12345678912  (commonly printed as '1234 56789 1 2')
        """
        if not val or not re.fullmatch(r"\d{11}", val):
            return False
        digits = [int(x) for x in val[:8]]
        check_digit = int(val[8])
        weights = [1, 3, 7, 9, 1, 3, 7, 9]
        return sum(a * b for a, b in zip(digits, weights)) % 10 == check_digit

    def is_valid_concession(self, val: str) -> bool:
        """
        Validate Concession (Health Care/Pensioner) OR DVA.
        Concession: exactly 9 digits + 1 letter (e.g., 123456789A)
        DVA: either 'QSM' + 5 digits (e.g., QSM12345) OR 4 letters + 4 digits (e.g., NBUR9080)
        """
        if not val:
            return False
        v = val.strip().upper()
        if re.fullmatch(r"\d{9}[A-Z]", v):  # Concession
            return True
        if re.fullmatch(r"QSM\d{5}", v):  # DVA (QSM + 5 digits)
            return True
        if re.fullmatch(r"[A-Z]{4}\d{4}", v):  # DVA (4 letters + 4 digits)
            return True
        return False

    def _update_packed_collected_lock(self):
        """Disable packed or collected date fields only if the other actually contains a real date."""
        if not hasattr(self, 'packed_input') or not hasattr(self, 'picked_input'):
            return

        packed_text = self.packed_input.text().strip()
        collected_text = self.picked_input.text().strip()

        # Normalize: ignore masked placeholders and underscores
        if packed_text in ("", "__/__/____"):
            packed_text = ""
        if collected_text in ("", "__/__/____"):
            collected_text = ""

        packed_valid = self.is_valid_date(packed_text)
        collected_valid = self.is_valid_date(collected_text)

        if collected_valid:
            # Collected filled → disable packed
            self.packed_input.setDisabled(True)
            self.picked_input.setDisabled(False)
            self.packed_input.setStyleSheet("background-color: #f0f0f0; color: #888;")
            self.picked_input.setStyleSheet("")
        elif packed_valid:
            # Packed filled → disable collected
            self.picked_input.setDisabled(True)
            self.packed_input.setDisabled(False)
            self.picked_input.setStyleSheet("background-color: #f0f0f0; color: #888;")
            self.packed_input.setStyleSheet("")
        else:
            # Both empty or invalid → enable both
            self.packed_input.setDisabled(False)
            self.picked_input.setDisabled(False)
            self.packed_input.setStyleSheet("")
            self.picked_input.setStyleSheet("")

    def print_this_week(self):
        """Send the preview table to the printer."""
        try:
            printer = QPrinter(QPrinter.HighResolution)
            dialog = QPrintDialog(printer, self)
            dialog.setWindowTitle("Print")
            if dialog.exec_() == QPrintDialog.Accepted:
                # Render the preview table directly to the printer
                painter = QPainter(printer)
                self.preview_table.render(painter)
                painter.end()
        except Exception as e:
            print(f"ERROR in print_this_week: {e}")

    def resolve_user_name(self, user_id):
        return USER_ID_MAP.get(user_id.strip().lower(), None)

    def handle_packer_input_change(self):
        self.packer_input.setStyleSheet("")  # no clearing

    def fill_packer_full_name(self):
        entered_id = self.packer_input.text().strip()
        name = self.resolve_user_name(entered_id)
        if name:
            self.packer_input.setText(name)

    def set_data_directory(self):
        directory = QFileDialog.getExistingDirectory(self, "Select Data Directory")
        if directory:
            _save_daacal_settings({"data_directory": directory})
            QMessageBox.information(self, "Saved", "Directory set to: " + directory + "\nRestart DAACal to apply.")

    def open_edit_drugs_dialog(self):
        print("DEBUG: open_edit_drugs_dialog called")
        try:
            # --- Cache pack-entry table data ONLY if pack-entry context exists ---
            cached_values = {}
            selected_row = -1
            scroll_value = 0
            table = None

            if self._pack_entry_is_live():
                table = self.pack_entry_drug_table
                selected_row = table.currentRow()
                scroll_value = table.verticalScrollBar().value()

                for r in range(table.rowCount()):
                    drug_item = table.item(r, 0)
                    if not drug_item:
                        continue
                    drug_name = drug_item.text().strip().lower()
                    be = table.item(r, 1)
                    disp = table.item(r, 2)
                    pack = table.item(r, 3)
                    cached_values[drug_name] = {
                        "be": be.text() if be else "",
                        "disp": disp.text() if disp else "",
                        "pack": pack.text() if pack else "",
                    }

            # --- Open Edit Drugs dialog ---
            parent = self.blank_pack_window if hasattr(self, "blank_pack_window") and self.blank_pack_window else self
            dlg = EditDrugsDialog(self.current_patient_id, self.current_med_names, self.conn, parent)
            dlg.drugs_updated.connect(self.refresh_current_patient_tab)
            dlg.excel_path = self.excel_path

            print("DEBUG: Dialog created, calling exec_()")
            result = dlg.exec_()
            print("DEBUG: Dialog closed with result", result)
            if result != QDialog.Accepted:
                return

            # --- Reload workbook + med names ---
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            med_names = []
            r = 4
            while True:
                val = ws.cell(row=r, column=1).value
                if val is None or str(val).strip() == "":
                    break
                med_names.append(val)
                r += 1
            self.current_med_names = med_names

            # --- Build active drug names from DB ---
            cur = self.conn.cursor()
            active_med_names = []
            for drug in self.current_med_names:
                cur.execute(
                    "SELECT active FROM patient_drug_status WHERE patient_id=? AND drug_name=?",
                    (self.current_patient_id, drug),
                )
                row = cur.fetchone()
                if row is not None and row[0]:
                    active_med_names.append(drug)

            # --- Repopulate pack-entry table + restore cached values (only if pack-entry exists) ---
            if self._pack_entry_is_live():
                self.repopulate_pack_entry_drug_table(active_med_names, self.current_patient_id, self.pack_entry_ws)

                table = self.pack_entry_drug_table
                for r in range(table.rowCount()):
                    drug_item = table.item(r, 0)
                    if not drug_item:
                        continue
                    drug_name = drug_item.text().strip().lower()
                    if drug_name in cached_values:
                        vals = cached_values[drug_name]
                        table.setItem(r, 1, QTableWidgetItem(vals["be"]))
                        table.setItem(r, 2, QTableWidgetItem(vals["disp"]))
                        table.setItem(r, 3, QTableWidgetItem(vals["pack"]))

                if 0 <= selected_row < table.rowCount():
                    table.setCurrentCell(selected_row, 0)
                    table.selectRow(selected_row)

                table.verticalScrollBar().setValue(scroll_value)
                table.repaint()
                print("DEBUG: Pack table restored successfully after Edit Drugs dialog.")

        except Exception as e:
            QMessageBox.critical(self, "Edit Active Drugs Error", str(e))
    def populate_drugs_table(self, med_names):
        if not hasattr(self, 'pack_entry_drug_table'):
            print("ERROR: No pack_entry_drug_table to update")
            return
        table = self.pack_entry_drug_table
        # Remove ALL rows
        table.setRowCount(0)
        # Add rows for current drugs
        for drug in med_names:
            row = table.rowCount()
            table.insertRow(row)
            table.setItem(row, 0, QTableWidgetItem(drug))
            # Optionally clear Batch/Expiry and Qty Packed columns
            table.setItem(row, 1, QTableWidgetItem(""))
            table.setItem(row, 2, QTableWidgetItem(""))
        table.update()
        table.repaint()
        print("DEBUG: Drugs table repopulated with:", med_names)

    def reload_active_drugs(self):
        print("DEBUG: reload_active_drugs CALLED")
        cur = self.conn.cursor()
        cur.execute("SELECT drug_name FROM patient_drug_status WHERE patient_id=? AND active=1",
                    (self.current_patient_id,))
        self.current_med_names = [row[0] for row in cur.fetchall()]
        print("DEBUG: Active drugs list after reload:", self.current_med_names)
        self.populate_drugs_table(self.current_med_names)

    def push_undo_action(self, patient_id, prev_data):
        self.undo_stack.append((patient_id, prev_data))
        if len(self.undo_stack) > 10:
            self.undo_stack.pop(0)
        self.redo_stack.clear()

    def perform_undo(self):
        if not self.undo_stack:
            return
        patient_id, old_data = self.undo_stack.pop()
        current_data = self.get_patient_row_data(patient_id)
        self.redo_stack.append((patient_id, current_data))
        self.set_patient_row_data(patient_id, old_data)

    def perform_redo(self):
        if not self.redo_stack:
            return
        patient_id, redo_data = self.redo_stack.pop()
        current_data = self.get_patient_row_data(patient_id)
        self.undo_stack.append((patient_id, current_data))
        self.set_patient_row_data(patient_id, redo_data)

    def get_patient_row_data(self, patient_id):
        row = self.find_patient_row_by_id(patient_id)
        if row is None:
            return None
        return {
            'name': self.table.item(row, 1).text(),
            'notes': self.table.item(row, 2).text(),
            'pack_date': self.table.item(row, 3).text(),
            'collect_date': self.table.item(row, 4).text(),
            'due_date': self.table.item(row, 5).text(),
            'given_by': self.table.item(row, 6).text()
        }

    def set_patient_row_data(self, patient_id, data):
        row = self.find_patient_row_by_id(patient_id)
        if row is None:
            return
        self.table.item(row, 1).setText(data['name'])
        self.table.item(row, 2).setText(data['notes'])
        self.table.item(row, 3).setText(data['pack_date'])
        self.table.item(row, 4).setText(data['collect_date'])
        self.table.item(row, 5).setText(data['due_date'])
        self.table.item(row, 6).setText(data['given_by'])
        self.save_patient_data_from_table(row)

    def find_patient_row_by_id(self, patient_id):
        for row in range(self.table.rowCount()):
            if self.table.item(row, 0).text() == str(patient_id):
                return row
        return None

    def try_close_tab_by_index(self, index):
        tab_widget = self.tabs.widget(index)
        if not tab_widget:
            return

        # --- Detect unsaved changes across all patient fields ---
        tracked_fields = [
            'first_name_input', 'last_name_input', 'notes_input',
            'packed_input', 'picked_input', 'medicare_input',
            'concession_input', 'given_by_input'
        ]

        modified = False

        # Check text-based widgets
        for field in tracked_fields:
            if hasattr(self, field):
                widget = getattr(self, field)
                if isinstance(widget, QLineEdit) and widget.isModified():
                    modified = True
                    break

        # Check checkboxes and dropdowns (compare to DB)
        if not modified:
            try:
                # Identify which patient tab is active
                if hasattr(self, "current_patient_number"):
                    number = str(self.current_patient_number).strip()
                    if number:
                        self.cur.execute("""
                                         SELECT charge, partial_supply, pack_size, weeks_per_blister
                                         FROM patients
                                         WHERE number = ?
                                         """, (number,))
                        current_db = self.cur.fetchone()
                        if current_db:
                            charge_db, partial_db, packsize_db, weeks_db = current_db

                            # Charge checkbox
                            if hasattr(self, 'charge_button'):
                                if self.charge_button.isChecked() != (charge_db == 'Y'):
                                    modified = True

                            # Partial supply checkbox
                            if hasattr(self, 'claim_button'):
                                if self.claim_button.isChecked() != (partial_db == 'Y'):
                                    modified = True

                            # Pack size dropdown
                            if hasattr(self, 'pack_size_dropdown'):
                                if self.pack_size_dropdown.currentText() != (packsize_db or ''):
                                    modified = True

                            # Weeks per blister dropdown
                            if hasattr(self, 'weeks_per_blister_dropdown'):
                                if self.weeks_per_blister_dropdown.currentText() != (weeks_db or ''):
                                    modified = True
            except Exception as e:
                print("DEBUG: Unsaved check failed:", e)

        # --- Prompt user if anything changed ---
        if modified:
            reply = QMessageBox.question(
                self,
                "Unsaved Changes",
                "Do you want to save changes before closing?",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel
            )
            if reply == QMessageBox.Cancel:
                return
            if reply == QMessageBox.Yes:
                for btn in tab_widget.findChildren(QPushButton):
                    if btn.text().lower().startswith("save"):
                        btn.click()
                        return

        # --- Close tab normally ---
        closing_calendar = (tab_widget is self.main_widget)
        self.tabs.removeTab(index)

        # --- If Calendar tab was closed, rebuild it from scratch ---
        if closing_calendar:
            try:
                # destroy old calendar page completely
                old_widget = self.main_widget
                self.main_widget = QWidget()
                self.layout = QVBoxLayout()
                self.main_widget.setLayout(self.layout)

                # rebuild toolbar
                self.top_buttons = QHBoxLayout()

                self.top_buttons.addWidget(self.print_button)
                self.top_buttons.addWidget(self.new_packed_button)
                self.top_buttons.addWidget(self.checked_button)
                self.top_buttons.addWidget(self.collected_button)
                self.top_buttons.addWidget(self.flag_button)
                self.top_buttons.addWidget(self.pause_button)
                self.top_buttons.addWidget(self.login_input)
                self.layout.addLayout(self.top_buttons)

                # rebuild table widget completely
                self.table = QTableWidget()
                self.table.verticalHeader().setVisible(False)
                self.table.cellClicked.connect(self.handle_warning_click)
                self.table.cellPressed.connect(self.table_mouse_click)
                self.table.cellDoubleClicked.connect(self.open_patient_tab)
                self.layout.addWidget(self.table)

                self.setup_table()
                self.table.setItemPrototype(DateItem("01/01/2000"))

                ceased = bool(
                    getattr(self, "view_ceased_action", None)
                    and self.view_ceased_action.isChecked()
                )
                self.load_data(ceased=ceased)

                self.tabs.insertTab(0, self.main_widget, "Calendar")
                self.tabs.setCurrentWidget(self.main_widget)

                old_widget.deleteLater()

            except Exception as e:
                print("DEBUG: Failed to rebuild Calendar tab:", e)

    @pyqtSlot()
    def emit_db_changed_signal(self):
        """Emit DB changed signal to safely trigger reload in main thread."""
        self.db_changed_signal.emit()

    @pyqtSlot()
    def schedule_db_reload(self):
        if not hasattr(self, "_db_reload_timer"):
            self._db_reload_timer = QTimer(self)
            self._db_reload_timer.setSingleShot(True)
            self._db_reload_timer.timeout.connect(self._perform_db_reload)

        self._db_reload_timer.start(250)

    @pyqtSlot()
    def _perform_db_reload(self):
        self.load_data(
            ceased=bool(getattr(self, "view_ceased_action", None) and self.view_ceased_action.isChecked())
        )


    def __init__(self, splash=None):
        self.splash = splash

        self.undo_button = QPushButton()
        self.redo_button = QPushButton()
        self.undo_button.hide()
        self.redo_button.hide()
        self.undo_stack = []
        self.redo_stack = []

        super().__init__()

        self._startup_step("Loading DAACal...\nStarting up...")

        # --- Ceased Patients Buttons (hidden until needed) ---
        self.delete_button = QPushButton("Delete")
        self.restore_button = QPushButton("Restore")
        self.delete_button.clicked.connect(self.handle_delete_ceased_patient)
        self.restore_button.clicked.connect(self.handle_restore_ceased_patient)

        self._startup_step("Loading DAACal...\nPreparing window...")

        # === Set Window Title & Icon from data directory ===
        self.setWindowTitle("DAACal")
        icon_path = os.path.join(BASE_DIR, "daacal.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        else:
            print(f"WARNING: Icon not found at {icon_path}")

        # Establish database connection AFTER directory is resolved
        if not os.path.exists(DB_FILE):
            QMessageBox.warning(self, "Missing DB", "webster_calendar.db not found. Please select a valid directory.")
            self.set_data_directory()
            return

        self._startup_step("Loading DAACal...\nOpening database...")

        self.conn = sqlite3.connect(DB_FILE)
        self.cur = self.conn.cursor()

        # --- MPS PDF wizard remembered settings (per patient) ---
        self.cur.execute("""
                         CREATE TABLE IF NOT EXISTS mps_patient_settings
                         (
                             patient_name
                             TEXT
                             PRIMARY
                             KEY,
                             weeks_per_blister
                             INTEGER
                             DEFAULT
                             1,
                             header_json
                             TEXT
                             DEFAULT
                             NULL,
                             updated_ts
                             TEXT
                             DEFAULT
                             NULL
                         )
                         """)
        self.conn.commit()

        sync_user_map_to_db(self.conn)

        self.cur.execute("""
                         CREATE TABLE IF NOT EXISTS patient_drug_status
                         (
                             patient_id
                             INTEGER,
                             drug_name
                             TEXT,
                             active
                             INTEGER
                             DEFAULT
                             1,
                             PRIMARY
                             KEY
                         (
                             patient_id,
                             drug_name
                         )
                             )
                         """)
        self.conn.commit()

        # Ensure 'patients' table exists
        try:
            self.cur.execute("""
                             CREATE TABLE IF NOT EXISTS patients
                             (
                                 name
                                 TEXT
                                 PRIMARY
                                 KEY,
                                 notes
                                 TEXT,
                                 charge
                                 TEXT,
                                 pack_date
                                 TEXT,
                                 collect_date
                                 TEXT,
                                 due_date
                                 TEXT,
                                 given_by
                                 TEXT,
                                 weeks_per_blister
                                 TEXT,
                                 flagged
                                 TEXT,
                                 paused
                                 TEXT
                             )
                             """)
            self.conn.commit()
        except Exception as e:
            QMessageBox.critical(self, "DB Error", "Failed to create patients table:\n" + str(e))
            return

        self._startup_step("Loading DAACal...\nChecking database structure...")

        # ------------------------------------------------------------------
        # KEEP YOUR EXISTING DB MIGRATION / ALTER TABLE BLOCK HERE
        # This is the section from your current __init__ that includes:
        # - mps_patient_settings migration
        # - packed_by / checked_by / partial_supply / claim
        # - medicare / concession / given_out_by / weeks_per_blister
        # - any pack_size related commits
        # ------------------------------------------------------------------

        screen = QApplication.primaryScreen()
        size = screen.availableGeometry()
        screen_geometry = QApplication.primaryScreen().availableGeometry()
        rect = QApplication.primaryScreen().availableGeometry()
        self.setGeometry(rect)


        self._startup_step("Loading DAACal...\nBuilding window...")

        self.tabs = QTabWidget()
        self.tabs.setTabBar(CustomTabBar())
        self.tabs.setTabsClosable(True)
        self.tabs.tabCloseRequested.connect(self.try_close_tab_by_index)
        self.setCentralWidget(self.tabs)

        self.main_widget = QWidget()
        self.layout = QVBoxLayout()
        self.main_widget.setLayout(self.layout)
        self.tabs.addTab(self.main_widget, "Calendar")

        self.menu_bar = QMenuBar()
        self.setMenuBar(self.menu_bar)

        manage_menu = self.menu_bar.addMenu("Manage")
        self.manage_users_action = QAction("Users", self)
        self.manage_users_action.triggered.connect(self.open_users_page)
        self.manage_patients_action = QAction("Patients", self)
        self.manage_patients_action.triggered.connect(self.open_patients_page)
        manage_menu.addAction(self.manage_users_action)
        manage_menu.addAction(self.manage_patients_action)

        self.manage_directory_action = QAction("Directory", self)
        self.manage_directory_action.triggered.connect(self.set_data_directory)
        manage_menu.addAction(self.manage_directory_action)

        self.set_printer_action = QAction("Set Printer", self)
        self.set_printer_action.triggered.connect(self.set_printer)
        manage_menu.addAction(self.set_printer_action)

        view_menu = self.menu_bar.addMenu("View")
        self.view_ceased_action = QAction("View Ceased Patients", self)
        self.view_ceased_action.setCheckable(True)
        self.view_ceased_action.triggered.connect(self.load_ceased_data)
        view_menu.addAction(self.view_ceased_action)

        # --- Printer cache ---
        self._cached_printer_infos = []
        self._cached_printers = []
        self._printer_cache_ready = False
        self._printer_refresh_in_progress = False
        self._printer_refresh_thread = None
        self._printer_refresh_worker = None

        # live printer dialog refs
        self._printer_dialog = None
        self._printer_combo_box = None
        self._printer_refresh_btn = None
        self._printer_ok_btn = None
        self._printer_status_label = None

        # Warm the cache DURING startup so the splash stays up until it is ready
        self._startup_step("Loading DAACal...\nLoading printer cache...")
        self._refresh_printer_cache()

        # --- CLAIMS: single click opens the Claims window immediately ---
        self.claims_action = self.menu_bar.addAction("Claims")
        self.claims_action.triggered.connect(self.open_claims_window)

        # --- MPS Connect dropdown ---
        self.mps_menu = self.menu_bar.addMenu("MPS Connect")
        self.mps_blister_action = QAction("Blister PDF Wizard…", self)
        self.mps_blister_action.triggered.connect(self.open_mps_blister_wizard)
        self.mps_menu.addAction(self.mps_blister_action)

        self._startup_step("Loading DAACal...\nBuilding toolbar...")

        self.top_buttons = QHBoxLayout()

        self.undo_button.clicked.connect(self.perform_undo)
        self.redo_button.clicked.connect(self.perform_redo)
        self.undo_button.setFixedSize(30, 30)
        self.redo_button.setFixedSize(30, 30)

        self.print_button = QPushButton('Print')
        self.print_button.clicked.connect(self.handle_print_action)

        self.checked_button = QPushButton('Checked')
        self.checked_button.clicked.connect(self.handle_checked_action)

        self.new_packed_button = QPushButton('Pack')
        self.new_packed_button.clicked.connect(self.handle_new_packed_action)

        self.collected_button = QPushButton('Collected')
        self.collected_button.clicked.connect(self.handle_collected_action)

        self.pause_button = QPushButton('Pause')
        self.pause_button.clicked.connect(self.toggle_pause)

        self.flag_button = QPushButton('Changes')
        self.flag_button.clicked.connect(self.toggle_flag)

        self.login_input = QLineEdit()
        self.login_input.setPlaceholderText('Login')
        self.login_input.setFixedWidth(990)
        self.login_input.installEventFilter(self)
        self.login_input.mousePressEvent = self.login_click_select_all
        self.login_input.textChanged.connect(self.clear_active_user_if_empty)

        self.top_buttons.addWidget(self.print_button)
        self.top_buttons.addWidget(self.new_packed_button)
        self.top_buttons.addWidget(self.checked_button)
        self.top_buttons.addWidget(self.collected_button)
        self.top_buttons.addWidget(self.flag_button)
        self.top_buttons.addWidget(self.pause_button)
        self.top_buttons.addWidget(self.login_input)
        self.layout.addLayout(self.top_buttons)

        self._startup_step("Loading DAACal...\nCreating calendar table...")

        self.table = QTableWidget()
        self.table.verticalHeader().setVisible(False)
        self.table.cellClicked.connect(self.handle_warning_click)
        self.table.cellPressed.connect(self.table_mouse_click)
        self.table.cellDoubleClicked.connect(self.open_patient_tab)
        self.layout.addWidget(self.table)

        self.setup_table()
        self.table.setItemPrototype(DateItem("01/01/2000"))

        self._startup_step("Loading DAACal...\nLoading calendar data...")

        self.load_data()
        self.db_changed_signal.connect(self.schedule_db_reload)

        self.update_button_states()
        self.active_user = None
        self.login_input.returnPressed.connect(self.assign_active_user)

        self.inactivity_timer = QTimer()
        self.inactivity_timer.setInterval(120000)

        self.inactivity_timer.timeout.connect(self.clear_active_user_due_to_timeout)
        self.installEventFilter(self)
        self.inactivity_timer.start()

        self.viewing_ceased = False
        self.global_event_filter = GlobalActivityFilter(self)
        QApplication.instance().installEventFilter(self.global_event_filter)

        self._startup_step("Loading DAACal...\nStarting watchers...")

        # --- Auto-reload on DB file changes ---
        try:
            db_path = os.path.abspath(DB_FILE)
            db_dir = os.path.dirname(db_path)

            self.db_handler = DBChangeHandler(self)
            self.db_observer = Observer()
            self.db_observer.schedule(self.db_handler, path=db_dir, recursive=False)
            self.db_observer.start()

            print(f"DEBUG: Watching for database changes in {db_dir}")
        except Exception as e:
            print(f"DEBUG: Failed to start DB watcher: {e}")

        self._startup_step("Loading DAACal...\nFinalising...")

    def clear_active_user_due_to_timeout(self):
        self.active_user = None
        self.login_input.clear()
        self.login_input.setReadOnly(False)

        if hasattr(self, 'patient_login_input'):
            self.patient_login_input.clear()
            self.patient_login_input.setReadOnly(False)

        self.disable_patient_controls()
        self.viewing_ceased = False


    def setup_table(self):
        headers = ["", "#", "$", "Name", "Date Packed", "Picked Up", "Due Date", "Days Till Due", "Notes"]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        fixed_widths = [30, 30, 30, 300, 150, 150, 150, 150]
        for i, width in enumerate(fixed_widths):
            self.table.setColumnWidth(i, width)
            self.table.horizontalHeader().setSectionResizeMode(i, QHeaderView.Fixed)

        self.table.horizontalHeader().setSectionResizeMode(8, QHeaderView.Stretch)
        font = self.table.horizontalHeader().font()
        font.setBold(True)
        self.table.horizontalHeader().setFont(font)

        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)

    def load_data(self, ceased=False):
        try:
            self.table.setSortingEnabled(False)
            self.table.setRowCount(0)

            self.cur.execute(
                "SELECT id, number, charge, name, date_packed, picked_up, due_date, days_till_due, notes, paused, flagged, packed_by, checked_by "
                "FROM patients WHERE ceased = ? ORDER BY days_till_due ASC", (1 if ceased else 0,)
            )
            rows = self.cur.fetchall()

            # --- Precompute days_till_due (prefer stored due_date; else picked_up + 21) ---
            today_d = datetime.today().date()
            rows = [list(r) for r in rows]
            for i, r in enumerate(rows):
                due_src = (r[6] or "").strip()  # due_date column from DB
                picked_src = (r[5] or "").strip()
                if due_src:
                    try:
                        due = datetime.strptime(due_src, "%d/%m/%Y").date()
                        rows[i][7] = (due - today_d).days
                    except Exception as e:
                        rows[i][7] = None
                        print(f"WARNING: Could not calculate Days Till Due for {r[3]} — Due='{r[6]}' — {e}")
                elif picked_src:
                    try:
                        due = datetime.strptime(picked_src, "%d/%m/%Y").date() + timedelta(days=21)
                        rows[i][7] = (due - today_d).days
                    except Exception as e:
                        rows[i][7] = None
                        print(f"WARNING: Could not calculate Days Till Due for {r[3]} — Picked Up='{r[5]}' — {e}")
                else:
                    rows[i][7] = None

            # --- Sort: priority to 'packed but unchecked', then paused/flagged, else by days_till_due then recency of packed ---
            def sort_key_override(r):
                packed_by = str(r[11] or '').strip()
                checked_by = str(r[12] or '').strip()
                has_unchecked_icon = bool(packed_by) and not checked_by
                if has_unchecked_icon:
                    return (-1, -1)  # highest priority
                if r[9]:  # paused
                    return (99999, float('-inf'))
                if r[10]:  # flagged
                    return (99998, float('-inf'))
                try:
                    packed_ts = -datetime.strptime(r[4], "%d/%m/%Y").timestamp() if r[4] else float('-inf')
                except Exception:
                    packed_ts = float('-inf')
                day_key = r[7] if isinstance(r[7], int) else 9999
                return (day_key, packed_ts)

            rows.sort(key=sort_key_override)
            self.table.setRowCount(len(rows))

            # --- Week windows for colouring ---
            today_dt = datetime.today()
            start_of_week = today_dt - timedelta(days=today_dt.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            next_week_start = start_of_week + timedelta(days=7)
            next_week_end = start_of_week + timedelta(days=13)

            for row_idx, row in enumerate(rows):
                pid, number, charge, name, date_packed, picked_up, due_date, _, notes, paused, flagged, packed_by, checked_by = row

                # Initialise display vars
                warning_icon = ""
                due_date_str = ""
                days_text = ""

                # Unchecked indicator
                packed_str = (packed_by or "").strip()
                checked_str = (checked_by or "").strip()
                if packed_str and not checked_str:
                    warning_icon = "🔲"

                # --- Row colour logic ---
                row_color = QColor(255, 255, 255)  # default white
                try:
                    if paused:
                        row_color = QColor(255, 0, 0)  # red
                    elif flagged:
                        row_color = QColor(255, 192, 203)  # pink
                    elif (due_date and str(due_date).strip()) or (picked_up and picked_up.strip()):
                        base_due_str = str(due_date).strip() if (due_date and str(due_date).strip()) \
                            else (datetime.strptime(picked_up.strip(), "%d/%m/%Y") + timedelta(days=21)).strftime(
                            "%d/%m/%Y")
                        due_dt = datetime.strptime(base_due_str, "%d/%m/%Y")
                        if start_of_week.date() <= due_dt.date() <= end_of_week.date():
                            row_color = QColor(255, 255, 0)  # yellow = this week
                        elif next_week_start.date() <= due_dt.date() <= next_week_end.date():
                            row_color = QColor(204, 153, 255)  # purple = next week
                    elif date_packed and not picked_up:
                        # Green for packed & uncollected
                        row_color = QColor(0, 255, 0)
                        try:
                            dp = datetime.strptime(date_packed.strip(), "%d/%m/%Y")
                            if (datetime.today() - dp).days > 30:
                                warning_icon = "⚠️"
                        except Exception:
                            pass
                except Exception as e:
                    print("DEBUG: colouring failed:", e)

                # --- Displayed Due Date & Days (prefer stored due_date) ---
                try:
                    if (due_date and str(due_date).strip()) or (picked_up and picked_up.strip()):
                        base_due_str = str(due_date).strip() if (due_date and str(due_date).strip()) \
                            else (datetime.strptime(picked_up.strip(), "%d/%m/%Y") + timedelta(days=21)).strftime(
                            "%d/%m/%Y")
                        due_dt = datetime.strptime(base_due_str, "%d/%m/%Y")
                        due_date_str = due_dt.strftime("%d/%m/%Y")
                        delta_days = (due_dt.date() - datetime.today().date()).days
                        days_text = str(delta_days)
                        try:
                            if int(delta_days) < 0:
                                warning_icon = "⚠️"
                        except Exception:
                            pass
                    elif date_packed and not picked_up:
                        # show packed date in Due column if nothing else
                        due_date_str = ""
                        days_text = ""
                except Exception as e:
                    print("DEBUG: due/days display failed:", e)
                    due_date_str = ""
                    days_text = ""

                # Override text for paused/flagged
                if paused:
                    days_text = "PAUSED"
                elif flagged:
                    days_text = "FLAGGED"

                # --- Populate table cells ---
                values = [warning_icon, number, charge, name, date_packed, picked_up, due_date_str, days_text, notes]
                for col_idx, value in enumerate(values):
                    if col_idx == 1:
                        item = NumericItem("" if not value else str(value))
                    elif col_idx == 3:
                        # ✅ Patient Name → display exactly as stored ("Last, First")
                        raw_name = str(value or "")
                        item = QTableWidgetItem(raw_name.title())

                    elif col_idx == 2:
                        item = QTableWidgetItem("" if not value else str(value))
                    elif col_idx == 7:
                        item = NumericItem("" if not value else str(value))
                    elif col_idx == 4:
                        item = DateItem("" if not value else str(value))
                    else:
                        item = QTableWidgetItem("" if not value else str(value))

                    # ✅ Left-align the Notes column; center everything else
                    # ✅ Left-align the Notes column with a small tab space
                    if col_idx == 8:  # Notes column
                        note_text = f"   {value}" if value else ""
                        item.setText(note_text)
                        item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    else:
                        item.setTextAlignment(Qt.AlignCenter)

                    if col_idx != 0:
                        item.setBackground(row_color)

                    # Charge column ($)
                    if col_idx == 2:
                        if str(value).strip().upper() == 'Y':
                            item.setText('$')
                            item.setForeground(QBrush(QColor(255, 255, 255)) if paused else QBrush(QColor(255, 0, 0)))
                        else:
                            item.setText('')

                    self.table.setItem(row_idx, col_idx, item)

        except Exception as e:
            with open("error_log.txt", "a") as log:
                log.write(f"Error in load_data: {e}\n")
            QMessageBox.warning(self, "Load Error", f"Failed to load data:\n{e}")

        self.table.setSortingEnabled(True)

    def toggle_pause(self):
        if not self.enforce_login(): return
        selected = self.table.currentRow()
        if selected < 0:
            return
        number_item = self.table.item(selected, 1)
        if not number_item:
            return
        number = number_item.text()
        number = number_item.text()
        self.cur.execute("SELECT id, name, notes, date_packed, picked_up, due_date, given_out_by FROM patients WHERE number = ?", (number,))
        row_data = self.cur.fetchone()
        if row_data:
            patient_id = row_data[0]
            prev_data = {
                'name': row_data[1], 'notes': row_data[2], 'pack_date': row_data[3],
                'collect_date': row_data[4], 'due_date': row_data[5], 'given_by': row_data[6]
            }
            self.push_undo_action(patient_id, prev_data)
        self.cur.execute("UPDATE patients SET paused = NOT paused WHERE number = ?", (number,))
        self.conn.commit()
        self.load_data()
        self.update_button_states()
        self.table.selectRow(selected)

    def toggle_flag(self):
        if not self.enforce_login(): return
        selected = self.table.currentRow()
        if selected < 0:
            return
        number_item = self.table.item(selected, 1)
        if not number_item:
            return
        number = number_item.text()
        number = number_item.text()
        self.cur.execute("SELECT id, name, notes, date_packed, picked_up, due_date, given_out_by FROM patients WHERE number = ?", (number,))
        row_data = self.cur.fetchone()
        if row_data:
            patient_id = row_data[0]
            prev_data = {
                'name': row_data[1], 'notes': row_data[2], 'pack_date': row_data[3],
                'collect_date': row_data[4], 'due_date': row_data[5], 'given_by': row_data[6]
            }
            self.push_undo_action(patient_id, prev_data)
        self.cur.execute("UPDATE patients SET flagged = NOT flagged WHERE number = ?", (number,))
        self.conn.commit()
        self.load_data()
        self.update_button_states()
        self.table.selectRow(selected)


    def load_ceased_data(self):
        showing_ceased = self.view_ceased_action.isChecked()
        # Rebuild top button contents only
        while self.top_buttons.count():
            item = self.top_buttons.takeAt(0)
            if item.widget(): item.widget().setParent(None)
        if showing_ceased:
            self.top_buttons.addWidget(self.delete_button)
            self.top_buttons.addWidget(self.restore_button)
            self.top_buttons.addWidget(self.login_input)
        else:
            self.top_buttons.addWidget(self.print_button)
            self.top_buttons.addWidget(self.new_packed_button)
            self.top_buttons.addWidget(self.checked_button)
            self.top_buttons.addWidget(self.collected_button)
            self.top_buttons.addWidget(self.flag_button)
            self.top_buttons.addWidget(self.pause_button)
            self.top_buttons.addWidget(self.login_input)
        self.load_data(ceased=showing_ceased)

    def _update_date_field_style(self, field):
        """Grey out visually (and optionally make read-only) when blank/masked."""
        text = field.text().strip()
        if not text or text == "__/__/____":
            # force the colour, overriding parent palette
            field.setStyleSheet("""
                QLineEdit {
                    background-color: #e6e6e6;
                    color: #555555;
                    border: 1px solid #c0c0c0;
                }
            """)
            field.setReadOnly(False)  # leave True if you want it uneditable
        else:
            field.setStyleSheet("")  # reset to normal
            field.setReadOnly(False)

    def open_patient_tab(self, row, col):
        if not self.enforce_login(): return
        number_item = self.table.item(row, 1)
        name_item = self.table.item(row, 3)
        if not number_item or not name_item:
            return
        number = number_item.text().strip()
        name_to_open = name_item.text().strip()
        for i in range(self.tabs.count()):
            tab = self.tabs.widget(i)
            if tab.property("patient_number") == number:
                self.tabs.setCurrentIndex(i)
                return

        number = self.table.item(row, 1).text()
        name_to_open = self.table.item(row, 3).text().strip()
        for i in range(self.tabs.count()):
            existing_tab = self.tabs.widget(i)
            if existing_tab.property("patient_number") == number:
                self.tabs.setCurrentIndex(i)
                return
        number = self.table.item(row, 1).text()
        self.current_patient_number = number
        self.cur.execute("""
                         SELECT name,
                                charge,
                                date_packed,
                                picked_up,
                                notes,
                                pack_size,
                                partial_supply,
                                weeks_per_blister,
                                given_out_by,
                                medicare,
                                concession
                         FROM patients
                         WHERE number = ?
                         """, (number,))
        result = self.cur.fetchone()
        name, charge, date_packed, picked_up, notes, pack_size, partial_supply_val, weeks_val, given_out_by, medicare_val, concession_val = result
        print("DEBUG: Retrieved weeks_per_blister =", repr(weeks_val))
        self.weeks_per_blister_dropdown = QComboBox()
        self.weeks_per_blister_dropdown.addItems(["1 week", "2 weeks", "4 weeks"])
        try:
            if weeks_val in ["1 week", "2 weeks", "4 weeks"]:
                self.weeks_per_blister_dropdown.setCurrentText(weeks_val)
                print("DEBUG: Set weeks dropdown to", self.weeks_per_blister_dropdown.currentText())
            else:
                print("DEBUG: Invalid or missing weeks_val:", repr(weeks_val))
        except Exception as e:
            print("DEBUG: Exception setting weeks dropdown:", e)
        original_notes = notes  # Ensure we capture legacy note early
        # Name stored as "Last, First" → parse correctly
        parts = [p.strip() for p in name.split(",", 1)]
        last = parts[0] if parts else ""
        first = parts[1] if len(parts) > 1 else ""

        self.current_patient_name = name
        patient_page = QWidget()
        patient_layout = QVBoxLayout()
        row_layout = QHBoxLayout()

        groupbox = QGroupBox("Patient Information")
        groupbox.setFixedHeight(350)
        groupbox.setStyleSheet("QGroupBox { border: 2px solid black; margin-top: 10px; }")
        grid = QGridLayout()
        grid.addWidget(QLabel("First Names:"), 0, 0)
        self.first_name_input = QLineEdit()
        self.first_name_input.setText(first)
        self.first_name_input.setPlaceholderText("First Names")
        grid.addWidget(self.first_name_input, 0, 1)
        grid.addWidget(QLabel("Blister Size:"), 0, 2)
        self.pack_size_dropdown = QComboBox()
        self.pack_size_dropdown.addItems(["Small", "Large"])
        if pack_size in ["Small", "Large"]:
            self.pack_size_dropdown.setCurrentText(pack_size)
        grid.addWidget(self.pack_size_dropdown, 0, 3)
        grid.addWidget(QLabel("Last Name:"), 1, 0)
        self.last_name_input = QLineEdit()
        self.last_name_input.setText(last)
        self.last_name_input.setPlaceholderText("Last Name")
        grid.addWidget(self.last_name_input, 1, 1)
        grid.addWidget(QLabel("Weeks/Blister:"), 1, 2)
        grid.addWidget(self.weeks_per_blister_dropdown, 1, 3)
        grid.addWidget(QLabel("Medicare:"), 2, 2)
        self.medicare_input = QLineEdit()
        self.checked_by_input = QLineEdit()
        grid.addWidget(self.medicare_input, 2, 3)
        self.medicare_input.setText(medicare_val or "")
        # --- Uniform toggle buttons (Claim / Charge) ---
        toggle_style = """
            QPushButton {
                background-color: #d9d9d9;
                border: 1px solid #999;
                border-radius: 4px;
                padding: 4px 12px;
            }
            QPushButton:checked {
                color: white;
                font-weight: bold;
            }
        """

        # Claim (green active)
        self.claim_button = QPushButton("Claim")
        self.claim_button.setCheckable(True)
        self.claim_button.setFixedWidth(100)  # ✅ match Charge button width
        self.claim_button.setChecked(partial_supply_val == 'Y')
        self.claim_button.setStyleSheet(toggle_style + "QPushButton:checked { background-color: #4CAF50; }")
        grid.addWidget(self.claim_button, 2, 0, alignment=Qt.AlignRight)

        # Charge (red active)
        self.charge_button = QPushButton("Charge")
        self.charge_button.setCheckable(True)
        self.charge_button.setFixedWidth(100)
        self.charge_button.setStyleSheet(toggle_style + "QPushButton:checked { background-color: #FF6666; }")
        self.charge_button.setChecked(charge and str(charge).strip().upper() == "Y")
        grid.addWidget(self.charge_button, 2, 1)

        def update_claim_fields_state():
            if not hasattr(self, "medicare_input") or not hasattr(self, "concession_input"):
                return
            if self.claim_button.isChecked():
                self.medicare_input.setText(medicare_val or "")
                self.concession_input.setText(concession_val or "")
                self.medicare_input.setEnabled(True)
                self.concession_input.setEnabled(True)
                self.medicare_input.setStyleSheet("")
                self.concession_input.setStyleSheet("")
            else:
                # Disable the fields but DO NOT clear them — keep values stored
                self.medicare_input.setEnabled(False)
                self.concession_input.setEnabled(False)
                self.medicare_input.setStyleSheet("background-color: #e6e6e6; color: #666;")
                self.concession_input.setStyleSheet("background-color: #e6e6e6; color: #666;")

        grid.addWidget(QLabel("Date Packed:"), 3, 0)
        self.packed_input = QLineEdit()
        self.packed_input.setInputMask("00/00/0000;_")
        self.packed_input.setPlaceholderText("dd/MM/yyyy")
        self.packed_input.setText(date_packed or "")
        self.packed_input.textChanged.connect(lambda: self._update_date_field_style(self.packed_input))
        self._update_date_field_style(self.packed_input)  # set initial colour
        grid.addWidget(self.packed_input, 3, 1)
        grid.addWidget(QLabel("Concession:"), 3, 2)
        self.concession_input = QLineEdit()
        self.packed_by_input = QLineEdit()

        def _validate_card_fields():
            med = self.medicare_input.text().replace(" ", "").strip().upper()
            conc = self.concession_input.text().replace(" ", "").strip().upper()

            # --- Medicare formatting (11 digits -> 1234 56789 1 2) ---
            if med:
                if re.fullmatch(r"\d{11}", med):
                    formatted = f"{med[:4]} {med[4:9]} {med[9:10]} {med[10:]}"
                    cursor_pos = self.medicare_input.cursorPosition()
                    self.medicare_input.blockSignals(True)
                    self.medicare_input.setText(formatted)
                    self.medicare_input.blockSignals(False)
                    self.medicare_input.setCursorPosition(cursor_pos)
                self.medicare_input.setStyleSheet("" if self.is_valid_medicare(med) else "border: 2px solid red;")
            else:
                self.medicare_input.setStyleSheet("")

            # --- Concession / DVA formatting ---
            if conc:
                if re.fullmatch(r"\d{9}[A-Z]", conc):
                    # Concession: 111 111 111A
                    formatted = f"{conc[:3]} {conc[3:6]} {conc[6:]}"
                    self.dva_label.setText("")
                elif re.fullmatch(r"QSM\d{5}", conc):
                    # DVA: QSM 12345
                    formatted = f"{conc[:3]} {conc[3:]}"
                    self.dva_label.setText("(DVA)")
                elif re.fullmatch(r"[A-Z]{4}\d{4}", conc):
                    # DVA: NBUR 9080
                    formatted = f"{conc[:4]} {conc[4:]}"
                    self.dva_label.setText("(DVA)")
                else:
                    formatted = conc
                    self.dva_label.setText("")

                cursor_pos = self.concession_input.cursorPosition()
                self.concession_input.blockSignals(True)
                self.concession_input.setText(formatted)
                self.concession_input.blockSignals(False)
                self.concession_input.setCursorPosition(cursor_pos)

                self.concession_input.setStyleSheet("" if self.is_valid_concession(conc) else "border: 2px solid red;")
            else:
                self.concession_input.setStyleSheet("")
                self.dva_label.setText("")

        # Connect validation triggers
        self.medicare_input.textChanged.connect(_validate_card_fields)
        self.concession_input.textChanged.connect(_validate_card_fields)

        grid.addWidget(self.concession_input, 3, 3)
        self.dva_label = QLabel("")
        self.dva_label.setStyleSheet("color: #555; font-weight: bold;")
        grid.addWidget(self.dva_label, 3, 4)
        self.concession_input.setText(concession_val or "")
        self.claim_button.toggled.connect(update_claim_fields_state)
        update_claim_fields_state()  # apply on load

        if date_packed and date_packed.strip():
            try:
                xl_path = os.path.join(COLLECTION_LOGS_DIR, f"{name_to_open.upper().replace(',', '').strip()}.xlsx")
                if os.path.exists(xl_path):
                    wb = openpyxl.load_workbook(xl_path, data_only=True)
                    ws = wb.active
                    for col in range(2, ws.max_column + 1, 6):
                        row = 4
                        row = 4
                        while True:
                            val = ws.cell(row=row, column=col).value
                            if val is None or str(val).strip() == '':
                                break
                            if isinstance(val, (int, float)):
                                cell_date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(val) - 2).strftime("%d/%m/%Y")
                            elif isinstance(val, datetime):
                                cell_date = val.strftime("%d/%m/%Y")
                            elif isinstance(val, str):
                                cell_date = val.strip()
                            else:
                                row += 1
                                continue
                            if cell_date == date_packed.strip():
                                initials = ""
                                r = 4
                                while True:
                                    val = ws.cell(row=r, column=col + 4).value
                                    if val is not None and str(val).strip():
                                        initials = str(val)
                                        break
                                    r += 1
                                break
                            row += 1
                        break
            except Exception as e:
                print("DEBUG: Failed to extract initials from Excel:", e)

        grid.addWidget(QLabel("Date Collected:"), 4, 0)
        self.picked_input = QLineEdit()
        self.picked_input.setInputMask("00/00/0000;_")
        self.picked_input.setPlaceholderText("dd/MM/yyyy")
        self.picked_input.setText(picked_up or "")
        self.picked_input.textChanged.connect(lambda: self._update_date_field_style(self.picked_input))
        self._update_date_field_style(self.picked_input)
        # Disable either field depending on which is filled
        self._update_packed_collected_lock()
        # Keep it in sync while user types
        self.packed_input.textChanged.connect(self._update_packed_collected_lock)
        self.picked_input.textChanged.connect(self._update_packed_collected_lock)

        grid.addWidget(self.picked_input, 4, 1)
        grid.addWidget(QLabel(" "), 4, 2)
        self.given_by_input = QLineEdit()
        self.given_by_input.setText(given_out_by or "")
        grid.addWidget(self.given_by_input, 4, 3)
        # Hide backend-linked fields from UI
        self.checked_by_input.hide()
        self.packed_by_input.hide()
        self.given_by_input.hide()

        groupbox.setLayout(grid)

        # Load latest pack entry from Excel log
        outer_left_layout = QVBoxLayout()
        outer_left_layout.addWidget(groupbox)
        # --- Add Edit Drug List button below Patient Information ---
        edit_drugs_btn = QPushButton("Edit Active Drugs")
        edit_drugs_btn.setStyleSheet("font-weight: bold;")
        def handle_edit_drugs():
            xl_path = os.path.join(COLLECTION_LOGS_DIR, f"{name_to_open.upper().replace(',', '').strip()}.xlsx")
            self.excel_path = xl_path
            self.current_patient_id = pid
            self.current_med_names = []
            if os.path.exists(xl_path):
                try:
                    wb = openpyxl.load_workbook(xl_path)
                    ws = wb.active
                    r = 4
                    while True:
                        val = ws.cell(row=r, column=1).value
                        if not val:
                            break
                        self.current_med_names.append(val)
                        r += 1
                except Exception as e:
                    print("DEBUG: Failed to load med names:", e)
            self.blank_pack_window = self  # placeholder parent
            self.open_edit_drugs_dialog()
        edit_drugs_btn.clicked.connect(handle_edit_drugs)
        outer_left_layout.addWidget(edit_drugs_btn)

        left = outer_left_layout
        # --- Injected Pack Viewer (Final Clean Build) ---
        self.pack_tables = []
        self.current_pack_index = 0
        self.pack_table_container = QVBoxLayout()
        # --- Resolve Excel log path robustly ---
        safe_name = name.upper().replace(",", "").strip()
        xl_path = os.path.join(COLLECTION_LOGS_DIR, f"{safe_name}.xlsx")

        # Try both LAST FIRST and FIRST LAST file orders
        if not os.path.exists(xl_path):
            parts = safe_name.split()
            if len(parts) == 2:
                alt_name = f"{parts[1]} {parts[0]}"
                alt_path = os.path.join(COLLECTION_LOGS_DIR, f"{alt_name}.xlsx")
                if os.path.exists(alt_path):
                    xl_path = alt_path

        print(f"DEBUG: Using Excel path: {xl_path}")
        if os.path.exists(xl_path):

            try:
                wb = openpyxl.load_workbook(xl_path, data_only=True)
                ws = wb.active
                med_names = []
                r = 4
                while True:
                    val = ws.cell(row=r, column=1).value
                    if val is None or str(val).strip() == '':
                        break
                    med_names.append(val)
                    r += 1
                total_cols = ws.max_column
                for block_start in reversed(range(2, total_cols + 1, 6)):
                    # Get pack date
                    date_val = None

                    try:
                        if isinstance(date_val, (int, float)):
                            pack_date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_val) - 2).strftime("%d/%m/%Y")
                        elif isinstance(date_val, datetime):
                            pack_date = date_val.strftime("%d/%m/%Y")
                        elif isinstance(date_val, str) and date_val.strip().isdigit():
                            pack_date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_val.strip()) - 2).strftime("%d/%m/%Y")
                        else:
                            pack_date = datetime.strptime(str(date_val).strip(), "%d/%m/%Y").strftime("%d/%m/%Y")
                    except:
                        pack_date = str(date_val)
                    for r in range(4, 20):
                        val = ws.cell(row=r, column=block_start).value
                        if val:
                            date_val = val
                            break
                    try:
                        if isinstance(date_val, (int, float)):
                            pack_date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_val) - 2).strftime("%d/%m/%Y")
                        elif isinstance(date_val, datetime):
                            pack_date = date_val.strftime("%d/%m/%Y")
                        else:
                            pack_date = str(date_val)
                    except:
                        pack_date = str(date_val)
                    # Get initials
                    initials = ""
                    for r in range(4, 20):
                        val = ws.cell(row=r, column=block_start + 4).value
                        if val:
                            initials = str(val)
                            break
                    # Build table
                    table = QTableWidget()
                    table.setColumnCount(6)
                    table.setHorizontalHeaderLabels(["Medication", "QTY Dispensed", "Batch", "Expiry", "QTY Packed", "QTY Remaining"])
                    header = table.horizontalHeader()
                    for i in range(6):
                        header.setSectionResizeMode(i, QHeaderView.Stretch)
                    table.setEditTriggers(QTableWidget.NoEditTriggers)
                    table.setRowCount(0)
                    display_row = 0
                    for row_idx, med in enumerate(med_names):
                        qty_packed = ws.cell(row=row_idx + 4, column=block_start + 3).value
                        if not qty_packed or str(qty_packed).strip() in ('', '0'):
                            pass
                        table.insertRow(display_row)
                        table.setItem(display_row, 0, QTableWidgetItem(str(med) if med else ""))
                        qty = ws.cell(row=row_idx + 4, column=block_start + 1).value
                        batch_exp = ws.cell(row=row_idx + 4, column=block_start + 2).value
                        qty_remaining = ws.cell(row=row_idx + 4, column=block_start + 5).value
                        batch, expiry = parse_batch_expiry(batch_exp)
                        qty_remaining = str(0) if qty_remaining in (None, '') else str(qty_remaining)
                        values = [qty, batch.upper(), expiry, qty_packed, qty_remaining]
                        for col_offset, val in enumerate(values):
                            item = QTableWidgetItem(str(val) if val else "")
                            item.setTextAlignment(Qt.AlignCenter)
                            table.setItem(display_row, col_offset + 1, item)
                        display_row += 1
                        qty_packed = ws.cell(row=row_idx + 4, column=block_start + 3).value
                        if not qty_packed or str(qty_packed).strip() in ('', '0'):
                            pass
                        qty = ws.cell(row=row_idx + 4, column=block_start + 1).value
                        batch_exp = ws.cell(row=row_idx + 4, column=block_start + 2).value
                        qty_packed = ws.cell(row=row_idx + 4, column=block_start + 3).value
                        qty_remaining = ws.cell(row=row_idx + 4, column=block_start + 5).value
                        batch, expiry = parse_batch_expiry(batch_exp)
                        values = [qty, batch.upper(), expiry, qty_packed, qty_remaining]
                        for col_offset, val in enumerate(values):
                            item = QTableWidgetItem(str(val) if val else "")
                            item.setTextAlignment(Qt.AlignCenter)
                    table.setVisible(False)
                    if date_val is not None and str(date_val).strip().lower() != "none":
                        try:
                            if isinstance(date_val, (int, float)):
                                label = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_val) - 2).strftime("%d/%m/%Y")
                            elif isinstance(date_val, datetime):
                                label = date_val.strftime("%d/%m/%Y")
                            elif isinstance(date_val, str) and date_val.strip().isdigit():
                                label = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_val.strip()) - 2).strftime("%d/%m/%Y")
                            else:
                                label = datetime.strptime(str(date_val).strip(), "%d/%m/%Y").strftime("%d/%m/%Y")
                        except:
                            label = str(date_val)
                        self.pack_tables.append((label, initials, table))
                        self.pack_table_container.addWidget(table)
            except Exception as e:
                print("PACK TABLE ERROR:", e)
        self.pack_header = QLabel()
        self.pack_header.setAlignment(Qt.AlignCenter)
        self.edit_btn = QPushButton("Edit")
        self.edit_btn.clicked.connect(self.edit_current_pack_entry)
        self.left_btn = QPushButton("←")
        self.right_btn = QPushButton("→")

        def show_pack(index: int):
            """Display the given pack entry and enable/disable arrows appropriately."""
            if not self.pack_tables:
                self.left_btn.setDisabled(True)
                self.right_btn.setDisabled(True)
                self.pack_header.setText("No Pack Entries Found")
                return

            # Clamp index
            index = max(0, min(index, len(self.pack_tables) - 1))
            self.current_pack_index = index

            # Show only the selected pack table
            for i, (_, _, tbl) in enumerate(self.pack_tables):
                tbl.setVisible(i == index)

            # Extract label and initials
            raw_label, initials, _ = self.pack_tables[index]
            try:
                val = str(raw_label).strip()
                if val.isdigit():
                    pack_label = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(val) - 2).strftime(
                        "%d/%m/%Y")
                elif "/" in val:
                    pack_label = datetime.strptime(val, "%d/%m/%Y").strftime("%d/%m/%Y")
                else:
                    pack_label = val
            except Exception:
                pack_label = str(raw_label).strip()

            # Header text
            if isinstance(initials, str) and "/" in initials:
                packed_by, checked_by = initials.split("/", 1)
                self.pack_header.setText(f"Pack Entry: {pack_label} (Packed by: {packed_by}, Checked by: {checked_by})")
            else:
                self.pack_header.setText(f"Pack Entry: {pack_label} (Packed: {initials})")

            # Disable arrows at edges (reversed order: index 0 =newest, last = oldest)
            self.left_btn.setDisabled(index == len(self.pack_tables) - 1)  # grey out ← when oldest reached
            self.right_btn.setDisabled(index == 0)  # grey out → when newest reached

        def go_left():
            """← Go back in time (older pack entry)."""
            if self.current_pack_index < len(self.pack_tables) - 1:
                self.current_pack_index += 1
                show_pack(self.current_pack_index)

        def go_right():
            """→ Go forward in time (newer pack entry)."""
            if self.current_pack_index > 0:
                self.current_pack_index -= 1
                show_pack(self.current_pack_index)

        # --- Correct connections ---
        self.left_btn.clicked.connect(go_left)
        self.right_btn.clicked.connect(go_right)

        nav = QHBoxLayout()
        nav.addWidget(self.left_btn)
        nav.addWidget(self.edit_btn)
        nav.addStretch()
        nav.addWidget(self.pack_header)
        nav.addStretch()
        nav.addWidget(self.right_btn)
        wrapper = QVBoxLayout()
        wrapper.addLayout(nav)
        wrapper.addLayout(self.pack_table_container)
        pack_box = QGroupBox("Pack Entries")
        pack_box.setLayout(wrapper)
        left.addWidget(pack_box)
        QTimer.singleShot(200, lambda: show_pack(self.current_pack_index))
        # --- End Injected ---


        right = QVBoxLayout()
        right.setAlignment(Qt.AlignTop)
        self.add_note_label = QLabel("Add Note:")
        self.notes_input = QLineEdit()
        self.notes_input.setPlaceholderText("Patient notes")
        self.notes_input.setText("")  # Leave blank on open
        self.notes_input.setPlaceholderText("Patient notes")

        self.notes_input.returnPressed.connect(lambda: self.add_note_to_log(pid))

        self.notes_input.returnPressed.connect(lambda: self.add_note_to_log(pid))

                # Load and display previous notes from log
        note_groupbox = QGroupBox("Patient Notes")
        note_groupbox.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        note_groupbox.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        note_groupbox.setStyleSheet("QGroupBox { border: 2px solid black; margin-top: 10px; }")
        self.cur.execute("SELECT id FROM patients WHERE number = ?", (number,))
        pid = self.cur.fetchone()[0]
        self.current_patient_id = pid
        note_log_layout = QVBoxLayout()
        note_log_layout.setAlignment(Qt.AlignTop)
        note_groupbox.setLayout(note_log_layout)
        add_note_label = QLabel("Add Note:")
        notes_input = QLineEdit()  # inside groupbox only
        notes_input.setPlaceholderText("Patient notes")
        notes_input.returnPressed.connect(lambda: self.add_note_to_log(pid))
        self.notes_input = notes_input
        note_log_layout.addWidget(add_note_label)
        note_log_layout.addWidget(notes_input)
        note_log_layout.addSpacing(10)
        note_log_layout.addWidget(QLabel("Notes:"))
        self.note_log_container = QWidget()
        self.note_log_layout = QVBoxLayout(self.note_log_container)
        self.note_log_layout.setAlignment(Qt.AlignTop)
        self.note_log_layout.setAlignment(Qt.AlignTop)

        self.cur.execute("SELECT note FROM notes_log WHERE patient_id = ?", (pid,))
        existing_notes = [row[0] for row in self.cur.fetchall()]
        if original_notes.strip() and original_notes.strip() not in existing_notes:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.cur.execute("INSERT INTO notes_log (patient_id, note, timestamp) VALUES (?, ?, ?)", (pid, original_notes.strip(), now))
            self.conn.commit()
        self.cur.execute("SELECT note FROM notes_log WHERE patient_id = ?", (pid,))
        existing_notes = [row[0] for row in self.cur.fetchall()]
        if original_notes.strip() and original_notes.strip() not in existing_notes:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.cur.execute("INSERT INTO notes_log (patient_id, note, timestamp) VALUES (?, ?, ?)", (pid, original_notes.strip(), now))
            self.conn.commit()
        self.load_note_log(pid)
        self.cur.execute("SELECT note FROM notes_log WHERE patient_id = ? ORDER BY timestamp DESC LIMIT 1", (pid,))
        latest_note = self.cur.fetchone()
        if latest_note:
            notes = latest_note[0] if latest_note[0] is not None else ""
        self.cur.execute("SELECT note FROM notes_log WHERE patient_id = ?", (pid,))
        existing_notes = [row[0] for row in self.cur.fetchall()]
        if original_notes.strip() and original_notes.strip() not in existing_notes:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.cur.execute("INSERT INTO notes_log (patient_id, note, timestamp) VALUES (?, ?, ?)", (pid, original_notes.strip(), now))
            self.conn.commit()
        self.cur.execute("SELECT note FROM notes_log WHERE patient_id = ?", (pid,))
        existing_notes = [row[0] for row in self.cur.fetchall()]
        if notes.strip() and notes.strip() not in existing_notes:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.cur.execute("INSERT INTO notes_log (patient_id, note, timestamp) VALUES (?, ?, ?)", (pid, notes.strip(), now))
            self.conn.commit()
        # Inject legacy notes into notes_log if missing
        self.cur.execute("SELECT COUNT(*) FROM notes_log WHERE patient_id = ?", (pid,))
        if self.cur.fetchone()[0] == 0 and notes.strip():
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.cur.execute("INSERT INTO notes_log (patient_id, note, timestamp) VALUES (?, ?, ?)", (pid, notes.strip(), now))
            self.conn.commit()

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setWidget(self.note_log_container)

        note_log_layout.addWidget(scroll)
        right.addWidget(note_groupbox)  # patched

        # Replace save_and_close to add note to log
        def save_and_close():
            nonlocal medicare_val, concession_val  # keep latest values across toggles
            self.save_allowed = True

            # --- Collect UI values ---
            first = (self.first_name_input.text() or "").strip()
            last = (self.last_name_input.text() or "").strip()
            name_to_save = f"{last}, {first}"

            pack_size_val = self.pack_size_dropdown.currentText().strip()
            weeks_sel = self.weeks_per_blister_dropdown.currentText().strip()

            # Charge toggle => 'Y'/'N'
            charge_val = 'Y' if self.charge_button.isChecked() else 'N'

            # Claim/Partial Supply toggle => 'Y'/'N'
            claim_val = 'Y' if self.claim_button.isChecked() else 'N'

            # Dates (stored as dd/MM/yyyy or blank)
            packed_txt = (self.packed_input.text() or "").strip().replace("_", "")
            if not self.is_valid_date(packed_txt):
                packed_txt = ""
            picked_txt = (self.picked_input.text() or "").strip().replace("_", "")
            if not self.is_valid_date(picked_txt):
                picked_txt = ""

            # Card fields
            medicare_val = (self.medicare_input.text() or "").strip()
            concession_val = (self.concession_input.text() or "").strip()

            # If not claiming, clear card fields
            #if claim_val == 'N':
             #   medicare_val = ""
              #  concession_val = ""

            # Notes (only the “add note” box on the right – legacy notes go via notes_log elsewhere)
            notes_current = (self.notes_input.text() or "").strip()

            # --- Validate (only if fields are non-empty) ---
            invalid_fields = []
            raw_med = medicare_val.replace(" ", "").upper()
            raw_conc = concession_val.replace(" ", "").upper()

            if raw_med and not self.is_valid_medicare(raw_med):
                self.medicare_input.setStyleSheet("border: 2px solid red;")
                invalid_fields.append("Medicare number")
            else:
                self.medicare_input.setStyleSheet("")

            if raw_conc and not self.is_valid_concession(raw_conc):
                self.concession_input.setStyleSheet("border: 2px solid red;")
                invalid_fields.append("Concession or DVA number")
            else:
                self.concession_input.setStyleSheet("")

            if invalid_fields:
                QMessageBox.warning(
                    self,
                    "Invalid Card Number",
                    "The following field(s) are invalid:\n\n• " + "\n• ".join(invalid_fields) +
                    "\n\nPlease correct these before saving and closing."
                )
                if "Medicare number" in invalid_fields:
                    self.medicare_input.setFocus()
                elif "Concession or DVA number" in invalid_fields:
                    self.concession_input.setFocus()
                self.save_allowed = False
                return  # abort save

            # --- Persist to DB ---
            try:
                # 1) Update patients table
                self.cur.execute("""
                                 UPDATE patients
                                 SET name=?,
                                     charge=?,
                                     date_packed=?,
                                     picked_up=?,
                                     notes=?,
                                     pack_size=?,
                                     partial_supply=?,
                                     weeks_per_blister=?,
                                     given_out_by=?,
                                     medicare=?,
                                     concession=?
                                 WHERE number = ?
                                 """, (
                                     name_to_save,
                                     charge_val,
                                     packed_txt,
                                     picked_txt,
                                     notes_current,  # keep simple notes field in sync (log is handled elsewhere)
                                     pack_size_val,
                                     claim_val,
                                     weeks_sel,
                                     (self.given_by_input.text() or "").strip(),
                                     medicare_val,
                                     concession_val,
                                     number
                                 ))
                self.conn.commit()

                # 2) Optional: append “add note” to notes_log if non-empty
                if notes_current:
                    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    self.cur.execute("INSERT INTO notes_log (patient_id, note, timestamp) VALUES (?, ?, ?)",
                                     (pid, notes_current, now))
                    self.conn.commit()
                    # Clear the add-note box after saving to log
                    self.notes_input.clear()
                    self.load_note_log(pid)

                # (Optional) refresh main table row if you have a helper; otherwise leave as-is
                if hasattr(self, "reload_calendar_view"):
                    self.reload_calendar_view()

            except Exception as e:
                QMessageBox.critical(self, "Save Failed", f"An error occurred while saving:\n\n{e}")
                self.save_allowed = False
                return

            # success — allow close
            self.save_allowed = True

        patient_page.setLayout(patient_layout)
        self.tabs.addTab(patient_page, name)

        # Create tab header with close button
        tab_index = self.tabs.indexOf(patient_page)
        tab_widget = QWidget()
        tab_layout = QHBoxLayout()
        tab_layout.setContentsMargins(0, 0, 0, 0)
        tab_label = QLabel(name)
        close_btn = QPushButton("❌")
        close_btn.setFixedSize(24, 24)
        close_btn.setStyleSheet("QPushButton { border: none; }")
        def try_close_tab():
            unsaved = (
                self.first_name_input.isModified() or
                self.last_name_input.isModified() or
                self.notes_input.isModified() or
                self.packed_input.isModified() or
                self.picked_input.isModified()
            )
            if unsaved:
                reply = QMessageBox.question(self, "Unsaved Changes",
                                             "Do you want to save changes before closing?",
                                             QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel)
                if reply == QMessageBox.Cancel:
                    return
                if reply == QMessageBox.Yes:
                    save_and_close()
                    return
            self.tabs.removeTab(self.tabs.indexOf(patient_page))
        close_btn.clicked.connect(try_close_tab)
        tab_layout.addWidget(tab_label)
        tab_layout.addWidget(close_btn)
        tab_widget.setLayout(tab_layout)
        self.tabs.setTabText(tab_index, name)
        patient_page.setProperty("patient_number", number)
        splitter = QSplitter(Qt.Horizontal)
        left_widget = QWidget()
        left_widget.setLayout(left)
        right_widget_wrapper = QWidget()
        right_widget_wrapper.setLayout(right)
        splitter.addWidget(left_widget)
        splitter.addWidget(right_widget_wrapper)
        splitter.setStretchFactor(0, 2)
        splitter.setStretchFactor(1, 1)
        patient_layout.addWidget(splitter)
        bottom = QHBoxLayout()
        bottom.setSpacing(10)
        self.patient_login_input = QLineEdit()
        self.patient_login_input.setPlaceholderText('Login')
        self.patient_login_input.setFixedWidth(990)
        self.patient_login_input.setText(self.login_input.text())
        self.patient_login_input.installEventFilter(self)
        self.patient_login_input.mousePressEvent = self.login_click_select_all
        self.patient_login_input.textChanged.connect(self.clear_active_user_if_empty)
        self.patient_login_input.returnPressed.connect(self.assign_active_user)
        bottom.addWidget(self.patient_login_input)
        close_btn = QPushButton('Save + Close')

        def handle_save_close():
            save_and_close()
            # Only close if validation passed (save_allowed True)
            if getattr(self, "save_allowed", True):
                idx = self.tabs.indexOf(patient_page)
                if idx != -1:
                    self.tabs.removeTab(idx)

        close_btn.clicked.connect(handle_save_close)

        bottom.addWidget(close_btn)
        patient_layout.addLayout(bottom)
        self.tabs.setCurrentWidget(patient_page)
        right_widget_wrapper.updateGeometry()
        QTimer.singleShot(100, lambda: self.tabs.currentWidget().layout().update())
    def load_note_log(self, patient_id):
        for i in reversed(range(self.note_log_layout.count())):
            self.note_log_layout.itemAt(i).widget().setParent(None)

        self.cur.execute("SELECT id, note, timestamp FROM notes_log WHERE patient_id = ? ORDER BY timestamp DESC", (patient_id,))
        for note_id, note_text, ts in self.cur.fetchall():
            note_line = QHBoxLayout()
            note_label = QTextEdit()
            note_label.setPlainText(f"{ts}: {note_text}")
            note_label.setReadOnly(True)
            doc = note_label.document()
            doc.setTextWidth(note_label.viewport().width())
            height = int(doc.size().height()) + 10
            note_label.setFixedHeight(height)
            note_label.setStyleSheet("background-color: #f9f9f9; border: 1px solid #ccc;")
            delete_btn = QPushButton("❌")
            star_btn = QPushButton("⭐")
            star_btn.setFixedWidth(30)

            def make_set_main_note(text, pid):
                def set_note():
                    self.cur.execute("UPDATE patients SET notes = ? WHERE id = ?", (text, pid))
                    self.notes_input.setText(text)
                    self.conn.commit()
                    self.load_data()
                return set_note

            star_btn.clicked.connect(make_set_main_note(note_text, patient_id))
            delete_btn.setFixedWidth(30)

            def make_delete_note(nid):
                def delete_note():
                    self.cur.execute("DELETE FROM notes_log WHERE id = ?", (nid,))
                    self.conn.commit()
                    self.cur.execute("SELECT note FROM notes_log WHERE patient_id = ? ORDER BY timestamp DESC LIMIT 1", (patient_id,))
                    latest_note = self.cur.fetchone()
                    self.cur.execute("UPDATE patients SET notes = ? WHERE id = ?", ((latest_note[0] if latest_note else ""), patient_id))
                    self.conn.commit()
                    self.load_note_log(patient_id)
                    self.load_data()
                return delete_note

            delete_btn.clicked.connect(make_delete_note(note_id))
            note_line.addWidget(note_label)
            note_line.addWidget(delete_btn)
            note_line.addWidget(star_btn)
            wrapper = QWidget()
            wrapper.setLayout(note_line)
            wrapper.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
            self.note_log_layout.addWidget(wrapper)

    def update_field(self, field, value):
        if not hasattr(self, "note_log_layout") or not hasattr(self, "current_patient_id"):
            print(f"SKIP: Tried to update '{field}' before patient page loaded.")
            return
        patient_id = self.current_patient_id
        if not hasattr(self, "note_log_layout"):
            print(f"SKIP: Tried to update '{field}' before patient page loaded.")
            return
        try:
            self.cur.execute(f"UPDATE patients SET {field} = ? WHERE number = ?", (value, int(self.current_patient_number)))
            self.conn.commit()
        except Exception as e:
            print(f"ERROR: Failed to update {field}: {e}")
        for i in reversed(range(self.note_log_layout.count())):
            self.note_log_layout.itemAt(i).widget().setParent(None)

        self.cur.execute("SELECT id, note, timestamp FROM notes_log WHERE patient_id = ? ORDER BY timestamp DESC", (patient_id,))
        for note_id, note_text, ts in self.cur.fetchall():
            note_line = QHBoxLayout()
            note_label = QTextEdit()
            note_label.setPlainText(f"{ts}: {note_text}")
            note_label.setReadOnly(True)
            doc = note_label.document()
            doc.setTextWidth(note_label.viewport().width())
            height = int(doc.size().height()) + 10
            note_label.setFixedHeight(height)
            note_label.setStyleSheet("background-color: #f9f9f9; border: 1px solid #ccc;")
            delete_btn = QPushButton("❌")
            star_btn = QPushButton("⭐")
            star_btn.setFixedWidth(30)
            def make_set_main_note(text, pid):
                def set_note():
                    self.cur.execute("UPDATE patients SET notes = ? WHERE id = ?", (text, pid))
                    self.notes_input.setText(text)
                    self.conn.commit()
                    self.load_data()
                return set_note
            star_btn.clicked.connect(make_set_main_note(note_text, patient_id))
            delete_btn.setFixedWidth(30)

            def make_delete_note(nid):
                def delete_note():
                    self.cur.execute("DELETE FROM notes_log WHERE id = ?", (nid,))
                    self.conn.commit()
                    self.cur.execute("SELECT note FROM notes_log WHERE patient_id = ? ORDER BY timestamp DESC LIMIT 1", (patient_id,))
                    latest_note = self.cur.fetchone()
                    self.cur.execute("UPDATE patients SET notes = ? WHERE id = ?", ((latest_note[0] if latest_note else ""), patient_id))
                    self.conn.commit()
                    self.load_note_log(patient_id)
                    self.load_data()
                return delete_note

            delete_btn.clicked.connect(make_delete_note(note_id))
            note_line.addWidget(note_label)
            note_line.addWidget(delete_btn)
            note_line.addWidget(star_btn)
            wrapper = QWidget()
            wrapper.setLayout(note_line)
            wrapper.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
            self.note_log_layout.addWidget(wrapper)


    def handle_warning_click(self, row, col):
        if not self.enforce_login(): return
        item = self.table.item(row, col)
        if item and col == 0 and item.text().strip() == "⚠️":
            name_item = self.table.item(row, 3)
            if name_item:
                due_val = self.table.item(row, 7).text()
                reason = "Pack overdue" if due_val.startswith("-") else "Pack not collected"
                self.show_warning_popup(name_item.text(), reason)

    def show_warning_popup(self, name, reason):
        self.warning_popup = QWidget()
        self.warning_popup.setWindowTitle("Warning")
        layout = QVBoxLayout()
        layout.addWidget(QLabel(f"⚠️ Warning for {name}: {reason}"))
        close_btn = QPushButton("OK")
        close_btn.clicked.connect(self.warning_popup.close)
        layout.addWidget(close_btn)
        self.warning_popup.setLayout(layout)
        self.warning_popup.setFixedSize(450, 120)
        self.warning_popup.show()
    def update_button_states(self):
        selected = self.table.currentRow()
        if selected < 0:
            self.pause_button.setText("Pause")
            self.pause_button.setStyleSheet("")
            self.flag_button.setText("Changes")
            self.flag_button.setStyleSheet("")
            return

        number_item = self.table.item(selected, 1)
        if number_item:
            number = number_item.text()
            self.cur.execute("SELECT paused, flagged FROM patients WHERE number = ?", (number,))
            result = self.cur.fetchone()
            if result:
                paused, flagged = result
                if paused:
                    self.pause_button.setText("Paused")
                    self.pause_button.setStyleSheet("background-color: red; color: white;")
                else:
                    self.pause_button.setText("Pause")
                    self.pause_button.setStyleSheet("")
                if flagged:
                    self.flag_button.setText("Flagged")
                    self.flag_button.setStyleSheet("background-color: pink; color: black;")
                else:
                    self.flag_button.setText("Changes")
                    self.flag_button.setStyleSheet("")
    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.table.clearSelection()
            self.update_button_states()

    def table_mouse_click(self, row, col):
        if self.table.selectionModel().isRowSelected(row):
            self.table.clearSelection()
        else:
            self.table.selectRow(row)
        self.update_button_states()



    def add_note_to_log(self, patient_id):
        new_note = self.notes_input.text().strip()
        if new_note:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.cur.execute("INSERT INTO notes_log (patient_id, note, timestamp) VALUES (?, ?, ?)", (patient_id, new_note, now))
            self.conn.commit()

            # Get latest note and update patients.notes
            self.cur.execute("SELECT note FROM notes_log WHERE patient_id = ? ORDER BY timestamp DESC LIMIT 1", (patient_id,))
            latest_note = self.cur.fetchone()
            if latest_note:
                self.cur.execute("UPDATE patients SET notes = ? WHERE id = ?", (latest_note[0], patient_id))
                self.conn.commit()

            self.notes_input.setText("")
            self.load_note_log(patient_id)
            self.load_data()


        self.users_window = QWidget()
        self.users_window.setWindowTitle("Manage Users")
        layout = QVBoxLayout()

        button_layout = QHBoxLayout()
        add_user_btn = QPushButton("Add User")
        remove_user_btn = QPushButton("Remove User")
        button_layout.addWidget(add_user_btn)
        button_layout.addWidget(remove_user_btn)
        layout.addLayout(button_layout)

        user_table = QTableWidget()
        user_table.setColumnCount(2)
        user_table.setHorizontalHeaderLabels(["User Name", "User ID"])
        user_table.horizontalHeader().setStretchLastSection(True)
        user_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(user_table)


    def open_patients_page(self):
        print("DEBUG: Opening Patient Manager")
        self.patients_window = QWidget()
        self.patients_window.setWindowTitle("Manage Patients")
        self.patients_window.setGeometry(150, 150, 600, 400)
        layout = QVBoxLayout()
        button_layout = QHBoxLayout()
        add_patient_btn = QPushButton("Add Patient")
        cease_patient_btn = QPushButton("Cease Patient")
        button_layout.addWidget(add_patient_btn)
        button_layout.addWidget(cease_patient_btn)
        add_patient_btn.clicked.connect(self.handle_add_patient)
        cease_patient_btn.clicked.connect(self.handle_cease_patient)
        layout.addLayout(button_layout)
        self.patient_table = QTableWidget()
        self.patient_table.setColumnCount(2)
        self.patient_table.setHorizontalHeaderLabels(["Number", "Patient Name"])
        self.patient_table.horizontalHeader().setStretchLastSection(True)
        self.patient_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.patient_table.verticalHeader().setVisible(False)
        # ✅ Allow sorting by clicking table headers
        self.patient_table.setSortingEnabled(True)
        self.patient_table.horizontalHeader().setSectionsClickable(True)
        self.patient_table.horizontalHeader().setSortIndicatorShown(True)
        layout.addWidget(self.patient_table)
        self.patient_table.setRowCount(0)
        self.cur.execute("SELECT number, name FROM patients WHERE ceased = 0")
        for row_data in self.cur.fetchall():
            row = self.patient_table.rowCount()
            self.patient_table.insertRow(row)
            item = QTableWidgetItem()
            item.setData(Qt.DisplayRole, int(row_data[0]))  # ✅ numeric value for proper sorting
            item.setTextAlignment(Qt.AlignCenter)
            self.patient_table.setItem(row, 0, item)
            # ✅ Display name in sentence case regardless of how it's stored
            name = row_data[1] or ""
            display_name = name.title()
            self.patient_table.setItem(row, 1, QTableWidgetItem(display_name))
        self.patients_window.setLayout(layout)
        # ✅ Default sort by patient number ascending
        self.patient_table.sortItems(0, Qt.AscendingOrder)
        self.patient_table.horizontalHeader().setSortIndicator(0, Qt.AscendingOrder)
        self.patients_window.show()

    def open_users_page(self):
        self.users_window = QWidget()
        self.users_window.setWindowTitle("Manage Users")
        layout = QVBoxLayout()
        button_layout = QHBoxLayout()
        add_user_btn = QPushButton("Add User")
        remove_user_btn = QPushButton("Remove User")
        button_layout.addWidget(add_user_btn)
        button_layout.addWidget(remove_user_btn)
        layout.addLayout(button_layout)
        self.user_table = QTableWidget()
        self.user_table.setColumnCount(2)
        self.user_table.setHorizontalHeaderLabels(["User Name", "User ID"])
        self.user_table.horizontalHeader().setStretchLastSection(True)
        self.user_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.user_table)
        self.user_table.setRowCount(0)
        self.cur.execute("SELECT name, user_id FROM users")
        for row_data in self.cur.fetchall():
            row = self.user_table.rowCount()
            self.user_table.insertRow(row)
            self.user_table.setItem(row, 0, QTableWidgetItem(row_data[0]))
            self.user_table.setItem(row, 1, QTableWidgetItem(row_data[1]))
        def add_user():
            name, ok1 = QInputDialog.getText(self, "Enter User Name", "User Name:")
            if not ok1 or not name.strip(): return
            user_id, ok2 = QInputDialog.getText(self, "Enter User ID", "User ID:")
            if not ok2 or not user_id.strip(): return
            self.cur.execute("SELECT COUNT(*) FROM users WHERE user_id = ?", (user_id.strip(),))
            if self.cur.fetchone()[0] > 0:
                QMessageBox.warning(self.users_window, "Duplicate User ID", "A user with this ID already exists.")
                return
            row = self.user_table.rowCount()
            self.user_table.insertRow(row)
            self.user_table.setItem(row, 0, QTableWidgetItem(name.strip()))
            self.user_table.setItem(row, 1, QTableWidgetItem(user_id.strip()))
            self.cur.execute("INSERT INTO users (name, user_id) VALUES (?, ?)", (name.strip(), user_id.strip()))
            self.conn.commit()
        def remove_user():
            selected = self.user_table.currentRow()
            if selected < 0:
                QMessageBox.warning(self.users_window, "No Selection", "Please select a user to remove.")
                return
            confirm = QMessageBox.question(self.users_window, "Confirm Removal", "Remove selected user?", QMessageBox.Yes | QMessageBox.Cancel)
            user_id_item = self.user_table.item(selected, 1)
            if user_id_item:
                self.cur.execute("DELETE FROM users WHERE user_id = ?", (user_id_item.text().strip(),))
                self.conn.commit()
            if confirm == QMessageBox.Yes:
                self.user_table.removeRow(selected)
        add_user_btn.clicked.connect(add_user)
        remove_user_btn.clicked.connect(remove_user)
        self.users_window.setLayout(layout)
        self.users_window.setFixedSize(400, 300)
        self.users_window.show()
        self.users_window.setLayout(layout)
        self.users_window.setFixedSize(400, 300)
        self.users_window.show()

    def handle_delete_ceased_patient(self):
        selected = self.table.currentRow()
        if selected < 0:
            return

        number_item = self.table.item(selected, 1)  # Column 1 = patient number
        name_item = self.table.item(selected, 3)
        if not number_item or not name_item:
            return

        number = number_item.text().strip()
        name = name_item.text().strip()

        confirm = QMessageBox.question(
            self,
            "Delete Permanently",
            f"Permanently delete {name} (#{number}) from the system?",
            QMessageBox.Yes | QMessageBox.Cancel
        )
        if confirm != QMessageBox.Yes:
            return

        try:
            # Use patient number instead of name for safety
            self.cur.execute("DELETE FROM patients WHERE number = ?", (number,))
            self.conn.commit()
            print(f"DEBUG: Permanently deleted patient #{number} ({name})")
        except Exception as e:
            QMessageBox.critical(self, "Delete Error", f"Failed to delete patient:\n{e}")
            return

        self.load_data(ceased=True)

    def handle_restore_ceased_patient(self):
        selected = self.table.currentRow()
        if selected < 0:
            return
        name_item = self.table.item(selected, 3)
        if not name_item:
            return
        name = name_item.text().strip()
        self.cur.execute("UPDATE patients SET ceased = 0 WHERE name = ?", (name,))
        self.conn.commit()
        self.load_data(ceased=True)

    def assign_active_user(self):
        if hasattr(self, 'patient_login_input') and self.patient_login_input.text().strip():
            self.login_input.setText(self.patient_login_input.text())
            self.login_input.setReadOnly(True)
        entered_id = self.login_input.text().strip().lower()
        if hasattr(self, 'patient_login_input') and self.patient_login_input.text().strip():
            self.login_input.setText(self.patient_login_input.text())
            self.login_input.setReadOnly(True)
            if hasattr(self, 'patient_login_input'):
                self.patient_login_input.setText(self.active_user)
                self.patient_login_input.setReadOnly(True)
                self.enable_patient_controls()
        if hasattr(self, 'patient_login_input') and self.sender() == self.patient_login_input:
            self.login_input.setText(self.patient_login_input.text())
            self.login_input.setReadOnly(True)
        if not entered_id:
            return
        self.cur.execute("SELECT name FROM users WHERE LOWER(user_id) = ?", (entered_id,))
        result = self.cur.fetchone()
        if result:
            self.active_user = result[0]
            self.login_input.setText(self.active_user)
            if hasattr(self, 'patient_login_input'):
                self.patient_login_input.setText(self.active_user)
                self.patient_login_input.setReadOnly(True)
            self.login_input.setReadOnly(True)
        else:
            self.login_input.clear()
            self.active_user = None
            if hasattr(self, "patient_login_input"):
                self.patient_login_input.clear()
                self.patient_login_input.setReadOnly(False)
            self.login_input.setReadOnly(False)


    
    def enforce_login(self):
        if self.active_user:
            return True

        dialog = QDialog(self, flags=Qt.FramelessWindowHint)
        dialog.setModal(True)
        dialog.setFixedSize(480, 500)
        dialog.setStyleSheet("background-color: rgb(228, 224, 219);")
        layout = QVBoxLayout()

        icon_path = os.path.join(BASE_DIR, "DAACal.png")
        if os.path.exists(icon_path):
            icon_label = QLabel()
            pixmap = QPixmap(icon_path)
            icon_label.setPixmap(pixmap.scaledToHeight(400, Qt.SmoothTransformation))
            icon_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(icon_label)

        login_field = QLineEdit()
        login_field.setPlaceholderText("User ID")
        login_field.setStyleSheet("background-color: white;")
        layout.addWidget(login_field)

        def try_login():
            user_id = login_field.text().strip().lower()
            self.cur.execute("SELECT name FROM users WHERE LOWER(user_id) = ?", (user_id,))
            result = self.cur.fetchone()
            if result:
                self.active_user = result[0]
                self.login_input.setText(self.active_user)
                self.login_input.setReadOnly(True)
                if hasattr(self, "patient_login_input"):
                    self.patient_login_input.setText(self.active_user)
                    self.patient_login_input.setReadOnly(True)
                dialog.accept()
            else:
                login_field.clear()

        login_field.returnPressed.connect(try_login)
        layout.setAlignment(Qt.AlignCenter)
        dialog.setLayout(layout)
        dialog.exec_()

        return bool(self.active_user)

        dialog = QDialog(self)
        dialog.setWindowTitle("User Login")
        dialog.setModal(True)
        layout = QVBoxLayout()

        icon_path = os.path.join(BASE_DIR, "DAACal.png")
        if os.path.exists(icon_path):
            icon_label = QLabel()
            pixmap = QPixmap(icon_path)
            icon_label.setPixmap(pixmap.scaledToHeight(400, Qt.SmoothTransformation))
            icon_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(icon_label)

        login_field = QLineEdit()
        login_field.setPlaceholderText("User ID")
        login_field.setStyleSheet("background-color: white;")
        layout.addWidget(login_field)

        button = QPushButton("Log In")
        layout.addWidget(button)

        def try_login():
            user_id = login_field.text().strip().lower()
            self.cur.execute("SELECT name FROM users WHERE LOWER(user_id) = ?", (user_id,))
            result = self.cur.fetchone()
            if result:
                self.active_user = result[0]
                self.login_input.setText(self.active_user)
                self.login_input.setReadOnly(True)
                if hasattr(self, "patient_login_input"):
                    self.patient_login_input.setText(self.active_user)
                    self.patient_login_input.setReadOnly(True)
                dialog.accept()
            else:
                login_field.clear()

        login_field.returnPressed.connect(try_login)
        button.clicked.connect(try_login)

        dialog.setLayout(layout)
        dialog.exec_()

        return bool(self.active_user)


    def is_authorized(self):
        if not self.active_user:
            QMessageBox.warning(self, "Access Denied", "You must log in as an active user to perform this action.")
            return False
        return True

    def clear_active_user_if_empty(self):
        if self.login_input.text().strip() == "" and not self.login_input.isReadOnly():
            self.active_user = None
            self.login_input.setReadOnly(False)
            if hasattr(self, 'patient_login_input'):
                self.patient_login_input.setReadOnly(False)
                self.patient_login_input.clear()
                self.disable_patient_controls()
            return




    def login_click_select_all(self, event):
        if self.login_input.isReadOnly():
            self.login_input.setReadOnly(False)
            self.active_user = None
        self.login_input.selectAll()
        QLineEdit.mousePressEvent(self.login_input, event)
        QLineEdit.mousePressEvent(self.login_input, event)



    def eventFilter(self, obj, event):
        # Handle backspace logout for both login fields
        if event.type() == event.KeyPress and event.key() == Qt.Key_Backspace:
            if obj == self.login_input:
                self.login_input.clear()
                self.active_user = None
                self.login_input.setReadOnly(False)
                self.disable_patient_controls()
                return True
            elif hasattr(self, 'patient_login_input') and obj == self.patient_login_input:
                self.patient_login_input.clear()
                self.active_user = None
                self.patient_login_input.setReadOnly(False)
                self.disable_patient_controls()
                return True
        if obj == self.login_input and event.type() == event.KeyPress:
            if event.key() == Qt.Key_Backspace:
                self.login_input.clear()
                self.active_user = None
                self.login_input.setReadOnly(False)
                self.disable_patient_controls()
                return True  # block default behavior

        if event.type() in (QEvent.KeyPress, QEvent.MouseMove, QEvent.MouseButtonPress):
            self.inactivity_timer.start()
            self.viewing_ceased = False

        return super().eventFilter(obj, event)

    def edit_current_pack_entry(self):
        print("DEBUG: Using REDEFINED edit_current_pack_entry")
        patient_name = getattr(self, 'current_patient_name', '').strip()
        if not patient_name:
            QMessageBox.warning(self, 'Missing Data', 'No current patient selected.')
            return
        xl_path = os.path.join(COLLECTION_LOGS_DIR, f"{patient_name.upper().replace(',', '').strip()}.xlsx")
        print(f'DEBUG: Trying to open Excel path: {xl_path}')
        if not os.path.exists(xl_path):
            QMessageBox.warning(self, 'Missing File', f'Could not find Excel file for {patient_name}.')
            return
        wb = openpyxl.load_workbook(xl_path)
        ws = wb.active
        med_names = []
        r = 4
        while True:
            val = ws.cell(row=r, column=1).value
            if val is None or str(val).strip() == '':
                break
            med_names.append(val)
            r += 1
        pack_date = self.pack_tables[self.current_pack_index][0]
        found_col = None
        for block_start in range(2, ws.max_column + 1, 6):
            val = ws.cell(row=4, column=block_start).value
            try:
                if isinstance(val, (int, float)):
                    cell_date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(val) - 2).date()
                elif isinstance(val, datetime):
                    cell_date = val.date()
                elif isinstance(val, str):
                    cell_date = datetime.strptime(val.strip(), "%d/%m/%Y").date()
                else:
                    continue
                if cell_date == datetime.strptime(pack_date, "%d/%m/%Y").date():
                    found_col = block_start
                    break
            except:
                continue
        if not found_col:
            QMessageBox.warning(self, 'Entry Not Found', f'Could not find matching pack date: {pack_date}')
            return
        self.blank_window = QWidget()
        self.blank_window.installEventFilter(self.global_event_filter)
        self.blank_window.setWindowTitle(f"Edit Pack Entry ({pack_date})")
        self.blank_window.setGeometry(150, 150, 800, 600)
        layout = QVBoxLayout()
        layout.addWidget(QLabel(f"Editing Pack Entry for: {patient_name} — {pack_date}"))
        table = QTableWidget()
        table.setColumnCount(4)
        table.setEditTriggers(QTableWidget.AllEditTriggers)
        table.setSelectionMode(QTableWidget.NoSelection)
        
        table.setHorizontalHeaderLabels(["Medication", "Batch/Expiry", "Qty Dispensed", "Qty Packed"])
        table.setEditTriggers(QTableWidget.AllEditTriggers)
        table.setSelectionMode(QTableWidget.NoSelection)
        table.verticalHeader().setVisible(False)
        header = table.horizontalHeader()
        for i in range(table.columnCount()):
            header.setSectionResizeMode(i, QHeaderView.Stretch)
        table.setRowCount(0)
        for i, drug in enumerate(med_names):
            if drug:
                table.insertRow(i)
                table.setItem(i, 0, QTableWidgetItem(str(drug)))
                be = ws.cell(row=i+4, column=found_col + 2).value
                disp = ws.cell(row=i+4, column=found_col + 1).value
                pack = ws.cell(row=i+4, column=found_col + 3).value
                table.setItem(i, 1, QTableWidgetItem(str(be) if be else ""))
                table.setItem(i, 2, QTableWidgetItem(str(disp) if disp else ""))
                table.setItem(i, 3, QTableWidgetItem(str(pack) if pack else ""))
        layout.addWidget(table)
        table.setStyleSheet("QTableWidget::item:selected { background-color: rgb(240, 240, 240); }")
        def highlight_row(row):
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item:
                    item.setBackground(QColor(240, 240, 240))
        def on_selection_or_edit():
            row = table.currentRow()
            if row >= 0:
                highlight_row(row)
        table.itemSelectionChanged.connect(on_selection_or_edit)
        table.itemChanged.connect(on_selection_or_edit)

        table.setStyleSheet("QTableWidget::item:selected { background-color: rgb(240, 240, 240); }")
        def highlight_row(row):
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item:
                    item.setBackground(QColor(240, 240, 240))
        def handle_cell_click(row, col):
            highlight_row(row)
        table.cellClicked.connect(handle_cell_click)

        def save_edits():
            for r in range(table.rowCount()):
                be = table.item(r, 1).text().strip()
                disp = table.item(r, 2).text().strip()
                pack = table.item(r, 3).text().strip()
                ws.cell(row=r + 4, column=found_col + 1).value = disp
                ws.cell(row=r + 4, column=found_col + 2).value = be
                ws.cell(row=r + 4, column=found_col + 3).value = pack
                try:
                    remaining = int(disp) - int(pack)
                except:
                    remaining = ""
                ws.cell(row=r + 4, column=found_col + 5).value = remaining
            wb.save(xl_path)
            self.blank_window.close()
            self.load_data()
            self.cur.execute("SELECT number FROM patients WHERE UPPER(name) = ?", (patient_name.upper(),))
            result = self.cur.fetchone()
            if result:
                number = result[0]
                row_index = None
                for row in range(self.table.rowCount()):
                    if self.table.item(row, 1).text() == str(number):
                        row_index = row
                        break
                if row_index is not None:
                    self.tabs.removeTab(self.tabs.currentIndex())
                    QTimer.singleShot(100, lambda: self.open_patient_tab(row_index, 0))
        btn_row = QHBoxLayout()
        save_btn = QPushButton("Save Edits")
        save_btn.clicked.connect(save_edits)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        layout.addLayout(btn_row)
        self.blank_window.setLayout(layout)
        self.blank_window.show()
    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, "tabs") and self.tabs.currentWidget():
            self.tabs.currentWidget().layout().update()
            for child in self.tabs.currentWidget().findChildren(QWidget):
                child.updateGeometry()

    def disable_patient_controls(self):
        for attr in ['first_name_input', 'last_name_input', 'charge_input', 'packed_input', 'picked_input', 'notes_input']:
            if hasattr(self, attr):
                getattr(self, attr).setDisabled(True)
        for dropdown in ['pack_size_dropdown', 'weeks_per_blister_dropdown']:
            if hasattr(self, dropdown):
                getattr(self, dropdown).setDisabled(True)
        if hasattr(self, 'claim_button'):
            self.claim_button.setDisabled(True)

    def enable_patient_controls(self):
        for attr in ['first_name_input', 'last_name_input', 'charge_input', 'packed_input', 'picked_input', 'notes_input']:
            if hasattr(self, attr):
                getattr(self, attr).setDisabled(False)
        for dropdown in ['pack_size_dropdown', 'weeks_per_blister_dropdown']:
            if hasattr(self, dropdown):
                getattr(self, dropdown).setDisabled(False)
        if hasattr(self, 'claim_button'):
            self.claim_button.setDisabled(True)

    def show_print_preview(self, html):

        preview_dialog = QDialog(self)
        preview_dialog.setWindowTitle("Print Preview")
        preview_dialog.resize(900, 600)
        layout = QVBoxLayout(preview_dialog)
        browser = QTextBrowser()
        browser.setHtml(html)
        button_layout = QHBoxLayout()
        print_btn = QPushButton("Print")
        cancel_btn = QPushButton("Close")
        button_layout.addWidget(cancel_btn)
        button_layout.addStretch()
        button_layout.addWidget(print_btn)
        layout.addWidget(browser)
        layout.addLayout(button_layout)
        def proceed_to_print():
            printer = QPrinter()
            printer.setOrientation(QPrinter.Landscape)
            printer.setPageSize(QPrinter.A4)
            dialog = QPrintDialog(printer, self)
            if dialog.exec_() == QPrintDialog.Accepted:
                doc = QTextDocument()
                doc.setHtml(html)
                doc.print_(printer)
            preview_dialog.accept()
        print_btn.clicked.connect(proceed_to_print)
        cancel_btn.clicked.connect(preview_dialog.reject)
        preview_dialog.exec_()

    def repopulate_pack_entry_drug_table(self, med_names, patient_id, ws):
        """
        Populates the pack-entry table with ACTIVE meds only, while preserving the
        original Excel row number (critical when active meds are a filtered subset).
        """
        table = self.pack_entry_drug_table
        table.setRowCount(0)

        cur = self.conn.cursor()

        for i, drug in enumerate(med_names):
            if not drug or not str(drug).strip():
                continue

            cur.execute(
                "SELECT active FROM patient_drug_status WHERE patient_id=? AND drug_name=?",
                (patient_id, drug),
            )
            row = cur.fetchone()
            if row is not None and not row[0]:
                continue

            excel_row = i + 4

            last_be = ""
            last_pack = ""
            for col in reversed(range(2, ws.max_column + 1, 6)):
                be_val = ws.cell(row=excel_row, column=col + 2).value
                pack_val = ws.cell(row=excel_row, column=col + 3).value
                if (be_val and str(be_val).strip()) or (pack_val and str(pack_val).strip()):
                    last_be = be_val if be_val else ""
                    last_pack = pack_val if pack_val else ""
                    break

            row_idx = table.rowCount()
            table.insertRow(row_idx)

            drug_item = QTableWidgetItem(str(drug))
            drug_item.setData(Qt.UserRole, excel_row)
            table.setItem(row_idx, 0, drug_item)

            table.setItem(row_idx, 1, QTableWidgetItem(str(last_be)))
            table.setItem(row_idx, 2, QTableWidgetItem(""))
            table.setItem(row_idx, 3, QTableWidgetItem(str(last_pack)))

    def handle_checked_action(self):
        print("DEBUG: Checked button clicked")

        if not self.enforce_login():
            print("DEBUG: Login required or failed")
            return
        print("DEBUG: Login OK")

        # Query DB for packed but unchecked patients
        self.cur.execute("""
                         SELECT number, name, packed_by, checked_by
                         FROM patients
                         WHERE packed_by IS NOT NULL
                           AND (checked_by IS NULL OR checked_by = '')
                           AND ceased = 0
                         """)
        rows = self.cur.fetchall()
        print(f"DEBUG: Rows returned from DB: {len(rows)}")
        if not rows:
            print("DEBUG: No patients match query → showing info popup")
            QMessageBox.information(self, "No Unchecked Packs", "There are no packs pending checking.")
            return



        # Create as a standalone top-level dialog window
        self.check_window = QWidget(None, Qt.Window | Qt.WindowTitleHint | Qt.WindowCloseButtonHint)
        self.check_window.setWindowTitle("Select Packs to Check")
        self.check_window.adjustSize()
        self.check_window.setWindowModality(Qt.NonModal)
        self.check_window.setWindowFlag(Qt.WindowStaysOnTopHint, True)

        layout = QVBoxLayout()
        layout.setSpacing(6)  # reduce space between widgets
        layout.setContentsMargins(10, 10, 10, 10)  # optional: adjust window padding

        # --- Checker Input ---
        checker_label = QLabel("Checker:")
        self.checker_input = OneBackspaceClearsLineEdit()
        self.checker_input.setPlaceholderText("CHECKER")
        if self.active_user:
            self.checker_input.setText(self.active_user)

        def handle_checker_input_change():
            self.checker_input.setStyleSheet("")

        self.checker_input.textChanged.connect(handle_checker_input_change)

        def fill_checker_full_name():
            entered_id = self.checker_input.text().strip()
            name = self.resolve_user_name(entered_id)
            if name:
                self.checker_input.setText(name)

        self.checker_input.returnPressed.connect(fill_checker_full_name)

        checker_row = QHBoxLayout()
        checker_row.setContentsMargins(0, 0, 0, 0)
        checker_row.setSpacing(6)
        checker_row.addWidget(checker_label)
        checker_row.addWidget(self.checker_input)
        layout.addLayout(checker_row)

        self.checker_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        # --- Patient checkboxes ---
        self.checkboxes_to_check = []
        for number, name, packed_by, checked_by in rows:
            cb = QCheckBox(f"{name} (Packed by: {packed_by})")
            cb.patient_number = number
            layout.addWidget(cb)
            self.checkboxes_to_check.append(cb)

        # --- Buttons ---
        button_layout = QHBoxLayout()
        button_layout.addStretch(1)  # Pushes buttons to center
        cancel_btn = QPushButton("Cancel")
        check_btn = QPushButton("Check")
        button_layout.addWidget(cancel_btn)
        button_layout.addSpacing(20)  # Gap between buttons
        button_layout.addWidget(check_btn)
        button_layout.addStretch(1)  # Keeps them centered
        layout.addLayout(button_layout)

        self.check_window.setLayout(layout)

        def apply_checking():
            checker_id = self.checker_input.text().strip()
            if not checker_id:
                QMessageBox.warning(self.check_window, "Input Required", "Checker ID is required.")
                return

            for cb in self.checkboxes_to_check:
                if cb.isChecked():
                    self.cur.execute(
                        "UPDATE patients SET checked_by = ? WHERE number = ?",
                        (checker_id, cb.patient_number)
                    )

                    # Excel update
                    self.cur.execute("SELECT name, date_packed FROM patients WHERE number = ?", (cb.patient_number,))
                    patient_info = self.cur.fetchone()
                    if not patient_info:
                        continue
                    name, packed_date = patient_info
                    xl_path = os.path.join(COLLECTION_LOGS_DIR, f"{name.upper().replace(',', '').strip()}.xlsx")
                    print(f'DEBUG: Trying to open Excel path: {xl_path}')
                    if not os.path.exists(xl_path):
                        continue
                    try:
                        wb = openpyxl.load_workbook(xl_path)
                        ws = wb["Year"]
                        for block_start in range(2, ws.max_column + 1, 6):
                            for r in range(4, 20):
                                cell_val = ws.cell(row=r, column=block_start).value
                                if cell_val:
                                    if isinstance(cell_val, (int, float)):
                                        date_val = (datetime(1900, 1, 1).toordinal() + int(cell_val) - 2)
                                        cell_date = datetime.fromordinal(date_val).strftime("%d/%m/%Y")
                                    elif isinstance(cell_val, datetime):
                                        cell_date = cell_val.strftime("%d/%m/%Y")
                                    else:
                                        cell_date = str(cell_val)
                                    if cell_date == packed_date:
                                        for r_check in range(4, 20):
                                            if ws.cell(row=r_check, column=1).value:
                                                initials_cell = ws.cell(row=r_check, column=block_start + 4)
                                                if initials_cell.value and '/' in initials_cell.value:
                                                    parts = initials_cell.value.split('/')
                                                    if len(parts) == 1 or (len(parts) == 2 and not parts[1]):
                                                        initials_cell.value = f"{parts[0]}/{checker_id}"
                                        wb.save(xl_path)
                                        break
                    except Exception as e:
                        print("CHECK WRITE ERROR:", e)

            self.conn.commit()
            self.load_data()
            self.check_window.close()

        cancel_btn.clicked.connect(self.check_window.close)
        check_btn.clicked.connect(apply_checking)

        # --- Show window ---
        self.check_window.adjustSize()
        self.check_window.show()
        self.check_window.raise_()
        self.check_window.activateWindow()

        # Center over main app
        fg = self.frameGeometry()
        center_point = fg.center()
        self.check_window.move(center_point - self.check_window.rect().center())

        print(f"DEBUG: Check window opened with {len(rows)} patients (separate top-level window)")

    def handle_collected_action(self):
        if not self.enforce_login():
            return

        selected_row = self.table.currentRow()
        if selected_row < 0:
            return

        # --- Skip if paused or flagged ---
        name_item = self.table.item(selected_row, 3)
        if name_item:
            patient_name = name_item.text().strip()
            self.cur.execute("SELECT paused, flagged FROM patients WHERE name = ?", (patient_name,))
            flags = self.cur.fetchone()
            if flags and (flags[0] or flags[1]):
                return  # silently do nothing

        # --- Skip if already collected ---
        collected_item = self.table.item(selected_row, 5)  # "Picked Up" column
        if collected_item and collected_item.text().strip():
            return  # silently do nothing

        number_item = self.table.item(selected_row, 1)
        if not number_item:
            return
        number = number_item.text().strip()

        # --- Build dialog with both inputs ---
        dlg = QDialog(self)
        dlg.setWindowTitle("Collected")
        layout = QVBoxLayout(dlg)

        # Collected date
        layout.addWidget(QLabel("Date collected:"))
        date_edit = QDateEdit(dlg)
        date_edit.setDisplayFormat("dd/MM/yyyy")
        date_edit.setCalendarPopup(True)
        date_edit.setDate(QDate.currentDate())
        layout.addWidget(date_edit)

        # Weeks dropdown
        layout.addWidget(QLabel("Weeks collected:"))
        weeks_combo = QComboBox(dlg)
        weeks_combo.addItems(["1 week", "2 weeks", "3 weeks", "4 weeks",
                              "5 weeks", "6 weeks", "7 weeks", "8 weeks", "Custom…"])
        weeks_combo.setCurrentText("4 weeks")
        layout.addWidget(weeks_combo)

        # Custom weeks field (hidden unless "Custom…" selected)
        custom_weeks_edit = QLineEdit(dlg)
        custom_weeks_edit.setPlaceholderText("Enter number of weeks")
        custom_weeks_edit.hide()
        layout.addWidget(custom_weeks_edit)

        def on_weeks_changed(_):
            custom_weeks_edit.setVisible("custom" in weeks_combo.currentText().lower())

        weeks_combo.currentIndexChanged.connect(on_weeks_changed)

        # Buttons
        btns = QHBoxLayout()
        ok_btn = QPushButton("OK", dlg)
        cancel_btn = QPushButton("Cancel", dlg)
        btns.addStretch(1)
        btns.addWidget(cancel_btn)
        btns.addWidget(ok_btn)
        layout.addLayout(btns)

        cancel_btn.clicked.connect(dlg.reject)
        ok_btn.clicked.connect(dlg.accept)

        if dlg.exec_() != QDialog.Accepted:
            return

        # --- Validate date directly from QDateEdit ---
        collected_dt = date_edit.date().toPyDate()

        label = weeks_combo.currentText().lower()
        if "custom" in label:
            txt = custom_weeks_edit.text().strip()
            if not txt.isdigit() or int(txt) <= 0:
                return  # silently ignore invalid input
            weeks = int(txt)
        else:
            weeks = int(label.split()[0])

        # --- Calculate due date ---
        base_due = collected_dt + timedelta(days=21)
        due_dt = base_due + timedelta(days=(weeks - 4) * 7)

        collected_date_str = collected_dt.strftime("%d/%m/%Y")
        due_date_str = due_dt.strftime("%d/%m/%Y")

        # Save to DB
        self.cur.execute(
            "UPDATE patients SET picked_up = ?, due_date = ?, date_packed = '' WHERE number = ?",
            (collected_date_str, due_date_str, number)
        )
        self.conn.commit()

        self.load_data()
        self.table.selectRow(selected_row)

    def handle_print_action(self):
        dlg = PrintPreviewDialog(self)
        dlg.exec_()

    def populate_drugs_table(self, med_names):
        self.table_widget.clearContents()
        self.table_widget.setRowCount(len(med_names))
        for row, drug in enumerate(med_names):
            item = QTableWidgetItem(drug)
            self.table_widget.setItem(row, 0, item)

    def handle_new_packed_action(self, patient_number=None, prefill_date=None):
        selected_row = self.table.currentRow()
        if selected_row < 0:
            # absolutely nothing happens if no row is selected
            self.current_patient_id = None
            self.current_patient_number = None
            return

        # --- Green row check ---
        date_packed_item = self.table.item(selected_row, 4)  # "Date Packed"
        collected_item = self.table.item(selected_row, 5)  # "Picked Up"

        date_packed = date_packed_item.text().strip() if date_packed_item else ""
        collected = collected_item.text().strip() if collected_item else ""

        if date_packed and not collected:
            reply = QMessageBox.question(
                self,
                "Uncollected Pack",
                "Previous supply uncollected. Create new entry?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return

        # --- Extract FULL NAME (for Excel filename) ---
        name_item = self.table.item(selected_row, 3)
        if not name_item:
            return
        full_name = name_item.text().strip()  # e.g. "Troy McKinnon"

        # Split name into first/last for display header
        parts = full_name.split()
        if len(parts) >= 2:
            first_names = parts[0]
            last_name = " ".join(parts[1:])
        else:
            first_names = full_name
            last_name = ""

        # --- Extract NUMBER (correct for DB lookup) ---
        number_item = self.table.item(selected_row, 1)
        if not number_item:
            return

        number = number_item.text().strip()

        # ------------------------------------------------------------------
        #  DB LOOKUP USING NUMBER (correct — names vary, numbers do not)
        # ------------------------------------------------------------------
        self.cur.execute("SELECT id FROM patients WHERE number = ?", (number,))
        result = self.cur.fetchone()

        if not result:
            QMessageBox.warning(
                self,
                "Not Found",
                f"Could not find patient #{number} in database."
            )
            return

        patient_id = result[0]

        # ------------------------------------------------------------------
        #  EXCEL FILENAME: based on DISPLAY NAME, not DB name
        # ------------------------------------------------------------------
        excel_filename = f"{full_name.upper().replace(',', '').strip()}.xlsx"
        xl_path = os.path.join(COLLECTION_LOGS_DIR, excel_filename)
        self.excel_path = xl_path

        print(f"DEBUG: Trying to open Excel path: {xl_path}")

        if not os.path.exists(xl_path):
            QMessageBox.warning(
                self,
                "Missing Collection Log",
                f"Collection log not found.\nExpected file:\n\n{excel_filename}"
            )
            return

        self.excel_path = xl_path
        print(f'DEBUG: Trying to open Excel path: {xl_path}')
        if not os.path.exists(xl_path):
            return

        wb = openpyxl.load_workbook(xl_path)
        ws = wb.active

        # Get drug names from column A starting at row 4
        med_names = []
        r = 4
        while True:
            val = ws.cell(row=r, column=1).value
            if val is None or str(val).strip() == '':
                break
            med_names.append(val)
            r += 1

        self.current_patient_id = patient_id
        self.current_med_names = med_names
        self.pack_entry_ws = ws



        self.blank_pack_window = QWidget()
        self.blank_pack_window.setAttribute(Qt.WA_DeleteOnClose, True)
        self.blank_pack_window.destroyed.connect(self._clear_pack_entry_context)
        self.blank_pack_window.setWindowTitle("New Pack Entry")
        self.blank_pack_window.setAttribute(Qt.WA_DeleteOnClose, True)
        self.blank_pack_window.setMinimumSize(800, 500)
        self.blank_pack_window.setWindowFlags(self.blank_pack_window.windowFlags() | Qt.WindowStaysOnTopHint)

        main_layout = QVBoxLayout()
        header_layout = QHBoxLayout()

        # Patient name left (large bold)
        name_box = QVBoxLayout()
        first_label = QLabel(first_names)
        last_label = QLabel(last_name)
        name_font = QFont()
        name_font.setBold(True)
        name_font.setPointSize(14)
        first_label.setFont(name_font)
        last_label.setFont(name_font)
        name_box.addWidget(first_label)
        name_box.addWidget(last_label)
        header_layout.addLayout(name_box, stretch=2)

        # Pack date / packer input fields
        form_box = QVBoxLayout()
        self.pack_date_input = QDateEdit()
        self.pack_date_input.setDisplayFormat("dd/MM/yyyy")
        self.pack_date_input.setCalendarPopup(True)
        self.pack_date_input.setKeyboardTracking(False)
        self.pack_date_input.setDate(QDate.currentDate())
        self.pack_date_input.setObjectName("pack_date_input")
        self.packer_input = OneBackspaceClearsLineEdit()
        self.packer_input.setText(self.active_user if self.active_user else "")
        self.packer_input.textChanged.connect(self.handle_packer_input_change)
        self.packer_input.returnPressed.connect(self.fill_packer_full_name)
        self.packer_input.setPlaceholderText("PACKER")
        form_box.addWidget(self.pack_date_input)
        form_box.addWidget(self.packer_input)
        header_layout.addLayout(form_box, stretch=1)
        main_layout.addLayout(header_layout)

        # Drug entry table
        table = QTableWidget()
        table.setColumnCount(4)
        table.setEditTriggers(QTableWidget.AllEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setSelectionMode(QTableWidget.SingleSelection)
        self.pack_entry_drug_table = table

        table.setHorizontalHeaderLabels(["Drug", "Batch/Expiry", "Qty\nDispensed", "Qty\nPacked"])
        header = table.horizontalHeader()
        font = header.font()
        font.setBold(True)
        header.setFont(font)

        header = table.horizontalHeader()
        table.setColumnWidth(0, 250)
        table.setColumnWidth(2, 65)
        table.setColumnWidth(3, 65)
        header.setSectionResizeMode(0, QHeaderView.Interactive)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.Interactive)
        header.setSectionResizeMode(3, QHeaderView.Interactive)
        table.horizontalHeader().setMinimumSectionSize(120)
        table.setColumnWidth(1, 220)
        table.horizontalHeader().resizeSection(1, 220)
        table.horizontalHeader().setMaximumSectionSize(340)

        table.setRowCount(0)

        # <<--- CALL TABLE POPULATION LOGIC HERE --->
        self.repopulate_pack_entry_drug_table(med_names, patient_id, ws)
        # <<--- END TABLE POPULATION --->

        main_layout.addWidget(table)

        table.setStyleSheet("QTableWidget::item:selected { background-color: rgb(240, 240, 240); }")

        def highlight_row(row):
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item:
                    item.setBackground(QColor(240, 240, 240))

        def on_selection_or_edit():
            row = table.currentRow()
            if row >= 0:
                highlight_row(row)

        table.itemSelectionChanged.connect(on_selection_or_edit)
        table.itemChanged.connect(on_selection_or_edit)

        def handle_cell_click(row, col):
            highlight_row(row)

        table.cellClicked.connect(handle_cell_click)

        self.edit_drugs_button = QPushButton("Edit Drugs")
        self.edit_drugs_button.clicked.connect(self.open_edit_drugs_dialog)
        print("DEBUG: Edit Drugs button created and connected.")

        clear_btn = QPushButton("Clear Highlights")
        clear_btn.setFixedWidth(clear_btn.sizeHint().width())

        def clear_highlights():
            table.clearSelection()
            table.setCurrentCell(-1, -1)
            table.blockSignals(True)
            table.clearSelection()
            for r in range(table.rowCount()):
                for c in range(table.columnCount()):
                    item = table.item(r, c)
                    if item:
                        item.setBackground(QColor("white"))
            table.blockSignals(False)

        if hasattr(table, 'clicked_rows'):
            table.clicked_rows.clear()
        clear_btn.clicked.connect(clear_highlights)
        # Bottom button row
        # --- Bottom button row ---
        button_row = QHBoxLayout()

        # Done button (left aligned)
        done_button = QPushButton("Done")

        def finish_pack_entry():


            pack_date = self.pack_date_input.text().strip()
            packer_id = self.packer_input.text().strip()

            if not pack_date or not packer_id:
                QMessageBox.warning(self.blank_pack_window, "Missing Info", "Pack date and packer must be filled.")
                return

            # Convert full name to initials if full name entered
            if " " in packer_id and "/" not in packer_id:
                packer_id = "".join([p[0].upper() for p in packer_id.split()])

            # Append '/' to leave unchecked
            if not packer_id.endswith("/"):
                packer_id = f"{packer_id}/"

            # Load workbook
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active

            # Find next available 6-col block in columns B–G, H–M, ...
            block_size = 6
            start_col = 2  # B
            max_col = ws.max_column
            next_block_start = max_col + 1
            # Align start_col to correct 6-col block offset
            offset = (next_block_start - start_col) % block_size
            if offset != 0:
                next_block_start += (block_size - offset)

            # Write data for each drug row in table
            for row in range(self.pack_entry_drug_table.rowCount()):
                drug_name = self.pack_entry_drug_table.item(row, 0)
                batch_expiry = self.pack_entry_drug_table.item(row, 1)
                qty_dispensed = self.pack_entry_drug_table.item(row, 2)
                qty_packed = self.pack_entry_drug_table.item(row, 3)

                if not drug_name or not drug_name.text().strip():
                    continue

                drug_cell = self.pack_entry_drug_table.item(row, 0)
                drug_row = drug_cell.data(Qt.UserRole) if drug_cell else None
                if not isinstance(drug_row, int):
                    continue

                ws.cell(row=drug_row, column=next_block_start + 0).value = pack_date
                ws.cell(row=drug_row,
                        column=next_block_start + 1).value = qty_dispensed.text().strip() if qty_dispensed else ""
                ws.cell(row=drug_row,
                        column=next_block_start + 2).value = batch_expiry.text().strip() if batch_expiry else ""
                ws.cell(row=drug_row,
                        column=next_block_start + 3).value = qty_packed.text().strip() if qty_packed else ""
                ws.cell(row=drug_row, column=next_block_start + 4).value = packer_id
                try:
                    rem = (int(qty_dispensed.text()) if qty_dispensed else 0) - (
                        int(qty_packed.text()) if qty_packed else 0)
                except:
                    rem = ""
                ws.cell(row=drug_row, column=next_block_start + 5).value = rem

            wb.save(self.excel_path)

            # Mark patient as unchecked in DB
            self.cur.execute("""
                             UPDATE patients
                             SET date_packed = ?,
                                 picked_up   = NULL, -- remove collection date
                                 due_date    = NULL, -- also clear due if needed
                                 packed_by   = ?,
                                 checked_by  = NULL
                             WHERE id = ?
                             """, (pack_date, packer_id, self.current_patient_id))

            self.conn.commit()

            # Refresh main table
            self.load_data()

            # Close the pack entry window
            self.blank_pack_window.close()

        done_button.clicked.connect(finish_pack_entry)
        button_row.addWidget(done_button, alignment=Qt.AlignLeft)

        # Spacer to push others right
        button_row.addStretch()

        # Existing right-aligned buttons
        button_row.addWidget(self.edit_drugs_button)
        button_row.addWidget(clear_btn)

        main_layout.addLayout(button_row)

        self.blank_pack_window.setLayout(main_layout)
        self.blank_pack_window.show()

        table.verticalHeader().setVisible(False)

    def handle_add_patient(self):
        """Open a window to add a new patient and create an Excel collection log file."""
        self.add_patient_window = QWidget()
        self.add_patient_window.setWindowTitle("Add Patient")
        self.add_patient_window.setGeometry(200, 200, 400, 250)
        self.add_patient_window.setWindowModality(Qt.ApplicationModal)
        self.add_patient_window.setWindowFlag(Qt.WindowStaysOnTopHint, True)

        layout = QVBoxLayout()
        form_layout = QGridLayout()

        # --- Determine next patient number ---
        try:
            # Ensure numeric comparison, ignore NULLs and non-integers
            self.cur.execute("""
                             SELECT MAX(CAST(number AS INTEGER))
                             FROM patients
                             WHERE number GLOB '[0-9]*'
                             """)
            result = self.cur.fetchone()
            next_number = (int(result[0]) + 1) if result and result[0] is not None else 1
        except Exception as e:
            print("DEBUG: Error fetching next patient number:", e)
            next_number = 1

        # --- Labels and Inputs ---
        first_label = QLabel("First Names:")
        self.first_name_field = QLineEdit()
        self.first_name_field.setPlaceholderText("Enter first names")

        last_label = QLabel("Last Name:")
        self.last_name_field = QLineEdit()
        self.last_name_field.setPlaceholderText("Enter last name")

        number_label = QLabel("Patient Number:")
        self.patient_number_field = QLineEdit()
        self.patient_number_field.setText(str(next_number))
        self.patient_number_field.setReadOnly(True)
        self.patient_number_field.setStyleSheet("background-color: #eee;")

        form_layout.addWidget(first_label, 0, 0)
        form_layout.addWidget(self.first_name_field, 0, 1)
        form_layout.addWidget(last_label, 1, 0)
        form_layout.addWidget(self.last_name_field, 1, 1)
        form_layout.addWidget(number_label, 2, 0)
        form_layout.addWidget(self.patient_number_field, 2, 1)
        layout.addLayout(form_layout)

        # --- Save Button ---
        save_btn = QPushButton("Save")
        save_btn.clicked.connect(self.save_new_patient)
        layout.addWidget(save_btn, alignment=Qt.AlignRight)

        self.add_patient_window.setLayout(layout)
        self.first_name_field.setFocus()
        self.add_patient_window.show()

    def save_new_patient(self):
        """Save new patient details and create their Excel log.

        If the entered name already exists:
          - active patient: ask to duplicate as "Name (2)", "Name (3)", etc.
          - ceased patient: ask to restore instead of creating a new record
        """
        first = self.first_name_field.text().strip()
        last = self.last_name_field.text().strip()
        number = self.patient_number_field.text().strip()

        if not first or not last or not number:
            return

        base_name = f"{last}, {first}"

        # --- Check for exact existing patient name ---
        self.cur.execute(
            "SELECT number, ceased FROM patients WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))",
            (base_name,)
        )
        existing_rows = self.cur.fetchall()

        final_name = base_name

        if existing_rows:
            active_match = next((r for r in existing_rows if int(r[1] or 0) == 0), None)
            ceased_match = next((r for r in existing_rows if int(r[1] or 0) == 1), None)

            # Exact active patient exists -> duplicate or cancel
            if active_match:
                msg = QMessageBox(self)
                msg.setWindowTitle("Patient Already Exists")
                msg.setText(f"{base_name.title()} already exists.")
                dup_btn = msg.addButton("Duplicate", QMessageBox.AcceptRole)
                cancel_btn = msg.addButton("Cancel", QMessageBox.RejectRole)
                msg.exec_()

                if msg.clickedButton() != dup_btn:
                    return

                # Find next available duplicate suffix: (2), (3), ...
                suffix = 2
                while True:
                    candidate_name = f"{base_name} ({suffix})"
                    self.cur.execute(
                        "SELECT COUNT(*) FROM patients WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))",
                        (candidate_name,)
                    )
                    if self.cur.fetchone()[0] == 0:
                        final_name = candidate_name
                        break
                    suffix += 1

            # Exact ceased patient exists -> restore / duplicate / cancel
            elif ceased_match:
                msg = QMessageBox(self)
                msg.setWindowTitle("Ceased Patient Found")
                msg.setText(f"{base_name.title()} exists in ceased patients.")
                restore_btn = msg.addButton("Restore", QMessageBox.AcceptRole)
                duplicate_btn = msg.addButton("Duplicate", QMessageBox.ActionRole)
                cancel_btn = msg.addButton("Cancel", QMessageBox.RejectRole)
                msg.exec_()

                clicked = msg.clickedButton()

                if clicked == restore_btn:
                    ceased_number = str(ceased_match[0]).strip()
                    self.cur.execute(
                        "UPDATE patients SET ceased = 0 WHERE number = ?",
                        (ceased_number,)
                    )
                    self.conn.commit()
                    self.add_patient_window.close()
                    self.open_patients_page()
                    return

                elif clicked == duplicate_btn:
                    suffix = 2
                    while True:
                        candidate_name = f"{base_name} ({suffix})"
                        self.cur.execute(
                            "SELECT COUNT(*) FROM patients WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))",
                            (candidate_name,)
                        )
                        if self.cur.fetchone()[0] == 0:
                            final_name = candidate_name
                            break
                        suffix += 1
                else:
                    return

        # --- Insert new patient record ---
        self.cur.execute("""
                         INSERT INTO patients (number, name, notes, charge, date_packed, picked_up,
                                               due_date, given_out_by, weeks_per_blister,
                                               flagged, paused, ceased, pack_size)
                         VALUES (?, ?, '', 'N', '', '', '', '', '4 weeks', 0, 0, 0, 'Large')
                         """, (number, final_name))

        self.conn.commit()

        # --- Create Excel log file ---
        try:
            safe_name = final_name.upper().replace(",", "").strip()
            excel_path = os.path.join(COLLECTION_LOGS_DIR, f"{safe_name}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Year"

            ws.cell(row=1, column=1, value="Medication List")
            ws.cell(row=3, column=1, value="Drug")
            headers = ["Date", "Qty Dispensed", "Batch/Expiry", "Qty Packed", "Initials", "Remaining"]
            for i, h in enumerate(headers, start=2):
                ws.cell(row=3, column=i, value=h)

            for r in range(4, 20):
                for c in range(1, 8):
                    ws.cell(row=r, column=c, value=None)

            wb.save(excel_path)

        except Exception as e:
            print("EXCEL CREATE ERROR:", e)

        self.add_patient_window.close()
        self.open_patients_page()


    def handle_cease_patient(self):
        """Mark a selected patient as ceased, regardless of case/format."""
        selected = self.patient_table.currentRow()
        if selected < 0:
            return

        name_item = self.patient_table.item(selected, 1)
        number_item = self.patient_table.item(selected, 0)
        if not name_item or not number_item:
            return

        # Prefer using patient number for reliable lookup
        number = number_item.text().strip()
        if not number:
            return

        try:
            # Update via patient number instead of name
            self.cur.execute("UPDATE patients SET ceased = 1 WHERE number = ?", (number,))
            self.conn.commit()
            print(f"DEBUG: Ceased patient #{number}")
        except Exception as e:
            print(f"ERROR: Failed to cease patient #{number}: {e}")

        # Refresh patient manager window
        self.open_patients_page()

    def closeEvent(self, event):
        """Ensure the file observer is stopped cleanly on exit."""
        try:
            if hasattr(self, "db_observer"):
                self.db_observer.stop()
                self.db_observer.join(timeout=1)
        except Exception as e:
            print(f"DEBUG: Error stopping DB watcher: {e}")
        event.accept()

def send_ipc_command(command: str, server_name: str = SINGLE_INSTANCE_KEY, timeout_ms: int = 500) -> bool:
    sock = QLocalSocket()
    sock.connectToServer(server_name)
    if not sock.waitForConnected(timeout_ms):
        return False
    payload = (command.strip() + "\n").encode("utf-8")
    sock.write(payload)
    sock.flush()
    sock.waitForBytesWritten(timeout_ms)
    sock.disconnectFromServer()
    return True


def attach_ipc_listener(server: QLocalServer, app: QApplication) -> None:
    def on_new_connection():
        while server.hasPendingConnections():
            conn = server.nextPendingConnection()
            if conn is None:
                return

            conn.waitForReadyRead(200)
            raw = bytes(conn.readAll()).decode("utf-8", errors="ignore").strip().upper()

            conn.disconnectFromServer()
            conn.deleteLater()

            if raw == "QUIT":
                app.quit()

    server.newConnection.connect(on_new_connection)

# --- file: DAA_Calendar.py ---

if __name__ == "__main__":
    # ---------------------------
    # IPC helpers (single-instance)
    # ---------------------------
    def send_ipc_command(command: str, server_name: str = SINGLE_INSTANCE_KEY, timeout_ms: int = 800) -> bool:
        """
        Sends a short command to the already-running instance over QLocalSocket.
        Returns True if delivered, False if couldn't connect.
        """
        sock = QLocalSocket()
        sock.connectToServer(server_name)
        if not sock.waitForConnected(timeout_ms):
            return False
        payload = (command.strip() + "\n").encode("utf-8")
        sock.write(payload)
        sock.flush()
        sock.waitForBytesWritten(timeout_ms)
        sock.disconnectFromServer()
        return True

    def attach_ipc_listener(server: "QLocalServer", app: "QApplication") -> None:
        """
        Makes THIS instance respond to IPC commands from later-launched instances.
        Supported:
          - QUIT: gracefully exits this instance.
        """
        def on_new_connection():
            while server.hasPendingConnections():
                conn = server.nextPendingConnection()
                if conn is None:
                    return

                conn.waitForReadyRead(1000)
                raw = bytes(conn.readAll()).decode("utf-8", errors="ignore").strip().upper()

                conn.disconnectFromServer()
                conn.deleteLater()

                if raw == "QUIT":
                    app.quit()

        server.newConnection.connect(on_new_connection)

    # ---------------------------
    # Load stored settings (your existing logic)
    # ---------------------------
    settings_path = os.path.join(os.path.dirname(__file__), "settings.json")
    collection_logs_dir = ""
    if os.path.exists(settings_path):
        try:
            with open(settings_path, "r") as f:
                settings = json.load(f)
                collection_logs_dir = settings.get("collection_logs_dir", "")
        except Exception as e:
            print(f"DEBUG: Failed to load settings: {e}")

    # ---------------------------
    # QApplication
    # ---------------------------
    app = QApplication.instance() or QApplication([])
    app.setWindowIcon(QIcon(resource_path("DAACal.ico")))

    # ---------------------------
    # Single instance check
    # ---------------------------
    server = ensure_single_instance()

    if server is None:
        # --- Build the "already running" dialog ---
        msg = QMessageBox()
        msg.setWindowTitle("DAACal")

        ico_path = os.path.join(BASE_DIR, "daacal.ico")
        png_path = os.path.join(BASE_DIR, "daacal.png")
        icon = None
        if os.path.exists(ico_path):
            icon = QIcon(ico_path)
        elif os.path.exists(png_path):
            icon = QIcon(png_path)

        if icon is not None:
            msg.setWindowIcon(icon)

        if os.path.exists(png_path):
            msg.setIconPixmap(QPixmap(png_path).scaled(64, 64))

        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msg.button(QMessageBox.Yes).setText(" Relaunch ")
        msg.button(QMessageBox.No).setText(" Cancel ")

        # --- Countdown auto-close: 15s (acts like Cancel) ---
        popup_timeout_secs = 15
        remaining_secs = {"v": popup_timeout_secs}

        base_text = (
            "<b>DAACal is already open</b><br>"
            "<span style='font-size:11pt;'>Another session is running on this device.</span><br><br>"
            "Would you like to relaunch DAACal?"
        )

        def render_popup_text() -> None:
            msg.setText(
                base_text
                + f"<br><br><span style='font-size:10pt;'>Closing in <b>{remaining_secs['v']}</b> seconds…</span>"
            )

        render_popup_text()

        countdown_timer = QTimer(msg)
        countdown_timer.setInterval(1000)

        def on_tick() -> None:
            remaining_secs["v"] = max(0, remaining_secs["v"] - 1)
            render_popup_text()
            if remaining_secs["v"] <= 0:
                countdown_timer.stop()
                msg.reject()

        countdown_timer.timeout.connect(on_tick)
        countdown_timer.start()

        clicked = msg.exec_()
        countdown_timer.stop()

        if clicked != QMessageBox.Yes:
            sys.exit(0)

        # ---------------------------
        # Relaunch flow
        # ---------------------------
        sent = send_ipc_command("QUIT")
        if not sent:
            QMessageBox.warning(None, "DAACal", "Could not contact the running DAACal instance.")
            sys.exit(0)

        # Waiting dialog while old instance closes / lock is released
        wait_box = QMessageBox()
        wait_box.setWindowTitle("DAACal")
        wait_box.setStandardButtons(QMessageBox.NoButton)
        if icon is not None:
            wait_box.setWindowIcon(icon)
        if os.path.exists(png_path):
            wait_box.setIconPixmap(QPixmap(png_path).scaled(64, 64))

        retries_ms_total = 15000
        started_at = time.time()
        retry_every_ms = 150
        server_box = {"server": None}

        wait_remaining = {"v": int(retries_ms_total / 1000)}

        def render_wait_text() -> None:
            wait_box.setText(
                "<b>Relaunching…</b><br>"
                "Please wait while the previous session closes."
                f"<br><br><span style='font-size:10pt;'>Waiting up to <b>{wait_remaining['v']}</b> seconds…</span>"
            )

        render_wait_text()

        wait_countdown_timer = QTimer(wait_box)
        wait_countdown_timer.setInterval(1000)

        def wait_tick() -> None:
            wait_remaining["v"] = max(0, wait_remaining["v"] - 1)
            render_wait_text()

        wait_countdown_timer.timeout.connect(wait_tick)
        wait_countdown_timer.start()

        def try_acquire() -> None:
            server_box["server"] = ensure_single_instance()
            if server_box["server"] is not None:
                wait_box.accept()
                return

            elapsed_ms = (time.time() - started_at) * 1000
            if elapsed_ms >= retries_ms_total:
                wait_box.reject()
                return

            QTimer.singleShot(retry_every_ms, try_acquire)

        QTimer.singleShot(0, try_acquire)
        wait_box.exec_()
        wait_countdown_timer.stop()

        server = server_box["server"]
        if server is None:
            QMessageBox.warning(
                None,
                "DAACal",
                "Previous instance is taking too long to close.\n\n"
                "Please close it manually, then relaunch DAACal."
            )
            sys.exit(0)

        print("Previous instance closed. Launching fresh session...")

    # keep server alive so this instance owns the name
    app.single_instance_server = server
    attach_ipc_listener(server, app)

    # --- AUTO-CLOSE AFTER 8 HOURS (WITH 10-MINUTE LIVE COUNTDOWN WARNING) ---
    total_ms = 8 * 60 * 60 * 1000  # 8 hours
    warning_ms = total_ms - 10 * 60 * 1000  # 10 minutes before close

    auto_close_timer = QTimer(app)
    auto_close_timer.setSingleShot(True)
    auto_close_warning_timer = QTimer(app)
    auto_close_warning_timer.setSingleShot(True)

    # Keep references on the app so they don't get GC'd
    app.auto_close_timer = auto_close_timer
    app.auto_close_warning_timer = auto_close_warning_timer

    def show_auto_close_warning():
        """
        Show DAACal-styled popup 10 minutes before auto-close.
        User can keep DAACal open (cancel auto-close) or close now.
        Includes a live MM:SS countdown until auto-close.
        """
        msg = QMessageBox()
        msg.setWindowTitle("DAACal")

        ico_path = os.path.join(BASE_DIR, "daacal.ico")
        png_path = os.path.join(BASE_DIR, "daacal.png")
        icon2 = None
        if os.path.exists(ico_path):
            icon2 = QIcon(ico_path)
        elif os.path.exists(png_path):
            icon2 = QIcon(png_path)

        if icon2 is not None:
            msg.setWindowIcon(icon2)
        if os.path.exists(png_path):
            msg.setIconPixmap(QPixmap(png_path).scaled(64, 64))

        def format_secs(seconds: int) -> str:
            m = max(0, seconds) // 60
            s = max(0, seconds) % 60
            return f"{m:02d}:{s:02d}"

        def update_text():
            if auto_close_timer.isActive():
                rem_ms = max(0, auto_close_timer.remainingTime())
            else:
                rem_ms = 0
            rem_sec = rem_ms // 1000
            msg.setText(
                "<b>DAACal will close soon</b><br>"
                "Session timer active.<br><br>"
                "Session will close automatically.<br>"
                f"Closing in <b>{format_secs(rem_sec)}</b>."
            )
            if rem_ms <= 0:
                countdown_timer.stop()

        keep_btn = msg.addButton("Keep DAACal open", QMessageBox.YesRole)
        close_btn = msg.addButton("Close now", QMessageBox.NoRole)

        countdown_timer = QTimer(msg)
        countdown_timer.timeout.connect(update_text)
        countdown_timer.start(1000)
        update_text()

        msg.exec_()
        countdown_timer.stop()

        clicked_btn = msg.clickedButton()
        if clicked_btn is keep_btn:
            auto_close_timer.stop()
            print("DEBUG: Auto-close cancelled by user from countdown dialog.")
        elif clicked_btn is close_btn:
            print("DEBUG: User chose to close DAACal immediately from countdown dialog.")
            app.quit()

    # Hook timers
    auto_close_warning_timer.timeout.connect(show_auto_close_warning)
    auto_close_timer.timeout.connect(app.quit)

    # Start timers
    auto_close_warning_timer.start(warning_ms)
    auto_close_timer.start(total_ms)

    # Set global icon for all dialogs and widgets
    icon_path = os.path.join(BASE_DIR, "DAACal.ico")
    if os.path.exists(icon_path):
        app.setWindowIcon(QIcon(icon_path))
    else:
        print(f"DEBUG: Icon not found at {icon_path}")

    # PATCH: Force all QMainWindow instances to also get this icon
    _original_init = QMainWindow.__init__

    def _patched_init(self, *args, **kwargs):
        _original_init(self, *args, **kwargs)
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

    QMainWindow.__init__ = _patched_init

    # Create splash screen
    png_path = os.path.join(BASE_DIR, "daacal.png")

    if os.path.exists(png_path):
        splash_pix = QPixmap(png_path).scaled(
            220, 220,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation
        )
    else:
        splash_pix = QPixmap(220, 220)
        splash_pix.fill(Qt.white)

    splash = StartupSplash(splash_pix)
    splash.show()
    splash.raise_()
    splash.activateWindow()
    app.processEvents()

    # Create and show main window
    mainWin = WebsterCalendarApp(splash=splash)
    mainWin.showMaximized()

    QTimer.singleShot(3000, lambda: start_silent_update(mainWin))

    app.exec_()







